import io
import os
from datetime import date, datetime
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st
from supabase import Client, create_client

# =========================================================
# 1. CONFIGURATION DE LA PAGE STREAMLIT
# =========================================================
st.set_page_config(
    page_title="Suivi & Contrôle Qualité Béton - LGV Casa Sud", 
    page_icon="🏗️",
    layout="wide"
)

# Liste prédéfinie des classes de béton
CLASSES_BETON_LISTE = [
    "C20/25",
    "C25/30",
    "C30/37",
    "C35/45",
    "C40/50",
    "C45/55",
]

LISTE_METEO = [
    "Ensoleillé ☀️", 
    "Nuageux ☁️", 
    "Pluvieux 🌧️", 
    "Vent fort 💨", 
    "Chaud / Canicule 🌡️"
]

# --- FONCTIONS UTILITAIRES DE CONVERSION SÉCURISÉE ---
def safe_float(val, default=0.0):
    try:
        if val is None or str(val).strip() == "":
            return default
        return float(val)
    except (ValueError, TypeError):
        return default

def safe_int(val, default=0):
    try:
        if val is None or str(val).strip() == "":
            return default
        return int(float(val))
    except (ValueError, TypeError):
        return default

# =========================================================
# FONCTION 1 : EXCEL RAPPORT JOURNALIER (FORMAT PORTRAIT)
# =========================================================
def generer_excel_lpee(df_jour, date_rapport="", est_historique=False):
    """Génère un fichier Excel (.xlsx) stylisé aux normes LPEE en format PORTRAIT."""
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_titre = Font(name="Calibri", size=12, bold=True, color="1F4E78")
    font_sub_ctr = Font(name="Calibri", size=9, bold=True, color="1F4E78")
    font_info = Font(name="Calibri", size=8, italic=True, color="333333")
    font_header = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=8, color="000000")
    font_sig_title = Font(name="Calibri", size=8, bold=True, color="1F4E78")

    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_mapping = {
        "N°": "N°",
        "num_bon_livraison": "N° BL",
        "ouvrage": "Ouvrage",
        "element_betonne": "Élément Bétonné",
        "volume_beton": "Quantité (m³)",
        "classe_beton": "Classe béton",
        "heure_fin_production_cab": "Fin Prod.",
        "heure_arrivee_chantier": "Arrivée Chantier",
        "tbf": "T° Béton (°C)",
        "ta": "T° Amb. (°C)",
        "affaissement": "Slump (cm)\nNF EN 12350-2",
        "meteo": "Météo",
        "prelevement": "Prélèvement",
    }

    if df_jour.empty:
        classes = ["Rapport Béton"]
    elif est_historique or "classe_beton" not in df_jour.columns:
        classes = ["Registre Général"]
    else:
        classes = df_jour["classe_beton"].dropna().unique().tolist()
        if not classes:
            classes = ["Général"]

    for cls in classes:
        sheet_name = str(cls).replace("/", "-").replace("\\", "-")[:30]
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True

        if est_historique or cls in ["Registre Général", "Rapport Béton"]:
            df_cls = df_jour.copy()
        else:
            df_cls = df_jour[df_jour["classe_beton"] == cls].copy()

        df_export = df_cls.copy()
        if "N°" in df_export.columns:
            df_export = df_export.drop(columns=["N°"])
        df_export.insert(0, "N°", range(1, len(df_export) + 1))

        for c_key in col_mapping.keys():
            if c_key not in df_export.columns:
                df_export[c_key] = ""

        cols_presentes = list(col_mapping.keys())
        nb_cols = len(cols_presentes)

        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.horizontalCentered = True
        ws.page_margins.left = 0.2
        ws.page_margins.right = 0.2
        ws.page_margins.top = 0.3
        ws.page_margins.bottom = 0.3
        ws.print_title_rows = "5:5"

        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 16
        ws.row_dimensions[3].height = 16

        if os.path.exists("logo.png"):
            try:
                img = Image("logo.png")
                img.width = 55
                img.height = 50
                ws.add_image(img, "A1")
            except Exception:
                pass

        start_text_col = 2 if nb_cols > 2 else 1

        ws.merge_cells(start_row=1, start_column=start_text_col, end_row=1, end_column=nb_cols)
        c1 = ws.cell(row=1, column=start_text_col, value="LPEE - LABORATOIRE PUBLIC D'ESSAIS ET D'ÉTUDES")
        c1.font = font_titre
        c1.alignment = align_center

        ws.merge_cells(start_row=2, start_column=start_text_col, end_row=2, end_column=nb_cols)
        c2 = ws.cell(
            row=2, 
            column=start_text_col, 
            value="CENTRE TECHNIQUE RÉGIONAL DE CASABLANCA-SETTAT-BÉNI MELLAL (CTR-CSB)"
        )
        c2.font = font_sub_ctr
        c2.alignment = align_center

        ws.merge_cells(start_row=3, start_column=start_text_col, end_row=3, end_column=nb_cols)
        if est_historique:
            txt_info = "Projet : LGV CASA SUD | Client : TGCC | Registre Général"
        else:
            txt_info = f"Projet : LGV CASA SUD | Client : TGCC | Classe : {cls} | Date : {date_rapport}"
        c3 = ws.cell(row=3, column=start_text_col, value=txt_info)
        c3.font = font_info
        c3.alignment = align_center

        start_row = 5
        ws.row_dimensions[start_row].height = 28
        for col_idx, col_key in enumerate(cols_presentes, start=1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = col_mapping[col_key]
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_header
            cell.border = thin_border

        current_row = start_row + 1
        for row_data in df_export[cols_presentes].values:
            ws.row_dimensions[current_row].height = 17
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = "" if pd.isna(val) else val
                cell.font = font_data
                cell.border = thin_border
                cell.alignment = align_center
            current_row += 1

        # Signatures
        sig_row = current_row + 2
        ws.row_dimensions[sig_row].height = 16
        mid_col = nb_cols // 2

        ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=mid_col)
        c_resp = ws.cell(row=sig_row, column=1, value="Responsable d'essai")
        c_resp.font = font_sig_title
        c_resp.alignment = align_center

        ws.merge_cells(start_row=sig_row, start_column=mid_col + 1, end_row=sig_row, end_column=nb_cols)
        c_chef = ws.cell(row=sig_row, column=mid_col + 1, value="Chef du laboratoire")
        c_chef.font = font_sig_title
        c_chef.alignment = align_center

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row < start_row:
                    continue
                if cell.value is not None:
                    lines = str(cell.value).split("\n")
                    for l in lines:
                        max_len = max(max_len, len(l))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 9)

    wb.save(output)
    return output.getvalue()

# =========================================================
# CALCUL STATUT BÉTON
# =========================================================
def evaluer_statut_beton(tbf, h_fin, h_arr, affaissement):
    try:
        cond_tbf = float(tbf) < 32.1
        cond_aff = 10 <= float(affaissement) <= 22  # En cm
        fmt = "%H:%M"
        t_fin = datetime.strptime(str(h_fin).strip(), fmt)
        t_arr = datetime.strptime(str(h_arr).strip(), fmt)
        diff_minutes = (t_arr - t_fin).total_seconds() / 60
        if diff_minutes < 0:
            diff_minutes += 24 * 60
        cond_delai = 0 <= diff_minutes <= 120
        return "✅ Conforme" if (cond_tbf and cond_aff and cond_delai) else "⚠️ Non Conforme"
    except Exception:
        return "✅ Conforme"

# =========================================================
# 2. AUTHENTIFICATION & CONFIGURATION
# =========================================================
DEFAULT_PROJET = "LGV Casa Sud"
DEFAULT_ENTREPRISE = "TGCC"
DEFAULT_CENTRALE = "TG PREFA"

PASSWORD_GENERAL = "lpee2026"
PASSWORD_ADMIN = "lpee_admin_2026"

if "authentifie" not in st.session_state:
    st.session_state["authentifie"] = False
if "is_admin" not in st.session_state:
    st.session_state["is_admin"] = False

if not st.session_state["authentifie"]:
    st.title("🔒 Accès Sécurisé - LPEE Smart Control Béton")
    pwd = st.text_input("Mot de passe", type="password")
    if st.button("Valider l'accès"):
        if pwd in [PASSWORD_GENERAL, PASSWORD_ADMIN]:
            st.session_state["authentifie"] = True
            if pwd == PASSWORD_ADMIN:
                st.session_state["is_admin"] = True
            st.rerun()
        else:
            st.error("❌ Mot de passe incorrect.")
    st.stop()

# =========================================================
# 3. SUPABASE
# =========================================================
URL = "https://yqijsvxyrdymcnqluipa.supabase.co"
CLE = "sb_publishable_m8g5mocsCDgk3JpS1lpuCQ_3wOPyet1"

try:
    supabase: Client = create_client(URL, CLE)
    response = supabase.table("controles_beton").select("*").execute()
    data_all = response.data or []
except Exception as e:
    st.error(f"Erreur Supabase : {e}")
    data_all = []

st.title("🏗️ Suivi et Contrôle Qualité Béton")
st.info("Module de saisie et de suivi des bons de livraison du béton — **LGV Casa Sud (TGCC / LPEE)**")

# =========================================================
# 4. SÉLECTION JOURNÉE ET EXPORT
# =========================================================
st.markdown("---")
col_date1, col_date2, col_date3 = st.columns([1.2, 1.5, 1.3])

with col_date1:
    date_choisie = st.date_input("📅 Choisir la journée de bétonnage :", value=date.today())
    str_date_choisie = date_choisie.strftime("%d/%m/%Y")

data_jour = [r for r in data_all if r.get("date_betonnage") == str_date_choisie]

with col_date2:
    st.info(f"📌 **Journée du {str_date_choisie}**\nTotal camions contrôlés : **{len(data_jour)}**")

with col_date3:
    if data_jour:
        df_jour_export = pd.DataFrame(data_jour)
        excel_jour = generer_excel_lpee(df_jour_export, date_rapport=str_date_choisie)
        st.download_button(
            "📄 Télécharger Rapport Journalier (.xlsx)",
            data=excel_jour,
            file_name=f"Rapport_Beton_{str_date_choisie.replace('/', '_')}.xlsx",
            type="primary",
            use_container_width=True
        )

# =========================================================
# 5. ONGLETS PRINCIPAUX
# =========================================================
tab_ajouter, tab_modifier, tab_supprimer, tab_historique = st.tabs([
    "➕ Saisie d'un contrôle de bétonnage",
    "✏️ Modifier (Admin)",
    "❌ Supprimer (Admin)",
    "📚 Historique Général",
])

# --- ONGLET 1 : SAISIE ---
with tab_ajouter:
    with st.form("form_controle_beton"):
        st.subheader("Saisie d'un contrôle de bétonnage")
        
        col1, col2 = st.columns(2)

        with col1:
            date_livraison = st.date_input("Date de livraison", value=date_choisie, key="add_date")
            num_bl = st.text_input("N° Bon de Livraison (BL)", value="BL-2026-001", key="add_bl")
            classe_beton = st.selectbox("Classe béton", CLASSES_BETON_LISTE, index=0, key="add_classe")
            
            # ➕ NOUVEAU CHAMP : Quantité de béton (m³)
            quantite_beton = st.number_input(
                "Quantité de béton (m³)", 
                min_value=0.0, 
                value=8.00, 
                step=0.5, 
                format="%.2f",
                key="add_vol"
            )
            element_ouvrage = st.text_input("Élément d'ouvrage / Emplacement", value="Voile / Semelle", key="add_elem")

        with col2:
            temperature = st.number_input("Température (°C)", value=20.00, step=0.5, format="%.2f", key="add_tbf")
            affaissement = st.number_input("Affaissement / Slump (cm)", value=15.00, step=0.5, format="%.2f", key="add_aff")
            prelevement = st.selectbox(
                "Prélèvement : NF EN 12390-2",
                ["OUI - Conforme (NF EN 12390-2)", "NON - Sans prélèvement"],
                key="add_prelev"
            )
            nb_eprouvettes = st.number_input("Nombre d'éprouvettes", min_value=0, value=6, step=1, key="add_eprouvettes")
            
            # ➕ NOUVEAU CHAMP : État de la météo
            etat_meteo = st.selectbox("État de la météo", LISTE_METEO, index=0, key="add_meteo")

        st.markdown("<br>", unsafe_allow_html=True)
        submit_button = st.form_submit_button("Enregistrer le contrôle béton", type="primary")

        if submit_button:
            str_date = date_livraison.strftime("%d/%m/%Y")
            statut_auto = evaluer_statut_beton(temperature, "08:00", "08:30", affaissement)
            
            data_to_insert = {
                "projet": DEFAULT_PROJET,
                "entreprise": DEFAULT_ENTREPRISE,
                "centrale_beton": DEFAULT_CENTRALE,
                "ouvrage": "LGV CASA SUD",
                "element_betonne": element_ouvrage,
                "volume_beton": quantite_beton,
                "num_bon_livraison": num_bl,
                "classe_beton": classe_beton,
                "date_betonnage": str_date,
                "meteo": etat_meteo,
                "observations": f"Éprouvettes: {nb_eprouvettes}",
                "heure_fin_production_cab": "08:00",
                "heure_arrivee_chantier": "08:30",
                "tbf": temperature,
                "ta": 25.0,
                "affaissement": affaissement,
                "prelevement": prelevement,
                "technicien": "LPEE",
                "statut": statut_auto,
            }
            try:
                supabase.table("controles_beton").insert(data_to_insert).execute()
                st.success(f"✅ Contrôle enregistré avec succès ! BL: {num_bl} | Volume: {quantite_beton} m³ | Météo: {etat_meteo}")
                st.rerun()
            except Exception as e:
                st.error(f"Erreur d'enregistrement Supabase : {e}")

# --- ONGLET 2 : MODIFIER ---
with tab_modifier:
    if not st.session_state["is_admin"]:
        pwd_admin = st.text_input("Code Administrateur :", type="password", key="mod_pwd")
        if st.button("Débloquer pour modification"):
            if pwd_admin == PASSWORD_ADMIN:
                st.session_state["is_admin"] = True
                st.rerun()
            else:
                st.error("❌ Code incorrect")
    else:
        if data_jour:
            options_edit = {
                f"N° {r.get('id')} | BL: {r.get('num_bon_livraison')} | Volume: {r.get('volume_beton')} m³ | Météo: {r.get('meteo')}": r
                for r in data_jour
            }
            choix = st.selectbox("Sélectionner le camion à modifier :", list(options_edit.keys()))
            row_s = options_edit[choix]
            
            with st.form("form_edit"):
                c1, c2 = st.columns(2)
                with c1:
                    e_bl = st.text_input("N° BL", value=str(row_s.get("num_bon_livraison") or ""))
                    curr_classe = str(row_s.get("classe_beton") or "C20/25")
                    e_classe = st.selectbox("Classe béton", CLASSES_BETON_LISTE, index=CLASSES_BETON_LISTE.index(curr_classe) if curr_classe in CLASSES_BETON_LISTE else 0)
                    e_vol = st.number_input("Quantité de béton (m³)", value=safe_float(row_s.get("volume_beton"), 8.00), step=0.5)
                    e_elem = st.text_input("Élément d'ouvrage", value=str(row_s.get("element_betonne") or ""))

                with c2:
                    e_tbf = st.number_input("Température (°C)", value=safe_float(row_s.get("tbf"), 20.00), step=0.5)
                    e_aff = st.number_input("Affaissement (cm)", value=safe_float(row_s.get("affaissement"), 15.00), step=0.5)
                    curr_meteo = str(row_s.get("meteo") or LISTE_METEO[0])
                    e_meteo = st.selectbox("État de la météo", LISTE_METEO, index=LISTE_METEO.index(curr_meteo) if curr_meteo in LISTE_METEO else 0)

                if st.form_submit_button("🔄 Mettre à jour"):
                    up_data = {
                        "num_bon_livraison": e_bl,
                        "classe_beton": e_classe,
                        "volume_beton": e_vol,
                        "element_betonne": e_elem,
                        "tbf": e_tbf,
                        "affaissement": e_aff,
                        "meteo": e_meteo,
                    }
                    try:
                        supabase.table("controles_beton").update(up_data).eq("id", row_s["id"]).execute()
                        st.success("Données mises à jour avec succès !")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erreur de mise à jour : {e}")
        else:
            st.info(f"Aucun camion enregistré pour le {str_date_choisie}.")

# --- ONGLET 3 : SUPPRIMER ---
with tab_supprimer:
    if not st.session_state["is_admin"]:
        pwd_admin_del = st.text_input("Code Administrateur :", type="password", key="del_pwd")
        if st.button("Débloquer pour suppression"):
            if pwd_admin_del == PASSWORD_ADMIN:
                st.session_state["is_admin"] = True
                st.rerun()
            else:
                st.error("❌ Code incorrect")
    else:
        if data_jour:
            options_del = {
                f"N° {r.get('id')} | BL: {r.get('num_bon_livraison')} | Volume: {r.get('volume_beton')} m³": r["id"]
                for r in data_jour
            }
            c_del = st.selectbox("Sélectionner la ligne à supprimer :", list(options_del.keys()))
            if st.button("🚨 Supprimer définitivement", type="primary"):
                try:
                    supabase.table("controles_beton").delete().eq("id", options_del[c_del]).execute()
                    st.success("Ligne supprimée avec succès !")
                    st.rerun()
                except Exception as e:
                    st.error(f"Erreur lors de la suppression : {e}")
        else:
            st.info(f"Aucun camion enregistré pour le {str_date_choisie}.")

# --- ONGLET 4 : HISTORIQUE GENERAL ---
with tab_historique:
    st.subheader("📚 Registre Général des Contrôles")
    if data_all:
        df_all = pd.DataFrame(data_all)
        cols_h = [
            "num_bon_livraison",
            "element_betonne",
            "volume_beton",
            "classe_beton",
            "date_betonnage",
            "tbf",
            "affaissement",
            "meteo",
            "statut"
        ]
        df_h_v = df_all[[c for c in cols_h if c in df_all.columns]]
        st.dataframe(df_h_v, use_container_width=True)
        
        excel_hist = generer_excel_lpee(df_all, est_historique=True)
        st.download_button(
            "📚 Télécharger tout le Registre Général (.xlsx)",
            data=excel_hist,
            file_name="Registre_General_Beton_LPEE.xlsx"
        )
    else:
        st.info("Aucune donnée enregistrée dans le registre.")
