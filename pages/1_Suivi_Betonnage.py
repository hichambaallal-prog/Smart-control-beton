import io
import os
from datetime import date, datetime
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st
from supabase import Client, create_client

# 1. Configuration de la page Streamlit
st.set_page_config(
    page_title="Suivi Béton - LGV Casa Sud (LPEE)", layout="wide"
)

# Liste prédéfinie des classes de béton
CLASSES_BETON_LISTE = [
    "C20/25",
    "C25/30",
    "C30/37",
    "C35/45",
    "C 40/50",
    "C45/55",
]


# --- FONCTIONS UTILITAIRES DE CONVERSION SÉCURISÉE ---
def safe_float(val, default=0.0):
  try:
    if val is None or str(val).strip() == "":
      return default
    return float(val)
  except (ValueError, TypeError):
    return default


def safe_int(val, default=0):
  try:
    if val is None or str(val).strip() == "":
      return default
    return int(float(val))
  except (ValueError, TypeError):
    return default


# =========================================================
# FONCTION 1 : EXCEL RAPPORT JOURNALIER (FORMAT PORTRAIT / SANS TECHNICIEN ET STATUT)
# =========================================================
def generer_excel_lpee(df_jour, date_rapport="", est_historique=False):
  """Génère un fichier Excel (.xlsx) stylisé aux normes LPEE en format PORTRAIT.

  Sans les colonnes Technicien et Statut. Contient la référence à la norme NF
  EN 12350-2 pour l'affaissement.
  """
  output = io.BytesIO()
  wb = openpyxl.Workbook()
  wb.remove(wb.active)  # Supprimer la feuille par défaut

  # Définition des Styles LPEE
  font_titre = Font(name="Calibri", size=12, bold=True, color="1F4E78")
  font_sub_ctr = Font(name="Calibri", size=9, bold=True, color="1F4E78")
  font_info = Font(name="Calibri", size=8, italic=True, color="333333")
  font_header = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
  font_data = Font(name="Calibri", size=8, color="000000")
  font_sig_title = Font(name="Calibri", size=8, bold=True, color="1F4E78")

  fill_header = PatternFill(
      start_color="1F4E78", end_color="1F4E78", fill_type="solid"
  )

  thin_border = Border(
      left=Side(style="thin", color="D9D9D9"),
      right=Side(style="thin", color="D9D9D9"),
      top=Side(style="thin", color="D9D9D9"),
      bottom=Side(style="thin", color="D9D9D9"),
  )

  align_center = Alignment(horizontal="center", vertical="center")
  align_header = Alignment(
      horizontal="center", vertical="center", wrap_text=True
  )

  col_mapping = {
      "N°": "N°",
      "num_bon_livraison": "N° BL",
      "ouvrage": "Ouvrage",
      "element_betonne": "Élément Bétonné",
      "volume_beton": "Vol. (m³)",
      "classe_beton": "Classe béton",
      "heure_fin_production_cab": "Fin Prod.",
      "heure_arrivee_chantier": "Arrivée Chantier",
      "tbf": "TBF (°C)",
      "ta": "TA (°C)",
      "affaissement": "Slump (mm)\nNF EN 12350-2",
      "prelevement": "Prélév.",
  }

  # Identification des classes de béton présentes
  if df_jour.empty:
    classes = ["Rapport Béton"]
  elif est_historique or "classe_beton" not in df_jour.columns:
    classes = ["Registre Général"]
  else:
    classes = df_jour["classe_beton"].dropna().unique().tolist()
    if not classes:
      classes = ["Général"]

  for cls in classes:
    sheet_name = str(cls).replace("/", "-").replace("\\", "-")[:30]
    ws = wb.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True

    if est_historique or cls in ["Registre Général", "Rapport Béton"]:
      df_cls = df_jour.copy()
    else:
      df_cls = df_jour[df_jour["classe_beton"] == cls].copy()

    df_export = df_cls.copy()
    if "N°" in df_export.columns:
      df_export = df_export.drop(columns=["N°"])
    df_export.insert(0, "N°", range(1, len(df_export) + 1))

    # Assurer la présence de toutes les colonnes du mapping
    for c_key in col_mapping.keys():
      if c_key not in df_export.columns:
        df_export[c_key] = ""

    cols_presentes = list(col_mapping.keys())
    nb_cols = len(cols_presentes)

    # --- CONFIGURATION IMPRESSION EXCEL (A4 PORTRAIT) ---
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.horizontalCentered = True
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.print_title_rows = "5:5"

    # --- 1. EN-TÊTE ET LOGO LPEE ---
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 16

    if os.path.exists("logo.png"):
      try:
        img = Image("logo.png")
        img.width = 55
        img.height = 50
        ws.add_image(img, "A1")
      except Exception:
        pass

    start_text_col = 2 if nb_cols > 2 else 1

    ws.merge_cells(
        start_row=1,
        start_column=start_text_col,
        end_row=1,
        end_column=nb_cols,
    )
    c1 = ws.cell(
        row=1,
        column=start_text_col,
        value="LPEE - LABORATOIRE PUBLIC D'ESSAIS ET D'ÉTUDES",
    )
    c1.font = font_titre
    c1.alignment = align_center

    ws.merge_cells(
        start_row=2,
        start_column=start_text_col,
        end_row=2,
        end_column=nb_cols,
    )
    c2 = ws.cell(
        row=2,
        column=start_text_col,
        value=(
            "CENTRE TECHNIQUE RÉGIONAL DE CASABLANCA-SETTAT-BÉNI MELLAL"
            " (CTR-CSB)"
        ),
    )
    c2.font = font_sub_ctr
    c2.alignment = align_center

    ws.merge_cells(
        start_row=3,
        start_column=start_text_col,
        end_row=3,
        end_column=nb_cols,
    )
    if est_historique:
      txt_info = "Projet : LGV CASA SUD | Client : TGCC | Registre Général"
    else:
      txt_info = (
          f"Projet : LGV CASA SUD | Client : TGCC | Classe : {cls} | Date :"
          f" {date_rapport}"
      )
    c3 = ws.cell(row=3, column=start_text_col, value=txt_info)
    c3.font = font_info
    c3.alignment = align_center

    # --- 2. TABLEAU DE DONNÉES ---
    start_row = 5
    ws.row_dimensions[start_row].height = 28
    for col_idx, col_key in enumerate(cols_presentes, start=1):
      cell = ws.cell(row=start_row, column=col_idx)
      cell.value = col_mapping[col_key]
      cell.font = font_header
      cell.fill = fill_header
      cell.alignment = align_header
      cell.border = thin_border

    current_row = start_row + 1
    for row_data in df_export[cols_presentes].values:
      ws.row_dimensions[current_row].height = 17
      for col_idx, val in enumerate(row_data, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = "" if pd.isna(val) else val
        cell.font = font_data
        cell.border = thin_border
        cell.alignment = align_center
      current_row += 1

    # --- 3. PIED DE PAGE : SIGNATURES ---
    sig_row = current_row + 2
    ws.row_dimensions[sig_row].height = 16

    mid_col = nb_cols // 2

    ws.merge_cells(
        start_row=sig_row, start_column=1, end_row=sig_row, end_column=mid_col
    )
    c_resp = ws.cell(row=sig_row, column=1, value="Responsable d'essai")
    c_resp.font = font_sig_title
    c_resp.alignment = align_center

    for r in range(sig_row, sig_row + 4):
      for c in range(1, mid_col + 1):
        cell = ws.cell(row=r, column=c)
        top = Side(style="thin", color="1F4E78") if r == sig_row else None
        bottom = Side(style="thin", color="1F4E78") if r == sig_row + 3 else None
        left = Side(style="thin", color="1F4E78") if c == 1 else None
        right = Side(style="thin", color="1F4E78") if c == mid_col else None
        cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    ws.merge_cells(
        start_row=sig_row,
        start_column=mid_col + 1,
        end_row=sig_row,
        end_column=nb_cols,
    )
    c_chef = ws.cell(row=sig_row, column=mid_col + 1, value="Chef du laboratoire")
    c_chef.font = font_sig_title
    c_chef.alignment = align_center

    for r in range(sig_row, sig_row + 4):
      for c in range(mid_col + 1, nb_cols + 1):
        cell = ws.cell(row=r, column=c)
        top = Side(style="thin", color="1F4E78") if r == sig_row else None
        bottom = Side(style="thin", color="1F4E78") if r == sig_row + 3 else None
        left = Side(style="thin", color="1F4E78") if c == mid_col + 1 else None
        right = Side(style="thin", color="1F4E78") if c == nb_cols else None
        cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    # Ajustement des largeurs de colonnes
    for col in ws.columns:
      max_len = 0
      col_letter = get_column_letter(col[0].column)
      for cell in col:
        if cell.row < start_row:
          continue
        if cell.value is not None:
          lines = str(cell.value).split("\n")
          for l in lines:
            max_len = max(max_len, len(l))
      ws.column_dimensions[col_letter].width = max(max_len + 3, 9)

  wb.save(output)
  return output.getvalue()


# =========================================================
# FONCTION 2 : EXCEL RÉCAPITULATIF MENSUEL
# =========================================================
def generer_excel_recap_mensuel_lpee(df_recap, mois):
  output = io.BytesIO()
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = f"Synthèse {str(mois).replace('/', '-')}"
  ws.views.sheetView[0].showGridLines = True

  nb_cols = max(len(df_recap.columns), 7)

  ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
  ws.page_setup.paperSize = ws.PAPERSIZE_A4
  ws.sheet_properties.pageSetUpPr.fitToPage = True
  ws.page_setup.fitToWidth = 1
  ws.page_setup.fitToHeight = 0
  ws.page_setup.horizontalCentered = True
  ws.page_margins.left = 0.3
  ws.page_margins.right = 0.3

  font_titre = Font(name="Calibri", size=13, bold=True, color="1F4E78")
  font_sub_ctr = Font(name="Calibri", size=10, bold=True, color="1F4E78")
  font_card_val = Font(name="Calibri", size=10, bold=True, color="1F4E78")
  font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
  font_total = Font(name="Calibri", size=10, bold=True, color="1F4E78")

  fill_header = PatternFill(
      start_color="1F4E78", end_color="1F4E78", fill_type="solid"
  )
  fill_card = PatternFill(
      start_color="F2F5F9", end_color="F2F5F9", fill_type="solid"
  )
  fill_total = PatternFill(
      start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
  )

  thin_border = Border(
      left=Side(style="thin", color="D9D9D9"),
      right=Side(style="thin", color="D9D9D9"),
      top=Side(style="thin", color="D9D9D9"),
      bottom=Side(style="thin", color="D9D9D9"),
  )
  total_border = Border(
      left=Side(style="thin", color="D9D9D9"),
      right=Side(style="thin", color="D9D9D9"),
      top=Side(style="medium", color="1F4E78"),
      bottom=Side(style="double", color="1F4E78"),
  )

  align_center = Alignment(horizontal="center", vertical="center")
  align_left = Alignment(horizontal="left", vertical="center")

  if os.path.exists("logo.png"):
    try:
      img = Image("logo.png")
      img.width = 65
      img.height = 60
      ws.add_image(img, "A1")
    except Exception:
      pass

  start_txt_col = 2 if nb_cols > 2 else 1

  ws.merge_cells(
      start_row=1, start_column=start_txt_col, end_row=1, end_column=nb_cols
  )
  c1 = ws.cell(
      row=1,
      column=start_txt_col,
      value="LPEE - LABORATOIRE PUBLIC D'ESSAIS ET D'ÉTUDES",
  )
  c1.font = font_titre
  c1.alignment = align_center

  ws.merge_cells(
      start_row=2, start_column=start_txt_col, end_row=2, end_column=nb_cols
  )
  c2 = ws.cell(
      row=2,
      column=start_txt_col,
      value=(
          "CENTRE TECHNIQUE RÉGIONAL DE CASABLANCA-SETTAT-BÉNI MELLAL (CTR-CSB)"
      ),
  )
  c2.font = font_sub_ctr
  c2.alignment = align_center

  mid_col = nb_cols // 2
  ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=mid_col)
  ws.cell(row=4, column=1, value="PROJET : LGV CASA SUD").font = font_card_val

  ws.merge_cells(
      start_row=4, start_column=mid_col + 1, end_row=4, end_column=nb_cols
  )
  ws.cell(
      row=4, column=mid_col + 1, value="CLIENT / ENTREPRISE : TGCC"
  ).font = font_card_val

  ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=mid_col)
  ws.cell(
      row=5, column=1, value="ORGANISME : LPEE - Laboratoire LGV Casa Sud"
  ).font = font_card_val

  ws.merge_cells(
      start_row=5, start_column=mid_col + 1, end_row=5, end_column=nb_cols
  )
  ws.cell(
      row=5, column=mid_col + 1, value=f"PÉRIODE DE SYNTHÈSE : {mois}"
  ).font = font_card_val

  for r in range(4, 6):
    for c in range(1, nb_cols + 1):
      cell = ws.cell(row=r, column=c)
      cell.fill = fill_card
      cell.border = thin_border

  headers = list(df_recap.columns)
  start_row = 7
  ws.row_dimensions[start_row].height = 24

  for col_idx, h_name in enumerate(headers, start=1):
    cell = ws.cell(row=start_row, column=col_idx)
    cell.value = h_name
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

  current_row = start_row + 1
  for row_data in df_recap.values:
    ws.row_dimensions[current_row].height = 19
    for col_idx, val in enumerate(row_data, start=1):
      cell = ws.cell(row=current_row, column=col_idx)
      cell.value = "" if pd.isna(val) else val
      cell.border = thin_border
      cell.alignment = (
          align_center if col_idx in [1, 3, 4, 5, 6, 7] else align_left
      )
    current_row += 1

  if len(df_recap) > 0:
    ws.row_dimensions[current_row].height = 22
    c_tot = ws.cell(row=current_row, column=1, value="TOTAL / SYNTHÈSE MOIS")
    c_tot.font = font_total
    c_tot.alignment = align_center
    c_tot.fill = fill_total
    c_tot.border = total_border

    c_empty = ws.cell(row=current_row, column=2, value="—")
    c_empty.font = font_total
    c_empty.alignment = align_center
    c_empty.fill = fill_total
    c_empty.border = total_border

    tot_ctrl = int(df_recap["Nombre de contrôles"].sum())
    c_ctrl = ws.cell(row=current_row, column=3, value=tot_ctrl)
    c_ctrl.font = font_total
    c_ctrl.alignment = align_center
    c_ctrl.fill = fill_total
    c_ctrl.border = total_border

    for c_idx, val in enumerate(
        [
            df_recap["Affaissement Min (mm)"].min(),
            df_recap["Affaissement Max (mm)"].max(),
            df_recap["TBF Min (°C)"].min(),
            df_recap["TBF Max (°C)"].max(),
        ],
        start=4,
    ):
      c_cell = ws.cell(
          row=current_row, column=c_idx, value="" if pd.isna(val) else val
      )
      c_cell.font = font_total
      c_cell.alignment = align_center
      c_cell.fill = fill_total
      c_cell.border = total_border

  for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
      if cell.row < start_row:
        continue
      if cell.value is not None:
        lines = str(cell.value).split("\n")
        for l in lines:
          max_len = max(max_len, len(l))
    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

  wb.save(output)
  return output.getvalue()


# =========================================================
# CALCUL STATUT BÉTON
# =========================================================
def evaluer_statut_beton(tbf, h_fin, h_arr, affaissement):
  try:
    cond_tbf = float(tbf) < 32.1
    cond_aff = 160 <= int(float(affaissement)) <= 220
    fmt = "%H:%M"
    t_fin = datetime.strptime(str(h_fin).strip(), fmt)
    t_arr = datetime.strptime(str(h_arr).strip(), fmt)
    diff_minutes = (t_arr - t_fin).total_seconds() / 60
    if diff_minutes < 0:
      diff_minutes += 24 * 60
    cond_delai = 0 <= diff_minutes <= 120
    return (
        "✅ Conforme"
        if (cond_tbf and cond_aff and cond_delai)
        else "⚠️ Non Conforme"
    )
  except Exception:
    return "⚠️ Non Conforme"


# =========================================================
# 2. VALEURS FIXES & AUTHENTIFICATION
# =========================================================
DEFAULT_PROJET = "LGV Casa Sud"
DEFAULT_ENTREPRISE = "TGCC"
DEFAULT_CENTRALE = "TG PREFA"

PASSWORD_GENERAL = "lpee2026"
PASSWORD_ADMIN = "lpee_admin_2026"

if "authentifie" not in st.session_state:
  st.session_state["authentifie"] = False
if "is_admin" not in st.session_state:
  st.session_state["is_admin"] = False

if not st.session_state["authentifie"]:
  st.title("🔒 Accès Sécurisé - LPEE Smart Control Béton")
  pwd = st.text_input("Mot de passe", type="password")
  if st.button("Valider l'accès"):
    if pwd in [PASSWORD_GENERAL, PASSWORD_ADMIN]:
      st.session_state["authentifie"] = True
      if pwd == PASSWORD_ADMIN:
        st.session_state["is_admin"] = True
      st.rerun()
    else:
      st.error("❌ Mot de passe incorrect.")
  st.stop()

# =========================================================
# 3. SUPABASE
# =========================================================
URL = "https://yqijsvxyrdymcnqluipa.supabase.co"
CLE = "sb_publishable_m8g5mocsCDgk3JpS1lpuCQ_3wOPyet1"

try:
  supabase: Client = create_client(URL, CLE)
  response = supabase.table("controles_beton").select("*").execute()
  data_all = response.data or []
except Exception as e:
  st.error(f"Erreur Supabase : {e}")
  data_all = []

st.title("🚧 Suivi Journalier de Bétonnage - LGV Casa Sud")
st.markdown("##### LPEE - CTR-CSB | Projet : **LGV CASA SUD** | Client : **TGCC**")

# =========================================================
# 4. SÉLECTION JOURNÉE
# =========================================================
st.markdown("---")
col_date1, col_date2 = st.columns([1, 2])
with col_date1:
  date_choisie = st.date_input(
      "📅 Choisir la journée de bétonnage :", value=date.today()
  )
  str_date_choisie = date_choisie.strftime("%d/%m/%Y")

data_jour = [r for r in data_all if r.get("date_betonnage") == str_date_choisie]

with col_date2:
  st.info(
      f"📌 **Journée du {str_date_choisie}** | Total camions contrôlés :"
      f" **{len(data_jour)}**"
  )

# =========================================================
# 5. ONGLETS PRINCIPAUX
# =========================================================
tab_ajouter, tab_modifier, tab_supprimer, tab_historique, tab_recap_mensuel = (
    st.tabs([
        "➕ Ajouter un camion",
        "✏️ Modifier (Admin)",
        "❌ Supprimer (Admin)",
        "📚 Historique Général",
        "📊 Récap Mensuel",
    ])
)

# --- ONGLET 1 : SAISIE ---
with tab_ajouter:
  st.subheader("📝 Saisie de suivi de bétonnage")
  with st.form("form_controle_ajouter"):
    c1, c2, c3 = st.columns(3)
    with c1:
      date_betonnage_saisie = st.date_input(
          "📅 Date de bétonnage :", value=date_choisie, key="add_date_saisie"
      )
      projet = st.text_input("Projet", value=DEFAULT_PROJET, disabled=True)
      entreprise = st.text_input(
          "Entreprise / Client", value=DEFAULT_ENTREPRISE, disabled=True
      )
      centrale_beton = st.text_input(
          "Centrale béton", value=DEFAULT_CENTRALE, disabled=True
      )
    with c2:
      ouvrage = st.text_input("Ouvrage", value="PRO745 OA1", key="add_ouvrage")
      element_betonne = st.text_input(
          "Élément bétonné", value="Semelle C0", key="add_elem"
      )
      volume_beton = st.text_input("Volume béton (m³)", value="8", key="add_vol")
    with c3:
      num_bon_livraison = st.text_input(
          "N° bon livraison (BL)", value="BL2548", key="add_bl"
      )
      # MODIFICATION ICI : Sélecteur déroulant pour la classe de béton
      classe_beton = st.selectbox(
          "Classe béton",
          options=CLASSES_BETON_LISTE,
          index=2,  # C30/37 sélectionné par défaut
          key="add_classe",
      )

    st.markdown("---")
    col_m1, col_m2, col_m3, col_m4 = st.columns(4)
    with col_m1:
      heure_fin_prod = st.text_input(
          "Heure fin prod. CAB (HH:MM)", value="15:28", key="add_h_fin"
      )
      heure_arrivee = st.text_input(
          "Heure arrivée chantier (HH:MM)", value="16:31", key="add_h_arr"
      )
    with col_m2:
      tbf = st.number_input(
          "TBF (°C)", value=32.0, step=0.1, format="%.1f", key="add_tbf"
      )
      ta = st.number_input(
          "TA (°C)", value=28.9, step=0.1, format="%.1f", key="add_ta"
      )
    with col_m3:
      affaissement = st.number_input(
          "Affaissement (mm) [NF EN 12350-2]",
          value=170,
          step=10,
          format="%d",
          key="add_aff",
      )
      meteo = st.selectbox(
          "Météo", ["Soleil", "Nuageux", "Pluie", "Vent"], key="add_meteo"
      )
    with col_m4:
      prelevement = st.selectbox(
          "Prélèvement", ["OUI", "NON"], key="add_prelev"
      )
      technicien = st.text_input(
          "Technicien Contrôleur", value="Ismail / Mohamed", key="add_tech"
      )

    observations = st.text_area("Observations", value="RAS", key="add_obs")
    submit_add = st.form_submit_button("💾 Enregistrer le camion")

    if submit_add:
      statut_auto = evaluer_statut_beton(
          tbf, heure_fin_prod, heure_arrivee, affaissement
      )
      str_date_saisie = date_betonnage_saisie.strftime("%d/%m/%Y")
      data_to_insert = {
          "projet": DEFAULT_PROJET,
          "entreprise": DEFAULT_ENTREPRISE,
          "centrale_beton": DEFAULT_CENTRALE,
          "ouvrage": ouvrage,
          "element_betonne": element_betonne,
          "volume_beton": volume_beton,
          "num_bon_livraison": num_bon_livraison,
          "classe_beton": classe_beton,
          "date_betonnage": str_date_saisie,
          "meteo": meteo,
          "observations": observations,
          "heure_fin_production_cab": heure_fin_prod,
          "heure_arrivee_chantier": heure_arrivee,
          "tbf": round(tbf, 1),
          "ta": round(ta, 1),
          "affaissement": int(affaissement),
          "prelevement": prelevement,
          "technicien": technicien,
          "statut": statut_auto,
      }
      try:
        supabase.table("controles_beton").insert(data_to_insert).execute()
        st.success(
            f"Camion BL : {num_bon_livraison} (Classe béton : {classe_beton})"
            f" enregistré pour le {str_date_saisie} !"
        )
        st.rerun()
      except Exception as e:
        st.error(f"Erreur d'enregistrement : {e}")

# --- ONGLET 2 : MODIFIER ---
with tab_modifier:
  if not st.session_state["is_admin"]:
    pwd_admin = st.text_input("Code Administrateur :", type="password", key="mod_pwd")
    if st.button("Débloquer pour modification"):
      if pwd_admin == PASSWORD_ADMIN:
        st.session_state["is_admin"] = True
        st.rerun()
      else:
        st.error("❌ Code incorrect")
  else:
    if data_jour:
      options_edit = {
          (
              f"N° {r.get('id')} | BL: {r.get('num_bon_livraison')} | Classe:"
              f" {r.get('classe_beton')} | Ouvrage: {r.get('ouvrage')}"
          ): r
          for r in data_jour
      }
      choix = st.selectbox(
          "Sélectionner le camion à modifier :", list(options_edit.keys())
      )
      row_s = options_edit[choix]
      with st.form("form_edit"):
        c1, c2, c3 = st.columns(3)
        with c1:
          e_proj = st.text_input(
              "Projet", value=str(row_s.get("projet") or DEFAULT_PROJET)
          )
          e_ent = st.text_input(
              "Client", value=str(row_s.get("entreprise") or DEFAULT_ENTREPRISE)
          )
          e_cent = st.text_input(
              "Centrale",
              value=str(row_s.get("centrale_beton") or DEFAULT_CENTRALE),
          )
        with c2:
          e_ouv = st.text_input("Ouvrage", value=str(row_s.get("ouvrage") or ""))
          e_elem = st.text_input(
              "Élément", value=str(row_s.get("element_betonne") or "")
          )
          e_vol = st.text_input(
              "Volume", value=str(row_s.get("volume_beton") or "")
          )
        with c3:
          e_bl = st.text_input(
              "BL", value=str(row_s.get("num_bon_livraison") or "")
          )

          # MODIFICATION ICI : Pré-sélection sécurisée dans la liste déroulante lors de l'édition
          curr_classe = str(row_s.get("classe_beton") or "C30/37")
          idx_classe = (
              CLASSES_BETON_LISTE.index(curr_classe)
              if curr_classe in CLASSES_BETON_LISTE
              else 0
          )
          e_classe = st.selectbox(
              "Classe béton", options=CLASSES_BETON_LISTE, index=idx_classe
          )

        st.markdown("---")
        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        with col_m1:
          e_h_fin = st.text_input(
              "Fin Prod CAB",
              value=str(row_s.get("heure_fin_production_cab") or ""),
          )
          e_h_arr = st.text_input(
              "Arrivée Chantier",
              value=str(row_s.get("heure_arrivee_chantier") or ""),
          )
        with col_m2:
          e_tbf = st.number_input(
              "TBF", value=safe_float(row_s.get("tbf"), 32.0), step=0.1
          )
          e_ta = st.number_input(
              "TA", value=safe_float(row_s.get("ta"), 28.0), step=0.1
          )
        with col_m3:
          e_aff = st.number_input(
              "Affaissement (mm) [NF EN 12350-2]",
              value=safe_int(row_s.get("affaissement"), 170),
              step=10,
              format="%d",
          )

          meteo_list = ["Soleil", "Nuageux", "Pluie", "Vent"]
          curr_meteo = str(row_s.get("meteo") or "Soleil")
          meteo_idx = (
              meteo_list.index(curr_meteo) if curr_meteo in meteo_list else 0
          )
          e_meteo = st.selectbox("Météo", meteo_list, index=meteo_idx)
        with col_m4:
          prelev_list = ["OUI", "NON"]
          curr_prelev = str(row_s.get("prelevement") or "OUI").upper()
          prelev_idx = (
              prelev_list.index(curr_prelev)
              if curr_prelev in prelev_list
              else 0
          )
          e_prelev = st.selectbox(
              "Prélèvement", prelev_list, index=prelev_idx
          )

          e_tech = st.text_input(
              "Technicien", value=str(row_s.get("technicien") or "")
          )

        e_obs = st.text_area(
            "Observations", value=str(row_s.get("observations") or "")
        )
        if st.form_submit_button("🔄 Mettre à jour"):
          statut_up = evaluer_statut_beton(e_tbf, e_h_fin, e_h_arr, e_aff)
          up_data = {
              "projet": e_proj,
              "entreprise": e_ent,
              "centrale_beton": e_cent,
              "ouvrage": e_ouv,
              "element_betonne": e_elem,
              "volume_beton": e_vol,
              "num_bon_livraison": e_bl,
              "classe_beton": e_classe,
              "date_betonnage": row_s.get("date_betonnage") or str_date_choisie,
              "meteo": e_meteo,
              "observations": e_obs,
              "heure_fin_production_cab": e_h_fin,
              "heure_arrivee_chantier": e_h_arr,
              "tbf": round(e_tbf, 1),
              "ta": round(e_ta, 1),
              "affaissement": int(e_aff),
              "prelevement": e_prelev,
              "technicien": e_tech,
              "statut": statut_up,
          }
          try:
            supabase.table("controles_beton").update(up_data).eq(
                "id", row_s["id"]
            ).execute()
            st.success("Données mises à jour avec succès !")
            st.rerun()
          except Exception as e:
            st.error(f"Erreur de mise à jour : {e}")
    else:
      st.info(f"Aucun camion enregistré pour le {str_date_choisie}.")

# --- ONGLET 3 : SUPPRIMER ---
with tab_supprimer:
  if not st.session_state["is_admin"]:
    pwd_admin_del = st.text_input(
        "Code Administrateur :", type="password", key="del_pwd"
    )
    if st.button("Débloquer pour suppression"):
      if pwd_admin_del == PASSWORD_ADMIN:
        st.session_state["is_admin"] = True
        st.rerun()
      else:
        st.error("❌ Code incorrect")
  else:
    if data_jour:
      options_del = {
          (
              f"N° {r.get('id')} | BL: {r.get('num_bon_livraison')} | Classe:"
              f" {r.get('classe_beton')} | Ouvrage: {r.get('ouvrage')}"
          ): r["id"]
          for r in data_jour
      }
      c_del = st.selectbox(
          "Sélectionner la ligne à supprimer :", list(options_del.keys())
      )
      if st.button("🚨 Supprimer définitivement", type="primary"):
        try:
          supabase.table("controles_beton").delete().eq(
              "id", options_del[c_del]
          ).execute()
          st.success("Ligne supprimée avec succès !")
          st.rerun()
        except Exception as e:
          st.error(f"Erreur lors de la suppression : {e}")
    else:
      st.info(f"Aucun camion enregistré pour le {str_date_choisie}.")

# --- ONGLET 4 : HISTORIQUE GENERAL ---
with tab_historique:
  st.subheader("📚 Registre Général des Contrôles")
  if data_all:
    df_all = pd.DataFrame(data_all)
    cols_h = [
        "num_bon_livraison",
        "ouvrage",
        "element_betonne",
        "volume_beton",
        "classe_beton",
        "date_betonnage",
        "tbf",
        "affaissement",
    ]
    df_h_v = df_all[[c for c in cols_h if c in df_all.columns]]
    st.dataframe(df_h_v, use_container_width=True)
  else:
    st.info("Aucune donnée enregistrée dans le registre.")

# --- ONGLET 5 : RÉCAP MENSUEL ---
with tab_recap_mensuel:
  st.subheader("📊 Récapitulatif Mensuel LPEE - CTR-CSB")
  if data_all:
    df_r = pd.DataFrame(data_all)
    df_r["dt"] = pd.to_datetime(
        df_r["date_betonnage"], format="%d/%m/%Y", errors="coerce"
    )
    df_v = df_r.dropna(subset=["dt"]).copy()
    if not df_v.empty:
      df_v["mois_annee"] = df_v["dt"].dt.strftime("%m/%Y")
      mois_sel = st.selectbox(
          "📅 Choisir le mois :",
          sorted(df_v["mois_annee"].unique(), reverse=True),
      )
      df_m = df_v[df_v["mois_annee"] == mois_sel].copy()
      df_m["ouvrage_elem"] = (
          df_m["ouvrage"].fillna("") + " - " + df_m["element_betonne"].fillna("")
      )

      # Assurer la conversion numérique
      df_m["affaissement"] = pd.to_numeric(
          df_m["affaissement"], errors="coerce"
      )
      df_m["tbf"] = pd.to_numeric(df_m["tbf"], errors="coerce")

      df_recap = (
          df_m.groupby(["dt", "date_betonnage", "ouvrage_elem"])
          .agg(
              nb_controles=("ouvrage_elem", "count"),
              aff_min=("affaissement", "min"),
              aff_max=("affaissement", "max"),
              tbf_min=("tbf", "min"),
              tbf_max=("tbf", "max"),
          )
          .reset_index()
          .sort_values(["dt"])
      )

      df_f = df_recap[[
          "date_betonnage",
          "ouvrage_elem",
          "nb_controles",
          "aff_min",
          "aff_max",
          "tbf_min",
          "tbf_max",
      ]].rename(
          columns={
              "date_betonnage": "Date",
              "ouvrage_elem": "Ouvrage / Élément Bétonné",
              "nb_controles": "Nombre de contrôles",
              "aff_min": "Affaissement Min (mm)",
              "aff_max": "Affaissement Max (mm)",
              "tbf_min": "TBF Min (°C)",
              "tbf_max": "TBF Max (°C)",
          }
      )
      st.dataframe(df_f, use_container_width=True)

      excel_m = generer_excel_recap_mensuel_lpee(df_f, mois_sel)
      st.download_button(
          "📊 Télécharger le Récap Mensuel Officiel (.xlsx)",
          data=excel_m,
          file_name=f"Recap_Mensuel_LPEE_{mois_sel.replace('/', '_')}.xlsx",
      )
    else:
      st.info("Aucune date valide trouvée pour la synthèse mensuelle.")
  else:
    st.info("Aucune donnée disponible pour le récapitulatif mensuel.")

# =========================================================
# 6. TABLEAUX ET FICHIER EXCEL JOURNALIER (FORMAT PORTRAIT)
# =========================================================
st.markdown("---")
st.subheader(f"📋 Rapport Journalier de Bétonnage — Date : {str_date_choisie}")

if data_jour:
  df_jour = pd.DataFrame(data_jour)
  classes_du_jour = df_jour["classe_beton"].dropna().unique().tolist()

  vol_tot_j = (
      pd.to_numeric(df_jour["volume_beton"], errors="coerce").fillna(0).sum()
  )
  m1, m2, m3 = st.columns(3)
  with m1:
    st.metric("🏗️ Volume Total du Jour", f"{vol_tot_j:.2f} m³")
  with m2:
    st.metric("🚛 Total Camions (Jour)", f"{len(df_jour)}")
  with m3:
    st.metric("🏷️ Classes de Béton", f"{len(classes_du_jour)} classe(s)")

  st.markdown("---")

  cls_selectionnee = st.selectbox(
      "🔎 Filtrer l'affichage par Classe béton :",
      ["TOUTES LES CLASSES"] + classes_du_jour,
      key="select_classe_view",
  )

  if cls_selectionnee == "TOUTES LES CLASSES":
    df_visu = df_jour.copy()
  else:
    df_visu = df_jour[df_jour["classe_beton"] == cls_selectionnee].copy()

  # Colonnes affichées (sans technicien ni statut)
  cols_vis = [
      "num_bon_livraison",
      "ouvrage",
      "element_betonne",
      "volume_beton",
      "classe_beton",
      "heure_fin_production_cab",
      "heure_arrivee_chantier",
      "tbf",
      "ta",
      "affaissement",
      "prelevement",
  ]
  df_visu_display = df_visu[[c for c in cols_vis if c in df_visu.columns]]
  df_visu_display.index = range(1, len(df_visu_display) + 1)

  st.dataframe(df_visu_display, use_container_width=True)

  excel_jour_bytes = generer_excel_lpee(df_jour, date_rapport=str_date_choisie)
  nom_excel_jour = (
      f"Rapport_Betonnage_LPEE_{date_choisie.strftime('%Y-%m-%d')}.xlsx"
  )

  st.download_button(
      label=(
          "📥 Télécharger le Rapport Journalier Officiel LPEE (.xlsx) —"
          " Impression A4 Portrait"
      ),
      data=excel_jour_bytes,
      file_name=nom_excel_jour,
      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
  )
else:
  st.warning(f"Aucun camion enregistré pour le {str_date_choisie}.")
    import streamlit as st

# 🔒 VÉRIFICATION DE SÉCURITÉ OBLIGATOIRE
if not st.session_state.get("authenticated", False):
    st.error("⛔ Accès refusé. Veuillez d'abord vous connecter sur la page d'accueil (app).")
    st.stop()
