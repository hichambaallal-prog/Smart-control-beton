from datetime import date, datetime
import io
import re
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import plotly.graph_objects as go
import streamlit as st

# =========================================================
# UTILITAIRES D'EXTRACTION NUMÉRIQUE & FCK
# =========================================================


def extract_numeric(val):
  if pd.isna(val) or val is None:
    return None
  val_str = str(val).replace(",", ".")
  match = re.search(r"[-+]?\d*\.\d+|\d+", val_str)
  if match:
    num = float(match.group())
    return int(num) if num.is_integer() else num
  return None


def get_default_fck(df_display):
  if "Classe Béton" in df_display.columns and not df_display.empty:
    first_class = df_display["Classe Béton"].dropna().first_valid_index()
    if first_class is not None:
      val = extract_numeric(df_display.loc[first_class, "Classe Béton"])
      if val is not None and val > 0:
        return float(val)
  elif "classe_beton" in df_display.columns and not df_display.empty:
    first_class = df_display["classe_beton"].dropna().first_valid_index()
    if first_class is not None:
      val = extract_numeric(df_display.loc[first_class, "classe_beton"])
      if val is not None and val > 0:
        return float(val)
  return 35.0


# =========================================================
# GÉNÉRATION DES FICHIERS POUR LA COURBE (EXCEL & PDF)
# =========================================================


def generate_excel_courbe(
    df_grouped, classe_selectionnee, fck_cible, mode_agreg
):
  output = io.BytesIO()
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "Courbe & Données"

  # En-têtes
  ws.append(
      ["Période", f"Cible ({fck_cible} MPa)", "RC 7J (MPa)", "RC 28J (MPa)"]
  )

  # Style en-têtes
  font_th = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
  fill_th = PatternFill(
      start_color="1F4E79", end_color="1F4E79", fill_type="solid"
  )
  for col_i in range(1, 5):
    cell = ws.cell(row=1, column=col_i)
    cell.font = font_th
    cell.fill = fill_th
    cell.alignment = Alignment(horizontal="center", vertical="center")

  # Insertion des données
  for _, row in df_grouped.iterrows():
    rc7 = float(row["RC_7J"]) if pd.notna(row["RC_7J"]) else None
    rc28 = float(row["RC_28J"]) if pd.notna(row["RC_28J"]) else None
    ws.append([str(row["X_Axis"]), float(fck_cible), rc7, rc28])

  # Formatage des cellules de données
  for r in range(2, len(df_grouped) + 2):
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    for c in range(2, 5):
      cell = ws.cell(row=r, column=c)
      cell.alignment = Alignment(horizontal="center")
      if cell.value is not None:
        cell.number_format = "0.0"

  # Ajustement des largeurs de colonnes
  ws.column_dimensions["A"].width = 16
  ws.column_dimensions["B"].width = 16
  ws.column_dimensions["C"].width = 16
  ws.column_dimensions["D"].width = 16

  # Création du graphique natif Excel
  chart = LineChart()
  chart.title = (
      f"Évolution des Résistances ({mode_agreg}) - Classe {classe_selectionnee}"
  )
  chart.style = 10
  chart.y_axis.title = "Résistance à la compression (MPa)"
  chart.x_axis.title = "Date / Période"
  chart.width = 18
  chart.height = 10

  data = Reference(
      ws, min_col=2, min_row=1, max_col=4, max_row=len(df_grouped) + 1
  )
  cats = Reference(ws, min_col=1, min_row=2, max_row=len(df_grouped) + 1)

  chart.add_data(data, titles_from_data=True)
  chart.set_categories(cats)

  # Positionnement du graphique dans la feuille
  ws.add_chart(chart, "F2")

  wb.save(output)
  output.seek(0)
  return output.getvalue()


def generate_pdf_courbe(
    df_grouped, classe_selectionnee, fck_cible, mode_agreg
):
  """Génère un fichier PDF haute définition contenant la courbe d'évolution des résistances."""
  fig_pdf, ax = plt.subplots(figsize=(10, 5), dpi=300)

  # Ligne de référence Cible
  ax.plot(
      df_grouped["X_Axis"],
      [fck_cible] * len(df_grouped),
      label=f"Cible ({fck_cible} MPa)",
      color="#1f77b4",
      linestyle="--",
      linewidth=2,
  )

  # Courbe RC 28J (filtration des NaN pour relier les points)
  df_28j = df_grouped.dropna(subset=["RC_28J"])
  if not df_28j.empty:
    ax.plot(
        df_28j["X_Axis"],
        df_28j["RC_28J"],
        label="RC 28J",
        color="#5DADE2",
        marker="o",
        linewidth=2.5,
    )
    for _, row in df_28j.iterrows():
      ax.annotate(
          f"{row['RC_28J']:.1f}",
          (row["X_Axis"], row["RC_28J"]),
          textcoords="offset points",
          xytext=(0, 7),
          ha="center",
          fontsize=8,
          fontweight="bold",
          color="#2C3E50",
      )

  # Courbe RC 7J (filtration des NaN pour relier les points)
  df_7j = df_grouped.dropna(subset=["RC_7J"])
  if not df_7j.empty:
    ax.plot(
        df_7j["X_Axis"],
        df_7j["RC_7J"],
        label="RC 7J",
        color="#E74C3C",
        marker="o",
        linewidth=2.5,
    )
    for _, row in df_7j.iterrows():
      ax.annotate(
          f"{row['RC_7J']:.1f}",
          (row["X_Axis"], row["RC_7J"]),
          textcoords="offset points",
          xytext=(0, -14),
          ha="center",
          fontsize=8,
          fontweight="bold",
          color="#900C3F",
      )

  ax.set_title(
      f"Évolution des Résistances ({mode_agreg}) - Classe {classe_selectionnee}",
      fontsize=12,
      fontweight="bold",
      pad=15,
  )
  ax.set_xlabel("Date / Période", fontsize=10, fontweight="bold")
  ax.set_ylabel(
      "Résistance à la compression (MPa)", fontsize=10, fontweight="bold"
  )
  ax.grid(True, linestyle=":", alpha=0.6)
  ax.legend(
      loc="upper center", bbox_to_anchor=(0.5, -0.18), ncol=3, frameon=True
  )

  # Alignement de l'axe X pour afficher l'ensemble des dates de la période
  ax.set_xticks(range(len(df_grouped)))
  ax.set_xticklabels(
      df_grouped["X_Axis"], rotation=45 if len(df_grouped) > 8 else 0
  )

  plt.tight_layout()

  output = io.BytesIO()
  plt.savefig(output, format="pdf", bbox_inches="tight")
  plt.close(fig_pdf)
  output.seek(0)
  return output.getvalue()


# =========================================================
# GRAPHIQUE PLOTLY : ÉVOLUTION DES RÉSISTANCES
# =========================================================


def afficher_evolution_resistances(df, key_suffix=""):
  st.markdown("### 📈 Évolution Chronologique des Résistances")

  if df.empty:
    st.info("Données insuffisantes pour afficher l'évolution.")
    return

  df_plot = df.copy()

  col_map = {
      "Date Coulée": "date_prelevement",
      "Date Prélèvement": "date_prelevement",
      "date_coulee": "date_prelevement",
      "date_prelevement": "date_prelevement",
      "Classe Béton": "classe_beton",
      "classe_beton": "classe_beton",
      "Classe": "classe_beton",
      "Moy. Fc (MPa) [7 Jours]": "rc_7j",
      "Fc (MPa) [7 Jours]": "rc_7j",
      "fc_mpa_7 jours": "rc_7j",
      "rc_7j": "rc_7j",
      "Moy. Fc (MPa) [28 Jours]": "rc_28j",
      "Fc (MPa) [28 Jours]": "rc_28j",
      "fc_mpa_28 jours": "rc_28j",
      "rc_28j": "rc_28j",
  }
  df_plot = df_plot.rename(
      columns={k: v for k, v in col_map.items() if k in df_plot.columns}
  )

  for col in df_plot.columns:
    c_str = str(col).lower().strip()
    if "date" in c_str and "date_prelevement" not in df_plot.columns:
      df_plot["date_prelevement"] = df_plot[col]
    elif "classe" in c_str and "classe_beton" not in df_plot.columns:
      df_plot["classe_beton"] = df_plot[col]
    elif (
        "7" in c_str
        and ("fc" in c_str or "jour" in c_str or "rc" in c_str)
        and "rc_7j" not in df_plot.columns
    ):
      df_plot["rc_7j"] = df_plot[col]
    elif (
        "28" in c_str
        and ("fc" in c_str or "jour" in c_str or "rc" in c_str)
        and "rc_28j" not in df_plot.columns
    ):
      df_plot["rc_28j"] = df_plot[col]

  if "date_prelevement" in df_plot.columns:
    df_plot["date_prelevement"] = pd.to_datetime(
        df_plot["date_prelevement"], errors="coerce"
    )
    df_plot = df_plot.dropna(subset=["date_prelevement"])
  else:
    st.info("Colonne de date non identifiée.")
    return

  if df_plot.empty:
    st.info("Aucune date valide pour le graphique.")
    return

  def parse_mean_value(val):
    if pd.isna(val) or val is None:
      return np.nan
    if isinstance(val, (int, float)):
      return float(val)
    found = re.findall(r"[-+]?\d*\.\d+|\d+", str(val).replace(",", "."))
    if found:
      return np.mean([float(x) for x in found])
    return np.nan

  for col_rc in ["rc_7j", "rc_28j"]:
    if col_rc in df_plot.columns:
      df_plot[col_rc] = df_plot[col_rc].apply(parse_mean_value)
    else:
      df_plot[col_rc] = np.nan

  if "classe_beton" in df_plot.columns:
    classes_disponibles = sorted(
        [
            str(c)
            for c in df_plot["classe_beton"].dropna().unique()
            if str(c).strip()
        ]
    )
  else:
    classes_disponibles = []

  if classes_disponibles:
    default_idx = 0
    for idx, cls in enumerate(classes_disponibles):
      sub_cls = df_plot[df_plot["classe_beton"] == cls]
      if sub_cls["rc_7j"].notna().any() or sub_cls["rc_28j"].notna().any():
        default_idx = idx
        break

    classe_selectionnee = st.selectbox(
        "Sélectionner la classe de béton :",
        classes_disponibles,
        index=default_idx,
        key=f"select_classe_beton_{key_suffix}",
    )
    df_classe = df_plot[df_plot["classe_beton"] == classe_selectionnee].copy()
  else:
    classe_selectionnee = "Tous"
    df_classe = df_plot.copy()

  if df_classe.empty:
    st.warning("Aucune donnée enregistrée pour cette classe de béton.")
    return

  col_cfg1, col_cfg2 = st.columns([1, 1])
  with col_cfg1:
    mode_agreg = st.radio(
        "Granularité d'affichage :",
        ["Par Jour", "Par Mois"],
        horizontal=True,
        key=f"agreg_mode_{key_suffix}",
    )

  with col_cfg2:
    if "fck" in df_classe.columns and pd.notnull(df_classe["fck"].iloc[0]):
      fck_defaut = float(df_classe["fck"].iloc[0])
    else:
      fck_defaut = float(get_default_fck(df_classe))

    fck_cible = st.number_input(
        "Valeur cible fck (MPa) :",
        value=fck_defaut,
        step=1.0,
        key=f"fck_input_{key_suffix}",
    )

  if mode_agreg == "Par Jour":
    df_classe["Period_Str"] = df_classe["date_prelevement"].dt.strftime(
        "%d/%m/%Y"
    )
    df_grouped = (
        df_classe.groupby(["date_prelevement", "Period_Str"])
        .agg(RC_7J=("rc_7j", "mean"), RC_28J=("rc_28j", "mean"))
        .reset_index()
        .sort_values("date_prelevement")
    )
    df_grouped["X_Axis"] = df_grouped["Period_Str"]
  else:
    df_classe["Period"] = df_classe["date_prelevement"].dt.to_period("M")
    df_grouped = (
        df_classe.groupby("Period")
        .agg(RC_7J=("rc_7j", "mean"), RC_28J=("rc_28j", "mean"))
        .reset_index()
        .sort_values("Period")
    )
    df_grouped["X_Axis"] = df_grouped["Period"].dt.strftime("%m/%Y")

  if df_grouped.empty or (
      df_grouped["RC_7J"].isna().all() and df_grouped["RC_28J"].isna().all()
  ):
    st.info("Aucune donnée d'écrasement disponible pour cette classe.")
    return

  fig = go.Figure()

  fig.add_trace(
      go.Scatter(
          x=df_grouped["X_Axis"],
          y=[fck_cible] * len(df_grouped),
          mode="lines",
          name=f"Cible ({fck_cible} MPa)",
          line=dict(color="#1f77b4", width=2.5, dash="dash"),
      )
  )

  if df_grouped["RC_28J"].notna().any():
    fig.add_trace(
        go.Scatter(
            x=df_grouped["X_Axis"],
            y=df_grouped["RC_28J"],
            mode="lines+markers+text",
            name="RC 28J",
            text=df_grouped["RC_28J"].apply(
                lambda v: f"{v:.1f}" if pd.notna(v) else ""
            ),
            textposition="top center",
            line=dict(color="#5DADE2", width=3),
            marker=dict(size=8),
            connectgaps=True,
        )
    )

  if df_grouped["RC_7J"].notna().any():
    fig.add_trace(
        go.Scatter(
            x=df_grouped["X_Axis"],
            y=df_grouped["RC_7J"],
            mode="lines+markers+text",
            name="RC 7J",
            text=df_grouped["RC_7J"].apply(
                lambda v: f"{v:.1f}" if pd.notna(v) else ""
            ),
            textposition="bottom center",
            line=dict(color="#E74C3C", width=3),
            marker=dict(size=8),
            connectgaps=True,
        )
    )

  fig.update_layout(
      title=dict(
          text=(
              f"Évolution des Résistances ({mode_agreg}) - Classe"
              f" {classe_selectionnee}"
          ),
          font=dict(size=18),
      ),
      xaxis=dict(
          title="Date / Période",
          type="category",
          showgrid=True,
          gridcolor="#E5E8E8",
      ),
      yaxis=dict(
          title="Résistance à la compression (MPa)",
          showgrid=True,
          gridcolor="#E5E8E8",
          zeroline=False,
      ),
      legend=dict(
          orientation="h", yanchor="bottom", y=-0.25, xanchor="center", x=0.5
      ),
      plot_bgcolor="white",
      height=500,
      margin=dict(l=40, r=40, t=60, b=60),
  )

  st.plotly_chart(fig, use_container_width=True)

  # Exportation Téléchargement : Excel + PDF
  col_dl1, col_dl2 = st.columns(2)

  excel_courbe_bytes = generate_excel_courbe(
      df_grouped, classe_selectionnee, fck_cible, mode_agreg
  )
  pdf_courbe_bytes = generate_pdf_courbe(
      df_grouped, classe_selectionnee, fck_cible, mode_agreg
  )
  clean_cls = str(classe_selectionnee).replace(" ", "_").replace("/", "-")

  with col_dl1:
    st.download_button(
        label="📊 Télécharger la courbe en Excel (.xlsx)",
        data=excel_courbe_bytes,
        file_name=f"Courbe_Evolution_Resistances_{clean_cls}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"download_fig_excel_{key_suffix}",
    )

  with col_dl2:
    st.download_button(
        label="📄 Télécharger la courbe en PDF (.pdf)",
        data=pdf_courbe_bytes,
        file_name=f"Courbe_Evolution_Resistances_{clean_cls}.pdf",
        mime="application/pdf",
        key=f"download_fig_pdf_{key_suffix}",
    )


# =========================================================
# 1. GENERATION EXCEL (CONFORME LPEE - LGV CASA SUD)
# =========================================================


def generate_excel_synthesis_betonnage(
    df_data, titre_periode, is_mensuel=False
):
  output = io.BytesIO()
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "Synthèse Bétonnage"

  ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
  ws.page_setup.paperSize = ws.PAPERSIZE_A4
  ws.sheet_properties.pageSetUpPr.fitToPage = True
  ws.page_setup.fitToWidth = 1
  ws.page_setup.fitToHeight = 0

  ws.page_margins.left = 0.3
  ws.page_margins.right = 0.3
  ws.page_margins.top = 0.4
  ws.page_margins.bottom = 0.4

  color_primary = "1F4E79"
  color_header = "2D572C"
  color_sub_header = "E2EFDA"
  color_card_bg = "F7F9FA"
  color_kpi_bg = "EDF2F8"
  color_stat_bg = "F2F2F2"

  font_title = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
  font_section = Font(name="Calibri", size=13, bold=True, color=color_primary)
  font_bold = Font(name="Calibri", size=11, bold=True)
  font_normal = Font(name="Calibri", size=11)
  font_th = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
  font_sub_th = Font(name="Calibri", size=10, bold=True, color="000000")
  font_kpi_val = Font(name="Calibri", size=14, bold=True, color=color_primary)

  fill_title = PatternFill(
      start_color=color_primary, end_color=color_primary, fill_type="solid"
  )
  fill_th = PatternFill(
      start_color=color_header, end_color=color_header, fill_type="solid"
  )
  fill_sub_th = PatternFill(
      start_color=color_sub_header, end_color=color_sub_header, fill_type="solid"
  )
  fill_card = PatternFill(
      start_color=color_card_bg, end_color=color_card_bg, fill_type="solid"
  )
  fill_kpi = PatternFill(
      start_color=color_kpi_bg, end_color=color_kpi_bg, fill_type="solid"
  )
  fill_stat = PatternFill(
      start_color=color_stat_bg, end_color=color_stat_bg, fill_type="solid"
  )

  thin_border_side = Side(style="thin", color="B0C4DE")
  thin_border = Border(
      left=thin_border_side,
      right=thin_border_side,
      top=thin_border_side,
      bottom=thin_border_side,
  )
  dark_thin_side = Side(style="thin", color="000000")
  dark_border = Border(
      left=dark_thin_side,
      right=dark_thin_side,
      top=dark_thin_side,
      bottom=dark_thin_side,
  )
  total_border = Border(
      top=Side(style="thin", color="000000"),
      bottom=Side(style="double", color="000000"),
  )

  is_multi = isinstance(df_data.columns, pd.MultiIndex)
  nb_cols = min(max(len(df_data.columns), 10), 10)
  last_col_letter = get_column_letter(nb_cols)
  mid_col_idx = max(nb_cols // 2, 1)
  mid_col_letter = get_column_letter(mid_col_idx)
  next_mid_letter = get_column_letter(mid_col_idx + 1)

  ws.merge_cells(f"A1:{last_col_letter}2")
  cell_title = ws["A1"]
  cell_title.value = (
      "LABORATOIRE PUBLIC D'ESSAIS ET D'ÉTUDES (LPEE) -"
      " CTR-CSB\nRAPPORT DE SYNTHÈSE DU BÉTONNAGE"
  )
  cell_title.font = font_title
  cell_title.fill = fill_title
  cell_title.alignment = Alignment(
      horizontal="center", vertical="center", wrap_text=True
  )
  ws.row_dimensions[1].height = 25
  ws.row_dimensions[2].height = 25

  ws.merge_cells(f"A4:{mid_col_letter}4")
  ws["A4"].value = "   CLIENT :   TGCC"
  ws["A4"].font = font_bold
  ws["A4"].fill = fill_card
  ws["A4"].alignment = Alignment(horizontal="left", vertical="center")

  ws.merge_cells(f"{next_mid_letter}4:{last_col_letter}4")
  ws[f"{next_mid_letter}4"].value = "   PROJET :   LGV CASA SUD"
  ws[f"{next_mid_letter}4"].font = font_bold
  ws[f"{next_mid_letter}4"].fill = fill_card
  ws[f"{next_mid_letter}4"].alignment = Alignment(
      horizontal="left", vertical="center"
  )

  ws.merge_cells(f"A5:{mid_col_letter}5")
  ws["A5"].value = f"   PÉRIODE :   {titre_periode}"
  ws["A5"].font = font_bold
  ws["A5"].fill = fill_card
  ws["A5"].alignment = Alignment(horizontal="left", vertical="center")

  ws.merge_cells(f"{next_mid_letter}5:{last_col_letter}5")
  ws[f"{next_mid_letter}5"].value = (
      f"   DATE ÉDITION :   {datetime.now().strftime('%d/%m/%Y')}"
  )
  ws[f"{next_mid_letter}5"].font = font_bold
  ws[f"{next_mid_letter}5"].fill = fill_card
  ws[f"{next_mid_letter}5"].alignment = Alignment(
      horizontal="left", vertical="center"
  )

  for r in range(4, 6):
    ws.row_dimensions[r].height = 28
    for c in range(1, nb_cols + 1):
      ws.cell(row=r, column=c).border = thin_border

  row_idx = 7
  ws.merge_cells(f"A{row_idx}:{last_col_letter}{row_idx}")
  ws[f"A{row_idx}"] = "📊 RÉSUMÉ GLOBAL"
  ws[f"A{row_idx}"].font = font_section
  ws.row_dimensions[row_idx].height = 28

  row_idx += 1
  if is_multi:
    vol_col = [c for c in df_data.columns if c[0] == "Quantité (m³)"]
    vol_tot = df_data[vol_col[0]].sum() if vol_col else 0
  else:
    vol_tot = (
        df_data["Quantité (m³)"].sum()
        if "Quantité (m³)" in df_data.columns
        else 0
    )

  ws.merge_cells(f"A{row_idx}:{last_col_letter}{row_idx}")
  ws[f"A{row_idx}"].value = "Volume Total Béton"
  ws[f"A{row_idx}"].font = font_bold
  ws[f"A{row_idx}"].fill = fill_kpi
  ws[f"A{row_idx}"].alignment = Alignment(
      horizontal="center", vertical="center"
  )

  ws.merge_cells(f"A{row_idx+1}:{last_col_letter}{row_idx+1}")
  ws[f"A{row_idx+1}"].value = f"{vol_tot:.1f} m³"
  ws[f"A{row_idx+1}"].font = font_kpi_val
  ws[f"A{row_idx+1}"].fill = fill_kpi
  ws[f"A{row_idx+1}"].alignment = Alignment(
      horizontal="center", vertical="center"
  )

  for r in range(row_idx, row_idx + 2):
    for c in range(1, nb_cols + 1):
      ws.cell(row=r, column=c).border = thin_border

  ws.row_dimensions[row_idx].height = 24
  ws.row_dimensions[row_idx + 1].height = 32
  row_idx += 3

  ws.merge_cells(f"A{row_idx}:{last_col_letter}{row_idx}")
  ws[f"A{row_idx}"] = "📋 DÉTAIL DES CONTRÔLES"
  ws[f"A{row_idx}"].font = font_section
  ws.row_dimensions[row_idx].height = 28
  row_idx += 1

  if is_multi:
    col_i = 1
    for top_cat, sub_cat in df_data.columns[:nb_cols]:
      cell_top = ws.cell(row=row_idx, column=col_i)
      cell_top.value = top_cat
      cell_top.font = font_th
      cell_top.fill = fill_th
      cell_top.alignment = Alignment(
          horizontal="center", vertical="center", wrap_text=True
      )
      cell_top.border = thin_border

      cell_sub = ws.cell(row=row_idx + 1, column=col_i)
      cell_sub.value = sub_cat
      cell_sub.font = font_sub_th
      cell_sub.fill = fill_sub_th
      cell_sub.alignment = Alignment(
          horizontal="center", vertical="center", wrap_text=True
      )
      cell_sub.border = thin_border
      col_i += 1

    col_i = 1
    real_len = min(len(df_data.columns), nb_cols)
    while col_i <= real_len:
      top_val = df_data.columns[col_i - 1][0]
      span = sum(
          1
          for c in df_data.columns[:nb_cols]
          if c[0] == top_val and top_val != ""
      )
      if span > 1:
        ws.merge_cells(
            start_row=row_idx,
            start_column=col_i,
            end_row=row_idx,
            end_column=col_i + span - 1,
        )
        col_i += span
      else:
        if df_data.columns[col_i - 1][1] == "":
          ws.merge_cells(
              start_row=row_idx,
              start_column=col_i,
              end_row=row_idx + 1,
              end_column=col_i,
          )
        col_i += 1

    ws.row_dimensions[row_idx].height = 25
    ws.row_dimensions[row_idx + 1].height = 22
    row_idx += 2
  else:
    for col_num, h_name in enumerate(df_data.columns[:nb_cols], 1):
      cell = ws.cell(row=row_idx, column=col_num)
      cell.value = str(h_name)
      cell.font = font_th
      cell.fill = fill_th
      cell.alignment = Alignment(
          horizontal="center", vertical="center", wrap_text=True
      )
      cell.border = thin_border

    ws.row_dimensions[row_idx].height = 35
    row_idx += 1

  start_data_row = row_idx
  for row_data in df_data.itertuples(index=False):
    for col_num, val in enumerate(row_data[:nb_cols], 1):
      cell = ws.cell(row=row_idx, column=col_num)
      cell.value = val if pd.notna(val) else ""
      cell.font = font_normal
      cell.border = thin_border
      cell.alignment = Alignment(
          horizontal="center", vertical="center", wrap_text=True
      )

    ws.row_dimensions[row_idx].height = 26
    row_idx += 1

  end_data_row = row_idx - 1

  headers_flat = [
      c[0] if is_multi else str(c) for c in df_data.columns[:nb_cols]
  ]
  target_keywords = [
      "temp. béton",
      "temp. ambiante",
      "affaissement",
      "temperature",
      "aff",
  ]

  def is_target_col(col_name):
    c_lower = col_name.lower()
    return any(kw in c_lower for kw in target_keywords)

  target_col_indices = [
      idx + 1 for idx, h in enumerate(headers_flat) if is_target_col(h)
  ]

  for stat_label, stat_func in [("MIN", "MIN"), ("MAX", "MAX")]:
    ws.row_dimensions[row_idx].height = 26

    lbl_cell = ws.cell(row=row_idx, column=1)
    lbl_cell.value = stat_label
    lbl_cell.font = font_bold
    lbl_cell.fill = fill_stat
    lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
    lbl_cell.border = dark_border

    for col_num in range(2, len(headers_flat) + 1):
      c = ws.cell(row=row_idx, column=col_num)
      c.border = dark_border
      c.font = font_bold
      c.fill = fill_stat
      c.alignment = Alignment(horizontal="center", vertical="center")

      if col_num in target_col_indices:
        col_ltr = get_column_letter(col_num)
        if start_data_row <= end_data_row:
          c.value = (
              f"={stat_func}({col_ltr}{start_data_row}:{col_ltr}{end_data_row})"
          )
          c.number_format = "0.0"
        else:
          c.value = "-"
      else:
        c.value = ""

    row_idx += 1

  ws.row_dimensions[row_idx].height = 30
  total_cell = ws.cell(row=row_idx, column=1)
  total_cell.value = "TOTAL"
  total_cell.font = font_bold
  total_cell.border = total_border
  total_cell.alignment = Alignment(
      horizontal="center", vertical="center", wrap_text=True
  )

  for col_num in range(1, len(headers_flat) + 1):
    c = ws.cell(row=row_idx, column=col_num)
    c.border = total_border
    c.font = font_bold
    col_name = headers_flat[col_num - 1]
    col_ltr = get_column_letter(col_num)
    if "quantité" in col_name.lower() or "volume" in col_name.lower():
      c.value = f"=SUM({col_ltr}{start_data_row}:{col_ltr}{end_data_row})"
      c.number_format = '0.0 "m³"'
      c.alignment = Alignment(
          horizontal="center", vertical="center", wrap_text=True
      )

  row_idx += 3
  ws.merge_cells(
      start_row=row_idx,
      start_column=1,
      end_row=row_idx,
      end_column=mid_col_idx,
  )
  cell_resp = ws.cell(row=row_idx, column=1, value="Responsable d'essai :")
  cell_resp.font = font_bold
  cell_resp.alignment = Alignment(horizontal="center", vertical="center")

  ws.merge_cells(
      start_row=row_idx,
      start_column=mid_col_idx + 1,
      end_row=row_idx,
      end_column=nb_cols,
  )
  cell_chef = ws.cell(
      row=row_idx, column=mid_col_idx + 1, value="Chef du laboratoire :"
  )
  cell_chef.font = font_bold
  cell_chef.alignment = Alignment(horizontal="center", vertical="center")

  ws.row_dimensions[row_idx].height = 20
  row_idx += 1

  ws.merge_cells(
      start_row=row_idx,
      start_column=1,
      end_row=row_idx,
      end_column=mid_col_idx,
  )
  cell_nam_resp = ws.cell(row=row_idx, column=1, value="O.IKKEN")
  cell_nam_resp.font = Font(name="Calibri", size=11, bold=True, color="1F4E79")
  cell_nam_resp.alignment = Alignment(horizontal="center", vertical="center")

  ws.merge_cells(
      start_row=row_idx,
      start_column=mid_col_idx + 1,
      end_row=row_idx,
      end_column=nb_cols,
  )
  cell_nam_chef = ws.cell(row=row_idx, column=mid_col_idx + 1, value="H.BAALLAL")
  cell_nam_chef.font = Font(name="Calibri", size=11, bold=True, color="1F4E79")
  cell_nam_chef.alignment = Alignment(horizontal="center", vertical="center")

  ws.row_dimensions[row_idx].height = 20

  ws.column_dimensions["A"].width = 14
  ws.column_dimensions["B"].width = 12
  ws.column_dimensions["C"].width = 11.22
  ws.column_dimensions["D"].width = 40.0
  ws.column_dimensions["E"].width = 11.22
  ws.column_dimensions["F"].width = 11.22
  ws.column_dimensions["G"].width = 11.22
  ws.column_dimensions["H"].width = 11.22
  ws.column_dimensions["I"].width = 12
  ws.column_dimensions["J"].width = 12

  wb.save(output)
  output.seek(0)
  return output.getvalue()


def generate_excel_synthesis_controle(df_data, titre_periode):
  output = io.BytesIO()
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "Synthèse Contrôle Béton"

  ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
  ws.page_setup.paperSize = ws.PAPERSIZE_A4
  ws.sheet_properties.pageSetUpPr.fitToPage = True
  ws.page_setup.fitToWidth = 1
  ws.page_setup.fitToHeight = 0

  ws.page_margins.left = 0.3
  ws.page_margins.right = 0.3
  ws.page_margins.top = 0.4
  ws.page_margins.bottom = 0.4

  color_primary = "1F4E79"
  color_header = "2D572C"
  color_card_bg = "F7F9FA"
  color_kpi_bg = "EDF2F8"
  color_stat_hdr = "1F4E79"

  font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
  font_section = Font(name="Calibri", size=12, bold=True, color=color_primary)
  font_bold = Font(name="Calibri", size=11, bold=True)
  font_normal = Font(name="Calibri", size=11)
  font_th = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
  font_kpi_val = Font(name="Calibri", size=13, bold=True, color=color_primary)

  fill_title = PatternFill(
      start_color=color_primary, end_color=color_primary, fill_type="solid"
  )
  fill_th = PatternFill(
      start_color=color_header, end_color=color_header, fill_type="solid"
  )
  fill_card = PatternFill(
      start_color=color_card_bg, end_color=color_card_bg, fill_type="solid"
  )
  fill_kpi = PatternFill(
      start_color=color_kpi_bg, end_color=color_kpi_bg, fill_type="solid"
  )
  fill_stat_hdr = PatternFill(
      start_color=color_stat_hdr, end_color=color_stat_hdr, fill_type="solid"
  )

  thin_border_side = Side(style="thin", color="B0C4DE")
  thin_border = Border(
      left=thin_border_side,
      right=thin_border_side,
      top=thin_border_side,
      bottom=thin_border_side,
  )

  nb_cols = min(max(len(df_data.columns), 9), 10)
  last_col_letter = get_column_letter(nb_cols)
  mid_col_idx = max(nb_cols // 2, 1)
  mid_col_letter = get_column_letter(mid_col_idx)
  next_mid_letter = get_column_letter(mid_col_idx + 1)

  ws.merge_cells(f"A1:{last_col_letter}2")
  ws["A1"].value = (
      "LABORATOIRE PUBLIC D'ESSAIS ET D'ÉTUDES (LPEE) -"
      " CTR-CSB\nRAPPORT DE SYNTHÈSE DU CONTRÔLE BÉTON"
  )
  ws["A1"].font = font_title
  ws["A1"].fill = fill_title
  ws["A1"].alignment = Alignment(
      horizontal="center", vertical="center", wrap_text=True
  )

  ws.row_dimensions[1].height = 25
  ws.row_dimensions[2].height = 25

  ws.merge_cells(f"A4:{mid_col_letter}4")
  ws["A4"].value = "   CLIENT :   TGCC"
  ws["A4"].font = font_bold
  ws["A4"].fill = fill_card
  ws["A4"].alignment = Alignment(horizontal="left", vertical="center")

  ws.merge_cells(f"{next_mid_letter}4:{last_col_letter}4")
  ws[f"{next_mid_letter}4"].value = "   PROJET :   LGV CASA SUD"
  ws[f"{next_mid_letter}4"].font = font_bold
  ws[f"{next_mid_letter}4"].fill = fill_card
  ws[f"{next_mid_letter}4"].alignment = Alignment(
      horizontal="left", vertical="center"
  )

  ws.merge_cells(f"A5:{mid_col_letter}5")
  ws["A5"].value = f"   PÉRIODE :   {titre_periode}"
  ws["A5"].font = font_bold
  ws["A5"].fill = fill_card
  ws["A5"].alignment = Alignment(horizontal="left", vertical="center")

  ws.merge_cells(f"{next_mid_letter}5:{last_col_letter}5")
  ws[f"{next_mid_letter}5"].value = (
      f"   DATE ÉDITION :   {datetime.now().strftime('%d/%m/%Y')}"
  )
  ws[f"{next_mid_letter}5"].font = font_bold
  ws[f"{next_mid_letter}5"].fill = fill_card
  ws[f"{next_mid_letter}5"].alignment = Alignment(
      horizontal="left", vertical="center"
  )

  for r in range(4, 6):
    ws.row_dimensions[r].height = 28
    for c in range(1, nb_cols + 1):
      ws.cell(row=r, column=c).border = thin_border

  row_idx = 7
  ws.merge_cells(f"A{row_idx}:{last_col_letter}{row_idx}")
  ws[f"A{row_idx}"] = "📊 RÉSUMÉ GLOBAL"
  ws[f"A{row_idx}"].font = font_section

  row_idx += 1
  ws.merge_cells(f"A{row_idx}:{last_col_letter}{row_idx}")
  ws[f"A{row_idx}"].value = "Nombre Total de Prélèvements Contrôlés"
  ws[f"A{row_idx}"].font = font_bold
  ws[f"A{row_idx}"].fill = fill_kpi
  ws[f"A{row_idx}"].alignment = Alignment(
      horizontal="center", vertical="center"
  )

  ws.merge_cells(f"A{row_idx+1}:{last_col_letter}{row_idx+1}")
  ws[f"A{row_idx+1}"].value = f"{len(df_data)} prélèvement(s)"
  ws[f"A{row_idx+1}"].font = font_kpi_val
  ws[f"A{row_idx+1}"].fill = fill_kpi
  ws[f"A{row_idx+1}"].alignment = Alignment(
      horizontal="center", vertical="center"
  )

  for r in range(row_idx, row_idx + 2):
    for c in range(1, nb_cols + 1):
      ws.cell(row=r, column=c).border = thin_border

  row_idx += 3
  ws.merge_cells(f"A{row_idx}:{last_col_letter}{row_idx}")
  ws[f"A{row_idx}"] = "📋 DÉTAIL DES ÉCRASEMENTS PAR ÉCHÉANCE"
  ws[f"A{row_idx}"].font = font_section
  row_idx += 1

  headers = list(df_data.columns[:nb_cols])
  for col_num, h_name in enumerate(headers, 1):
    cell = ws.cell(row=row_idx, column=col_num)
    cell.value = str(h_name)
    cell.font = font_th
    cell.fill = fill_th
    cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

  ws.row_dimensions[row_idx].height = 35
  row_idx += 1

  start_data_row = row_idx
  for row_data in df_data.itertuples(index=False):
    for col_num, val in enumerate(row_data[:nb_cols], 1):
      cell = ws.cell(row=row_idx, column=col_num)
      cell.value = val if pd.notna(val) else ""
      cell.font = font_normal
      cell.border = thin_border
      cell.alignment = Alignment(
          horizontal="center", vertical="center", wrap_text=True
      )
    ws.row_dimensions[row_idx].height = 28
    row_idx += 1
  end_data_row = row_idx - 1

  row_idx += 2
  ws.merge_cells(f"A{row_idx}:{last_col_letter}{row_idx}")
  ws[f"A{row_idx}"] = "📈 SYNTHÈSE STATISTIQUE DES PARAMÈTRES"
  ws[f"A{row_idx}"].font = font_section
  row_idx += 1

  stat_headers = [
      "Indicateur",
      "Affaissement (mm)",
      "Temp. Béton (°C)",
      "Fc (MPa) [7 Jours]",
      "Fc (MPa) [28 Jours]",
  ]

  for idx, h in enumerate(stat_headers):
    col_start = 1 if idx == 0 else 2 + (idx - 1) * 2
    col_end = 1 if idx == 0 else col_start + 1

    if col_start <= nb_cols:
      if col_start == col_end or col_end > nb_cols:
        c = ws.cell(row=row_idx, column=col_start)
        c.value = h
      else:
        ws.merge_cells(
            start_row=row_idx,
            start_column=col_start,
            end_row=row_idx,
            end_column=col_end,
        )
        c = ws.cell(row=row_idx, column=col_start)
        c.value = h

      c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
      c.fill = fill_stat_hdr
      c.alignment = Alignment(
          horizontal="center", vertical="center", wrap_text=True
      )

  for c_i in range(1, nb_cols + 1):
    ws.cell(row=row_idx, column=c_i).border = thin_border

  ws.row_dimensions[row_idx].height = 25
  row_idx += 1

  col_map_excel = {"aff": "E", "temp": "F", "fc7": "H", "fc28": "I"}

  row_moy = row_idx + 1
  row_std = row_idx + 3

  if start_data_row <= end_data_row:
    f_min_aff = (
        f"=MIN({col_map_excel['aff']}{start_data_row}:{col_map_excel['aff']}{end_data_row})"
    )
    f_moy_aff = (
        f"=AVERAGE({col_map_excel['aff']}{start_data_row}:{col_map_excel['aff']}{end_data_row})"
    )
    f_max_aff = (
        f"=MAX({col_map_excel['aff']}{start_data_row}:{col_map_excel['aff']}{end_data_row})"
    )

    f_min_temp = (
        f"=MIN({col_map_excel['temp']}{start_data_row}:{col_map_excel['temp']}{end_data_row})"
    )
    f_moy_temp = (
        f"=AVERAGE({col_map_excel['temp']}{start_data_row}:{col_map_excel['temp']}{end_data_row})"
    )
    f_max_temp = (
        f"=MAX({col_map_excel['temp']}{start_data_row}:{col_map_excel['temp']}{end_data_row})"
    )

    f_min_fc7 = (
        f"=MIN({col_map_excel['fc7']}{start_data_row}:{col_map_excel['fc7']}{end_data_row})"
    )
    f_moy_fc7 = (
        f"=AVERAGE({col_map_excel['fc7']}{start_data_row}:{col_map_excel['fc7']}{end_data_row})"
    )
    f_max_fc7 = (
        f"=MAX({col_map_excel['fc7']}{start_data_row}:{col_map_excel['fc7']}{end_data_row})"
    )

    f_min_fc28 = (
        f"=MIN({col_map_excel['fc28']}{start_data_row}:{col_map_excel['fc28']}{end_data_row})"
    )
    f_moy_fc28 = (
        f"=AVERAGE({col_map_excel['fc28']}{start_data_row}:{col_map_excel['fc28']}{end_data_row})"
    )
    f_max_fc28 = (
        f"=MAX({col_map_excel['fc28']}{start_data_row}:{col_map_excel['fc28']}{end_data_row})"
    )

    f_std_fc28 = f"=STDEV.S({col_map_excel['fc28']}{start_data_row}:{col_map_excel['fc28']}{end_data_row})"
    f_cv_fc28 = f"=IFERROR(({col_map_excel['fc28']}{row_std}/{col_map_excel['fc28']}{row_moy})*100, 0)"
  else:
    f_min_aff = f_moy_aff = f_max_aff = "-"
    f_min_temp = f_moy_temp = f_max_temp = "-"
    f_min_fc7 = f_moy_fc7 = f_max_fc7 = "-"
    f_min_fc28 = f_moy_fc28 = f_max_fc28 = f_std_fc28 = f_cv_fc28 = "-"

  stat_rows = [
      ("MIN", f_min_aff, f_min_temp, f_min_fc7, f_min_fc28),
      ("MOY", f_moy_aff, f_moy_temp, f_moy_fc7, f_moy_fc28),
      ("MAX", f_max_aff, f_max_temp, f_max_fc7, f_max_fc28),
      ("σ", "-", "-", "-", f_std_fc28),
      ("CV %", "-", "-", "-", f_cv_fc28),
  ]

  for label, v_aff, v_temp, v_fc7, v_fc28 in stat_rows:
    c_lbl = ws.cell(row=row_idx, column=1, value=label)
    c_lbl.font = font_bold
    c_lbl.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    vals = [v_aff, v_temp, v_fc7, v_fc28]
    for idx, val in enumerate(vals):
      c_start = 2 + idx * 2
      c_end = c_start + 1
      if c_start <= nb_cols:
        if c_end <= nb_cols:
          ws.merge_cells(
              start_row=row_idx,
              start_column=c_start,
              end_row=row_idx,
              end_column=c_end,
          )
        c = ws.cell(row=row_idx, column=c_start)
        c.value = val
        c.font = font_normal
        c.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        if isinstance(val, str) and val.startswith("="):
          c.number_format = "0.00" if label in ["σ", "CV %"] else "0.0"

    for c_i in range(1, nb_cols + 1):
      ws.cell(row=row_idx, column=c_i).border = thin_border

    ws.row_dimensions[row_idx].height = 22
    row_idx += 1

  row_idx += 2
  ws.merge_cells(
      start_row=row_idx,
      start_column=1,
      end_row=row_idx,
      end_column=mid_col_idx,
  )
  cell_resp = ws.cell(row=row_idx, column=1, value="Responsable d'essai :")
  cell_resp.font = font_bold
  cell_resp.alignment = Alignment(horizontal="center", vertical="center")

  ws.merge_cells(
      start_row=row_idx,
      start_column=mid_col_idx + 1,
      end_row=row_idx,
      end_column=nb_cols,
  )
  cell_chef = ws.cell(
      row=row_idx, column=mid_col_idx + 1, value="Chef du laboratoire :"
  )
  cell_chef.font = font_bold
  cell_chef.alignment = Alignment(horizontal="center", vertical="center")

  ws.row_dimensions[row_idx].height = 20
  row_idx += 1

  ws.merge_cells(
      start_row=row_idx,
      start_column=1,
      end_row=row_idx,
      end_column=mid_col_idx,
  )
  cell_nam_resp = ws.cell(row=row_idx, column=1, value="O.IKKEN")
  cell_nam_resp.font = Font(name="Calibri", size=11, bold=True, color="1F4E79")
  cell_nam_resp.alignment = Alignment(horizontal="center", vertical="center")

  ws.merge_cells(
      start_row=row_idx,
      start_column=mid_col_idx + 1,
      end_row=row_idx,
      end_column=nb_cols,
  )
  cell_nam_chef = ws.cell(row=row_idx, column=mid_col_idx + 1, value="H.BAALLAL")
  cell_nam_chef.font = Font(name="Calibri", size=11, bold=True, color="1F4E79")
  cell_nam_chef.alignment = Alignment(horizontal="center", vertical="center")

  ws.row_dimensions[row_idx].height = 20

  ws.column_dimensions["A"].width = 14
  ws.column_dimensions["C"].width = 40.0
  cols_12 = ["B", "D", "E", "F", "G", "H", "I", "J"]
  for col_l in cols_12:
    ws.column_dimensions[col_l].width = 12

  wb.save(output)
  output.seek(0)
  return output.getvalue()


# =========================================================
# 2. CHARGEMENT & TRAITEMENT SUPABASE (VALEURS INDIVIDUELLES)
# =========================================================


def load_and_process_controle_data(supabase):
  res_ecrasement = (
      supabase.table("suivi_controle_beton")
      .select("*")
      .order("id", desc=True)
      .execute()
  )
  df_raw = (
      pd.DataFrame(res_ecrasement.data)
      if res_ecrasement and res_ecrasement.data
      else pd.DataFrame()
  )

  if df_raw.empty:
    return pd.DataFrame()

  res_betonnage = supabase.table("suivi_betonnage").select("*").execute()
  df_betonnage = (
      pd.DataFrame(res_betonnage.data)
      if res_betonnage and res_betonnage.data
      else pd.DataFrame()
  )

  col_mapping_input = {
      "affaissement": "affaissement_mm",
      "affaissement_mm": "affaissement_mm",
      "temperature": "temp_beton_C",
      "temp_beton": "temp_beton_C",
      "temp_beton_C": "temp_beton_C",
  }
  df_raw = df_raw.rename(
      columns={k: v for k, v in col_mapping_input.items() if k in df_raw.columns}
  )

  if "affaissement_mm" not in df_raw.columns:
    df_raw["affaissement_mm"] = None
  if "temp_beton_C" not in df_raw.columns:
    df_raw["temp_beton_C"] = None

  if not df_betonnage.empty:
    df_betonnage = df_betonnage.rename(
        columns={
            "affaissement": "affaissement_mm_b",
            "temperature": "temp_beton_C_b",
            "prelevement": "ref_controle_b",
            "date_livraison": "date_coulee_b",
            "ouvrage": "ouvrage_b",
        }
    )

  def clean_echeance(val):
    val_str = str(val).lower().strip()
    if "3" in val_str:
      return "3 jours"
    elif "7" in val_str:
      return "7 jours"
    elif "28" in val_str:
      return "28 jours"
    return val_str

  if "echeance" in df_raw.columns:
    df_raw["echeance_clean"] = df_raw["echeance"].apply(clean_echeance)
  else:
    df_raw["echeance_clean"] = ""

  if "fc_mpa" in df_raw.columns:
    df_raw["fc_mpa"] = pd.to_numeric(df_raw["fc_mpa"], errors="coerce")

  base_group_cols = ["ref_controle", "date_coulee", "classe_beton", "ouvrage"]
  existing_group_cols = [c for c in base_group_cols if c in df_raw.columns]

  if not existing_group_cols:
    return pd.DataFrame()

  if "date_coulee" in df_raw.columns:
    df_raw["date_dt"] = pd.to_datetime(df_raw["date_coulee"], errors="coerce")
  elif "date_ecrasement" in df_raw.columns:
    df_raw["date_dt"] = pd.to_datetime(
        df_raw["date_ecrasement"], errors="coerce"
    )
  else:
    df_raw["date_dt"] = pd.NaT

  pivot_rows = []
  grouped = df_raw.groupby(existing_group_cols, dropna=False)

  for group_key, group_df in grouped:
    first_row = group_df.iloc[0]
    row_dict = {col: first_row[col] for col in existing_group_cols}

    aff_idx = group_df["affaissement_mm"].dropna().first_valid_index()
    temp_idx = group_df["temp_beton_C"].dropna().first_valid_index()

    aff_val = (
        group_df.loc[aff_idx, "affaissement_mm"]
        if aff_idx is not None
        else None
    )
    temp_val = (
        group_df.loc[temp_idx, "temp_beton_C"] if temp_idx is not None else None
    )

    if (pd.isna(aff_val) or pd.isna(temp_val)) and not df_betonnage.empty:
      ref = str(row_dict.get("ref_controle", "")).strip()
      dt_str = str(row_dict.get("date_coulee", "")).strip()
      ovr = str(row_dict.get("ouvrage", "")).strip()

      matched_b = pd.DataFrame()
      if ref and "ref_controle_b" in df_betonnage.columns:
        matched_b = df_betonnage[
            df_betonnage["ref_controle_b"].astype(str).str.strip() == ref
        ]

      if (
          matched_b.empty
          and dt_str
          and ovr
          and "date_coulee_b" in df_betonnage.columns
          and "ouvrage_b" in df_betonnage.columns
      ):
        matched_b = df_betonnage[
            (df_betonnage["date_coulee_b"].astype(str).str.strip() == dt_str)
            & (df_betonnage["ouvrage_b"].astype(str).str.strip() == ovr)
        ]

      if not matched_b.empty:
        b_row = matched_b.iloc[0]
        if pd.isna(aff_val) and "affaissement_mm_b" in b_row:
          aff_val = b_row["affaissement_mm_b"]
        if pd.isna(temp_val) and "temp_beton_C_b" in b_row:
          temp_val = b_row["temp_beton_C_b"]

    row_dict["affaissement_mm"] = extract_numeric(aff_val)
    row_dict["temp_beton_C"] = extract_numeric(temp_val)

    valid_dates = group_df["date_dt"].dropna()
    row_dict["date_dt"] = valid_dates.min() if not valid_dates.empty else pd.NaT

    for ech in ["3 jours", "7 jours", "28 jours"]:
      vals = group_df[group_df["echeance_clean"] == ech]["fc_mpa"].dropna()
      if not vals.empty:
        if ech == "7 jours":
          # Calcul de la moyenne des résultats à 7 jours
          row_dict[f"fc_mpa_{ech}"] = round(vals.mean(), 1)
        else:
          if len(vals) == 1:
            row_dict[f"fc_mpa_{ech}"] = round(vals.iloc[0], 1)
          else:
            row_dict[f"fc_mpa_{ech}"] = " - ".join(
                [f"{round(v, 1):.1f}" for v in vals]
            )
      else:
        row_dict[f"fc_mpa_{ech}"] = None

    pivot_rows.append(row_dict)

  df_pivoted = pd.DataFrame(pivot_rows)

  desired_order = [
      "ref_controle",
      "date_coulee",
      "classe_beton",
      "ouvrage",
      "affaissement_mm",
      "temp_beton_C",
      "fc_mpa_3 jours",
      "fc_mpa_7 jours",
      "fc_mpa_28 jours",
      "date_dt",
  ]

  existing_cols = [c for c in desired_order if c in df_pivoted.columns]
  df_pivoted = df_pivoted[existing_cols]

  rename_map = {
      "ref_controle": "Réf. Contrôle",
      "date_coulee": "Date Coulée",
      "classe_beton": "Classe Béton",
      "ouvrage": "Ouvrage",
      "affaissement_mm": "Affaissement (mm)",
      "temp_beton_C": "Temp. Béton (°C)",
      "fc_mpa_3 jours": "Fc (MPa) [3 Jours]",
      "fc_mpa_7 jours": "Fc (MPa) [7 Jours]",
      "fc_mpa_28 jours": "Fc (MPa) [28 Jours]",
  }

  df_pivoted = df_pivoted.rename(columns=rename_map)
  return df_pivoted


def format_controle_dataframe(df_filtered):
  df_display = df_filtered.copy()
  if "date_dt" in df_display.columns:
    df_display = df_display.drop(columns=["date_dt"])
  return df_display


# =========================================================
# CALCULS STATISTIQUES PANDAS (ADAPTÉS AUX VALEURS INDIVIDUELLES)
# =========================================================


def compute_statistics_df(df_display):
  possible_cols = [
      ("Affaissement (mm)", ["Affaissement (mm)"]),
      ("Temp. Béton (°C)", ["Temp. Béton (°C)"]),
      ("Fc (MPa) [7 Jours]", ["Fc (MPa) [7 Jours]", "Moy. Fc (MPa) [7 Jours]"]),
      (
          "Fc (MPa) [28 Jours]",
          ["Fc (MPa) [28 Jours]", "Moy. Fc (MPa) [28 Jours]"],
      ),
  ]

  stats_data = {"Indicateur": ["MIN", "MOY", "MAX", "σ", "CV %"]}

  for label_col, candidates in possible_cols:
    actual_col = next((c for c in candidates if c in df_display.columns), None)

    if actual_col:
      raw_vals = df_display[actual_col].dropna()
      numeric_vals = []
      for val in raw_vals:
        if isinstance(val, (int, float)):
          if not np.isnan(val):
            numeric_vals.append(float(val))
        else:
          found = re.findall(r"[-+]?\d*\.\d+|\d+", str(val).replace(",", "."))
          for f in found:
            try:
              numeric_vals.append(float(f))
            except ValueError:
              pass

      s = pd.Series(numeric_vals)

      if not s.empty:
        v_min = round(s.min(), 1)
        v_moy = round(s.mean(), 1)
        v_max = round(s.max(), 1)

        if "28 Jours" in label_col:
          v_std = round(s.std(ddof=1), 2) if len(s) > 1 else 0.0
          v_cv = round((v_std / v_moy) * 100, 1) if v_moy > 0 else 0.0
          stats_data[label_col] = [v_min, v_moy, v_max, v_std, f"{v_cv} %"]
        else:
          stats_data[label_col] = [v_min, v_moy, v_max, "-", "-"]
      else:
        stats_data[label_col] = ["-", "-", "-", "-", "-"]
    else:
      stats_data[label_col] = ["-", "-", "-", "-", "-"]

  return pd.DataFrame(stats_data)


# =========================================================
# 3. STREAMLIT APP VUES
# =========================================================


def show(supabase):
  st.title("📊 Module de Synthèses du Béton")

  main_tab_betonnage, main_tab_controle = st.tabs([
      "🏗️ Synthèse de Suivi de Bétonnage",
      "🧪 Synthèse de Contrôle Béton",
  ])

  with main_tab_betonnage:
    st.subheader("Bilan du Suivi de Bétonnage")
    tab_j_b, tab_m_b = st.tabs(["📅 Bilan Journalier", "📅 Bilan Mensuel"])

    with tab_j_b:
      st.markdown("### Filtrage par jour, classe de béton et ouvrage")
      col1, col2, col3 = st.columns(3)
      with col1:
        selected_date = st.date_input(
            "Sélectionnez une date :", value=date.today(), key="b_date_j"
        )

      try:
        res = (
            supabase.table("suivi_betonnage")
            .select("*")
            .eq("date_livraison", str(selected_date))
            .execute()
        )
        data = res.data if res else []

        classes_j = ["Toutes"]
        ouvrages_j = ["Tous"]
        if data:
          df_temp = pd.DataFrame(data)
          if "classe_beton" in df_temp.columns:
            classes_j += sorted(
                list(df_temp["classe_beton"].dropna().unique())
            )
          if "ouvrage" in df_temp.columns:
            ouvrages_j += sorted(list(df_temp["ouvrage"].dropna().unique()))

        with col2:
          selected_class = st.selectbox(
              "Filtrer par classe de béton :", classes_j, key="b_class_j"
          )
        with col3:
          selected_ouvrage = st.selectbox(
              "Filtrer par ouvrage :", ouvrages_j, key="b_ouv_j"
          )

        if data:
          df = pd.DataFrame(data)
          if selected_class != "Toutes":
            df = df[df["classe_beton"] == selected_class]
          if selected_ouvrage != "Tous":
            df = df[df["ouvrage"] == selected_ouvrage]

          if df.empty:
            st.info("Aucun coulage enregistré pour les critères sélectionnés.")
          else:
            cols_drop = [
                c
                for c in [
                    "id",
                    "created_at",
                    "created",
                    "heure_fin_coulage",
                    "client",
                    "centrale_beton",
                    "technicien",
                    "observations",
                    "nb_eprouvettes",
                    "duree_transport",
                    "duree_transport_min",
                    "Durée de transport",
                ]
                if c in df.columns
            ]
            df = df.drop(columns=cols_drop)

            renames = {
                "date_livraison": "Date Livraison",
                "heure_arrivee": "Heure d'arrivée",
                "bl_num": "N° BL",
                "ouvrage": "Ouvrage",
                "quantite_m3": "Quantité (m³)",
                "classe_beton": "Classe",
                "temperature": "Temp. Béton (°C)",
                "temperature_ambiante": "Temp. Ambiante (°C)",
                "affaissement": "Affaissement (mm)",
                "prelevement": "Prélèvement",
                "meteo": "Météo",
            }
            df_display = df.rename(columns=renames)

            desired_col_order = [
                "Date Livraison",
                "Heure d'arrivée",
                "N° BL",
                "Ouvrage",
                "Quantité (m³)",
                "Classe",
                "Temp. Béton (°C)",
                "Temp. Ambiante (°C)",
                "Affaissement (mm)",
                "Prélèvement",
                "Météo",
            ]

            final_cols = [
                c for c in desired_col_order if c in df_display.columns
            ]
            remaining = [c for c in df_display.columns if c not in final_cols]
            df_display = df_display[final_cols + remaining]

            st.markdown("---")
            k1, k2 = st.columns(2)
            k1.metric(
                "Volume Total", f"{df_display['Quantité (m³)'].sum():.1f} m³"
            )
            if "Affaissement (mm)" in df_display.columns:
              k2.metric(
                  "Affaissement Moyen",
                  f"{pd.to_numeric(df_display['Affaissement (mm)'], errors='coerce').mean():.0f}"
                  " mm",
              )
            st.markdown("---")

            excel_file = generate_excel_synthesis_betonnage(
                df_display,
                f"Journée du {selected_date.strftime('%d/%m/%Y')}",
                is_mensuel=False,
            )
            st.download_button(
                label="📥 Télécharger la Synthèse Excel Bétonnage (A4 Portrait)",
                data=excel_file,
                file_name=f"Synthese_Betonnage_{selected_date}.xlsx",
                mime=(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
            )
            st.dataframe(df_display, use_container_width=True)
        else:
          st.info("Aucun coulage enregistré pour les critères sélectionnés.")
      except Exception as e:
        st.error(f"Erreur de chargement : {e}")

    with tab_m_b:
      st.markdown("### Bilan mensuel agrégé avec colonnes Min / Max séparées")
      col_m1, col_m2, col_m3 = st.columns(3)
      with col_m1:
        annee = date.today().year
        mois_liste = [
            "Janvier",
            "Février",
            "Mars",
            "Avril",
            "Mai",
            "Juin",
            "Juillet",
            "Août",
            "Septembre",
            "Octobre",
            "Novembre",
            "Décembre",
        ]
        mois_selected = st.selectbox(
            "Sélectionnez le mois :",
            mois_liste,
            index=date.today().month - 1,
            key="b_mois_m",
        )
        mois_num = mois_liste.index(mois_selected) + 1

      try:
        date_debut = f"{annee}-{mois_num:02d}-01"
        dernier_jour = (
            31
            if mois_num in [1, 3, 5, 7, 8, 10, 12]
            else (30 if mois_num in [4, 6, 9, 11] else 28)
        )
        date_fin = f"{annee}-{mois_num:02d}-{dernier_jour}"

        res_m = (
            supabase.table("suivi_betonnage")
            .select("*")
            .gte("date_livraison", date_debut)
            .lte("date_livraison", date_fin)
            .execute()
        )
        data_m = res_m.data if res_m else []

        classes_m = ["Toutes"]
        ouvrages_m = ["Tous"]
        if data_m:
          df_m_temp = pd.DataFrame(data_m)
          if "classe_beton" in df_m_temp.columns:
            classes_m += sorted(
                list(df_m_temp["classe_beton"].dropna().unique())
            )
          if "ouvrage" in df_m_temp.columns:
            ouvrages_m += sorted(list(df_m_temp["ouvrage"].dropna().unique()))

        with col_m2:
          selected_class_m = st.selectbox(
              "Filtrer par classe :", classes_m, key="b_class_m"
          )
        with col_m3:
          selected_ouvrage_m = st.selectbox(
              "Filtrer par ouvrage :", ouvrages_m, key="b_ouv_m"
          )

        if data_m:
          df_m = pd.DataFrame(data_m)
          if selected_class_m != "Toutes":
            df_m = df_m[df_m["classe_beton"] == selected_class_m]
          if selected_ouvrage_m != "Tous":
            df_m = df_m[df_m["ouvrage"] == selected_ouvrage_m]

          if df_m.empty:
            st.info("Aucun coulage enregistré pour ces critères.")
          else:
            df_m["quantite_m3"] = pd.to_numeric(
                df_m["quantite_m3"], errors="coerce"
            )
            df_m["temperature"] = pd.to_numeric(
                df_m["temperature"], errors="coerce"
            )
            df_m["temperature_ambiante"] = pd.to_numeric(
                df_m["temperature_ambiante"], errors="coerce"
            )
            df_m["affaissement"] = pd.to_numeric(
                df_m["affaissement"], errors="coerce"
            )
            df_m["date_dt"] = pd.to_datetime(
                df_m["date_livraison"], errors="coerce"
            )

            grouped_rows = []
            for (classe, ovr), group in df_m.groupby(
                ["classe_beton", "ouvrage"], dropna=False
            ):
              d_min = group["date_dt"].min()
              d_max = group["date_dt"].max()

              if pd.isna(d_min):
                date_str = "-"
              elif d_min == d_max:
                date_str = d_min.strftime("%d/%m/%Y")
              else:
                date_str = (
                    f"{d_min.strftime('%d/%m/%Y')} -"
                    f" {d_max.strftime('%d/%m/%Y')}"
                )

              vol_sum = group["quantite_m3"].sum()

              def get_min_max(series):
                s_valid = series.dropna()
                if s_valid.empty:
                  return "-", "-"
                return int(round(s_valid.min())), int(round(s_valid.max()))

              aff_min, aff_max = get_min_max(group["affaissement"])
              tb_min, tb_max = get_min_max(group["temperature"])
              ta_min, ta_max = get_min_max(group["temperature_ambiante"])

              grouped_rows.append({
                  ("Période", ""): date_str,
                  ("Classe", ""): classe,
                  ("Ouvrage", ""): ovr,
                  ("Quantité (m³)", ""): round(vol_sum, 1),
                  ("Affaissement", "Min"): aff_min,
                  ("Affaissement", "Max"): aff_max,
                  ("Temp. Béton (°C)", "Min"): tb_min,
                  ("Temp. Béton (°C)", "Max"): tb_max,
                  ("Temp. Ambiante (°C)", "Min"): ta_min,
                  ("Temp. Ambiante (°C)", "Max"): ta_max,
              })

            df_m_display = pd.DataFrame(grouped_rows)
            df_m_display.columns = pd.MultiIndex.from_tuples(
                df_m_display.columns
            )

            st.markdown("---")
            vol_total_m = df_m_display[("Quantité (m³)", "")].sum()
            st.metric("Volume Cumulé du Mois", f"{vol_total_m:.1f} m³")
            st.markdown("---")

            excel_file_m = generate_excel_synthesis_betonnage(
                df_m_display,
                f"Mois de {mois_selected} {annee}",
                is_mensuel=True,
            )
            st.download_button(
                label=(
                    "📥 Télécharger la Synthèse Mensuelle Excel Bétonnage (A4"
                    " Portrait)"
                ),
                data=excel_file_m,
                file_name=(
                    f"Synthese_Mensuelle_Betonnage_{mois_selected}_{annee}.xlsx"
                ),
                mime=(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
            )
            st.dataframe(df_m_display, use_container_width=True)
        else:
          st.info("Aucun coulage enregistré pour ce mois.")
      except Exception as e:
        st.error(f"Erreur de chargement : {e}")

  with main_tab_controle:
    st.subheader(
        "Bilan du Contrôle Béton (Valeurs Individuelles par Échéance)"
    )
    tab_j_c, tab_m_c = st.tabs(["📅 Bilan Journalier", "📅 Bilan Mensuel"])

    try:
      df_merged = load_and_process_controle_data(supabase)
    except Exception as e:
      st.error(f"Erreur lors de la préparation des données : {e}")
      df_merged = pd.DataFrame()

    classes_dispo = ["Toutes"]
    ouvrages_dispo = ["Tous"]
    if not df_merged.empty:
      if "Classe Béton" in df_merged.columns:
        classes_dispo += sorted(
            list(df_merged["Classe Béton"].dropna().unique())
        )
      if "Ouvrage" in df_merged.columns:
        ouvrages_dispo += sorted(list(df_merged["Ouvrage"].dropna().unique()))

    with tab_j_c:
      st.markdown("### Filtrage journalier par classe et ouvrage")
      col1, col2, col3 = st.columns(3)
      with col1:
        selected_date_c = st.date_input(
            "Sélectionnez une date :", value=date.today(), key="c_date_j"
        )
      with col2:
        selected_class_cj = st.selectbox(
            "Filtrer par classe :", classes_dispo, key="c_class_j"
        )
      with col3:
        selected_ouvrage_cj = st.selectbox(
            "Filtrer par ouvrage :", ouvrages_dispo, key="c_ouv_j"
        )

      if df_merged.empty:
        st.info("Aucune donnée disponible.")
      else:
        df_j_c = df_merged[df_merged["date_dt"].dt.date == selected_date_c]
        if (
            selected_class_cj != "Toutes"
            and "Classe Béton" in df_j_c.columns
        ):
          df_j_c = df_j_c[df_j_c["Classe Béton"] == selected_class_cj]
        if (
            selected_ouvrage_cj != "Tous"
            and "Ouvrage" in df_j_c.columns
        ):
          df_j_c = df_j_c[df_j_c["Ouvrage"] == selected_ouvrage_cj]

        if df_j_c.empty:
          st.info("Aucun contrôle enregistré pour les critères sélectionnés.")
        else:
          df_display_cj = format_controle_dataframe(df_j_c)

          st.markdown("---")
          st.metric("Nombre de Prélèvements", f"{len(df_display_cj)}")
          st.markdown("---")

          excel_file_cj = generate_excel_synthesis_controle(
              df_display_cj,
              f"Journée du {selected_date_c.strftime('%d/%m/%Y')}",
          )
          st.download_button(
              label="📥 Télécharger la Synthèse Contrôle Excel (A4 Portrait)",
              data=excel_file_cj,
              file_name=f"Synthese_Controle_Beton_{selected_date_c}.xlsx",
              mime=(
                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
              ),
          )
          st.dataframe(df_display_cj, use_container_width=True)

          st.markdown("### 📈 Synthèse Statistique")
          df_stats_j = compute_statistics_df(df_display_cj)
          st.dataframe(df_stats_j, use_container_width=True)

          st.markdown("---")
          afficher_evolution_resistances(df_j_c, key_suffix="journalier")

    with tab_m_c:
      st.markdown("### Bilan mensuel par classe et ouvrage")
      col_m1, col_m2, col_m3 = st.columns(3)
      with col_m1:
        annee_c = date.today().year
        mois_liste = [
            "Janvier",
            "Février",
            "Mars",
            "Avril",
            "Mai",
            "Juin",
            "Juillet",
            "Août",
            "Septembre",
            "Octobre",
            "Novembre",
            "Décembre",
        ]
        mois_selected_c = st.selectbox(
            "Sélectionnez le mois :",
            mois_liste,
            index=date.today().month - 1,
            key="c_mois_m",
        )
        mois_num_c = mois_liste.index(mois_selected_c) + 1
      with col_m2:
        selected_class_cm = st.selectbox(
            "Filtrer par classe :", classes_dispo, key="c_class_m"
        )
      with col_m3:
        selected_ouvrage_cm = st.selectbox(
            "Filtrer par ouvrage :", ouvrages_dispo, key="c_ouv_m"
        )

      if df_merged.empty:
        st.info("Aucune donnée disponible.")
      else:
        df_m_c = df_merged[
            (df_merged["date_dt"].dt.year == annee_c)
            & (df_merged["date_dt"].dt.month == mois_num_c)
        ]
        if (
            selected_class_cm != "Toutes"
            and "Classe Béton" in df_m_c.columns
        ):
          df_m_c = df_m_c[df_m_c["Classe Béton"] == selected_class_cm]
        if (
            selected_ouvrage_cm != "Tous"
            and "Ouvrage" in df_m_c.columns
        ):
          df_m_c = df_m_c[df_m_c["Ouvrage"] == selected_ouvrage_cm]

        if df_m_c.empty:
          st.info("Aucun contrôle enregistré pour ces critères.")
        else:
          df_display_cm = format_controle_dataframe(df_m_c)

          st.markdown("---")
          st.metric("Total Prélèvements du Mois", f"{len(df_display_cm)}")
          st.markdown("---")

          excel_file_cm = generate_excel_synthesis_controle(
              df_display_cm, f"Mois de {mois_selected_c} {annee_c}"
          )
          st.download_button(
              label=(
                  "📥 Télécharger la Synthèse Mensuelle Contrôle Excel (A4"
                  " Portrait)"
              ),
              data=excel_file_cm,
              file_name=(
                  f"Synthese_Mensuelle_Controle_{mois_selected_c}_{annee_c}.xlsx"
              ),
              mime=(
                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
              ),
          )
          st.dataframe(df_display_cm, use_container_width=True)

          st.markdown("### 📈 Synthèse Statistique Mensuelle")
          df_stats_m = compute_statistics_df(df_display_cm)
          st.dataframe(df_stats_m, use_container_width=True)

          st.markdown("---")
          afficher_evolution_resistances(df_m_c, key_suffix="mensuel")
