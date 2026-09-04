import base64
import io
import re
import unicodedata
from datetime import date, datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
import qrcode
from audit_log import enregistrer_modification, afficher_historique_modifications
import projets_config

try:
    from supabase import create_client, Client
except ImportError:
    Client = None

# Constante des onglets incluant la Phase 3
OPTIONS_ONGLETS = [
    "📋 Phase 0 : Réception & Validation",
    "📅 Phase 1 : Programmation",
    "💥 Phase 2 : Planning Daily & Saisie (Par Lot)",
    "🛡️ Phase 3 : Validation Admin (PVs)",
]


# ==============================================================================
# FONCTIONS UTILITAIRES : EXTRACTION DE DATES ET DÉLAIS (DAP-PANDAS & REGEX)
# ==============================================================================
def calculer_date_ecrasement(df):
    """
    Calcule la 'Date Écrasement Prévue' à partir de 'Date Coulée'
    et du délai spécifié dans 'Échéance Visée' ou 'echeance' (ex: '7 jours', '28 jours').
    """
    df_result = df.copy()

    # Détection de la colonne d'échéance disponible
    col_ech = next((c for c in ['Échéance Visée', 'echeance', 'Échéance'] if c in df_result.columns), None)
    col_coul = next((c for c in ['Date Coulée', 'date_coulee'] if c in df_result.columns), None)

    if not col_coul or not col_ech:
        return df_result

    # 1. Conversion de Date Coulée au format datetime
    df_result['Date Coulée'] = pd.to_datetime(df_result[col_coul], errors='coerce')

    # 2. Extraction du nombre de jours
    nb_jours = (
        df_result[col_ech]
        .astype(str)
        .str.extract(r'(\d+)')
        .fillna(28)[0]
        .astype(int)
    )

    # 3. Calcul de la Date Écrasement Prévue en ajoutant le nombre de jours
    df_result['Date Écrasement Prévue'] = df_result['Date Coulée'] + pd.to_timedelta(nb_jours, unit='D')

    # 4. Formate les dates au format YYYY-MM-DD
    df_result['Date Coulée'] = df_result['Date Coulée'].dt.strftime('%Y-%m-%d')
    df_result['Date Écrasement Prévue'] = df_result['Date Écrasement Prévue'].dt.strftime('%Y-%m-%d')

    return df_result


def extraire_nb_jours(echeance_str, default=28):
    """Extrait le nombre de jours numérique à partir d'une chaîne (ex: '28 jours', '7 J')."""
    if pd.isna(echeance_str) or not echeance_str:
        return default
    match = re.search(r'\d+', str(echeance_str))
    return int(match.group()) if match else default


def calculer_resistance(force_kn, type_essai="Compression", d=150.0, L=300.0, section=17671.46):
    """
    Calcule la résistance mécanique en MPa selon le type d'essai :
    - Compression : f_c = (F_kN * 1000) / section ≈ F_kN * 10 / 176.71
    - Traction par fendage : f_ct = (2 * F_kN * 1000) / (pi * L * d) ≈ F_kN / 70.6858
    """
    if not force_kn or force_kn <= 0:
        return 0.0

    if "fendage" in str(type_essai).lower() or "traction" in str(type_essai).lower():
        # f_ct (MPa) = (2 * F (N)) / (pi * L * d)
        return round(force_kn / 70.6858, 2)
    else:
        # f_c (MPa) = (F (N)) / A (mm²) = (F (kN) * 1000) / 17671.46
        return round((force_kn * 1000.0) / section, 1)


# ==============================================================================
# 1. GESTION DES UTILISATEURS ET CONNEXION SUPABASE
# ==============================================================================
def connecter_utilisateur(supabase, nom_utilisateur, mot_de_passe):
    """Vérifie l'utilisateur et initialise le session_state."""
    try:
        res = (
            supabase.table("users")
            .select("*")
            .eq("username", nom_utilisateur)
            .eq("password", mot_de_passe)
            .execute()
        )
        if res.data:
            u = res.data[0]
            st.session_state.update({
                "user_logged": True,
                "user": u,
                "username": u.get("username"),
                "role": u.get("role"),
                "user_role": u.get("role"),
                "can_edit": bool(u.get("can_edit", False)),
            })
            return True
        st.error("Nom d'utilisateur ou mot de passe incorrect.")
    except Exception as e:
        st.error(f"Erreur lors de la connexion : {e}")
    return False


def afficher_ecran_connexion(supabase):
    """Affiche le formulaire de connexion si l'utilisateur n'est pas authentifié."""
    st.title("🔐 Connexion au Laboratoire Smart Control Béton")
    
    if st.session_state.get("pending_qr_rec") or st.session_state.get("pending_qr_bid"):
        st.info("🎯 **Scan QR Code détecté !** Connectez-vous pour accéder directement à la Phase 2 de la fiche scannée.")

    with st.form("form_connexion", clear_on_submit=False):
        nom_u = st.text_input("Nom d'utilisateur", key="input_user")
        mdp = st.text_input("Mot de passe", type="password", key="input_pass")
        submit = st.form_submit_button("Se connecter", type="primary", use_container_width=True)

        if submit:
            if connecter_utilisateur(supabase, nom_u, mdp):
                st.toast("✅ Connexion réussie !", icon="🔓")
                st.rerun()


# =========================================================
# FONCTION UTILITAIRE : GÉNÉRATION ET GESTION DES QR CODES
# =========================================================
def generer_qr_code(data_url):
    """Génère un QR Code en mémoire sous forme d'image PNG en octets."""
    qr = qrcode.QRCode(box_size=8, border=2)
    qr.add_data(data_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    return buffer.getvalue()


# =========================================================
# FONCTION UTILITAIRE : VÉRIFICATION DES DOUBLONS DU N° DE RÉCEPTION
# =========================================================
def verifier_doublon_num_reception(supabase, num_reception, current_beton_id=None, projet_id=None):
    """Vérifie si le num_reception existe déjà dans suivi_betonnage, DANS LE
    MÊME PROJET uniquement."""
    if not num_reception or str(num_reception).strip() in ["", "-", "None", "NaN", "N/A"]:
        return False
    num_clean = str(num_reception).strip()
    try:
        query = (
            supabase.table("suivi_betonnage")
            .select("id, num_reception")
            .eq("num_reception", num_clean)
        )
        if projet_id:
            query = query.eq("projet_id", projet_id)
        res = query.execute()
        for m in (res.data or []):
            if current_beton_id is None or int(m.get("id")) != int(current_beton_id):
                return True
    except Exception as e:
        st.warning(f"Note lors de la vérification des doublons : {e}")
    return False


# =========================================================
# FONCTION UTILITAIRE : EXTRACTION SÉCURISÉE DU N° BL & DATE COULÉE
# =========================================================
def extraire_num_bl(*sources):
    """Inspecte récursivement les sources pour extraire le N° de Bon de Livraison (BL)."""
    keys = ["num_bl", "bl", "num_bon_livraison", "n_bl", "bon_livraison", "num_bl_p", "n_bon", "bon_de_livraison", "code_bl"]
    invalid = {"N/A", "NONE", "NAN", "-", ""}
    for src in sources:
        if isinstance(src, dict):
            for k in keys:
                val = str(src.get(k) or "").strip()
                if val and val.upper() not in invalid:
                    return val
            for k, v in src.items():
                if "bl" in k.lower() or "bon" in k.lower():
                    val = str(v or "").strip()
                    if val and val.upper() not in invalid:
                        return val
        elif isinstance(src, str):
            match = re.search(r"BL\s*:\s*([^\|]+)", src, re.IGNORECASE)
            if match:
                val = match.group(1).strip()
                if val and val.upper() not in invalid:
                    return val
    return "-"


def extraire_date_coulee(item):
    """Extrait la date de coulée / livraison exacte depuis l'objet bétonnage."""
    if not item or not isinstance(item, dict):
        return str(date.today())
    for k in ["date_coulee", "date_livraison", "date_prelevement"]:
        val = str(item.get(k) or "").strip()
        if val and val.upper() not in ["N/A", "NONE", "NAN", "-", ""]:
            return val[:10]
    return str(date.today())


# =========================================================
# 2. GÉNÉRATION DU PROCÈS-VERBAL EXCEL (FORMAT EXACT LPEE)
# =========================================================
def generer_pv_excel(export_data, infos_header):
    """Génère un Procès-Verbal (PV) d'écrasement de béton répliquant le modèle LPEE."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PV Écrasement LPEE"

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4, header=0.2, footer=0.2)

    font_bold = Font(name="Calibri", size=9, bold=True)
    font_bold_white = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    font_title_white = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=8.5)
    font_small = Font(name="Calibri", size=8)

    fill_dark = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_table = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_label = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    align_top_center = Alignment(horizontal="center", vertical="top", wrap_text=True)

    border_cell = Border(left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"),
                         top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"))

    default_bl = extraire_num_bl(infos_header)

    def remplacer_na(valeur, fallback=None):
        val_str = str(valeur).strip() if valeur is not None else ""
        return fallback if fallback is not None else default_bl if val_str.upper() in ["N/A", "NONE", "NAN", "", "-"] else valeur

    def format_cell(cell, font=font_regular, align=align_center, fill=None, border=border_cell):
        if font: cell.font = font
        if align: cell.alignment = align
        if fill: cell.fill = fill
        if border: cell.border = border

    # ENTÊTE
    ws.merge_cells("A1:D1")
    ws["A1"] = "LPEE / CTR CSB"
    format_cell(ws["A1"], font_bold_white, align_center)

    ws.merge_cells("A2:D3")
    ws["A2"] = "Laboratoire de Contrôle Externe"
    format_cell(ws["A2"], font_bold_white, align_center)

    for r in range(1, 4):
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = fill_dark

    ws["E1"] = "RE N° :"
    format_cell(ws["E1"], font_bold, align_left)
    ws.merge_cells("F1:G1")
    ws["F1"] = remplacer_na(infos_header.get("re_num"), "25/260/LGV/ B/")
    format_cell(ws["F1"], font_regular, align_left)
    
    ws["H1"] = "BETON"
    format_cell(ws["H1"], font_bold, align_center)

    ws.merge_cells("F2:H2")
    ws["F2"] = remplacer_na(infos_header.get("dossier"), "2025-260-05985-2025-0247")
    format_cell(ws["F2"], font_regular, align_left)

    ws["E3"] = "CLIENT :"
    format_cell(ws["E3"], font_bold, align_left)
    ws.merge_cells("F3:H3")
    ws["F3"] = remplacer_na(infos_header.get("client"), "TGCC")
    format_cell(ws["F3"], font_bold, align_left)

    for r in range(1, 4):
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = border_cell

    # TITRE & TYPE D'ESSAI
    ws.merge_cells("A4:H4")
    ws["A4"] = "ESSAIS MECANIQUES SUR BETON HYDRAULIQUE"
    format_cell(ws["A4"], font_title_white, align_center, fill_dark)
    for c in range(1, 9):
        ws.cell(row=4, column=c).fill = fill_dark
        ws.cell(row=4, column=c).border = border_cell

    # Vérification dynamique du type d'essai dans le lot
    type_essai_global = "Compression"
    for sample_i in export_data:
        t_e = str(sample_i.get("type_essai", "")).lower()
        if "fendage" in t_e or "traction" in t_e:
            type_essai_global = "Traction par fendage"
            break

    ws.merge_cells("A5:D5")
    ws["A5"] = f"[{'X' if type_essai_global == 'Compression' else ' '}] COMPRESSION NF EN 12390-3 (2019)"
    format_cell(ws["A5"], font_bold, align_center)

    ws.merge_cells("E5:H5")
    ws["E5"] = f"[{'X' if type_essai_global == 'Traction par fendage' else ' '}] TRACTION PAR FENDAGE NF EN 12390-6 (2019)"
    format_cell(ws["E5"], font_bold, align_center)

    for c in range(1, 9): ws.cell(row=5, column=c).border = border_cell

    ws.merge_cells("A6:F6")
    ws["A6"] = "Presse : Marque: Controls"
    format_cell(ws["A6"], font_bold, align_right)

    ws.merge_cells("G6:H6")
    ws["G6"] = "Classe : A"
    format_cell(ws["G6"], font_bold, align_center)

    for c in range(1, 9): ws.cell(row=6, column=c).border = border_cell

    # FICHE TECHNIQUE
    ws["A7"] = "Date de\nprélèvement"
    format_cell(ws["A7"], font_bold, align_center)
    ws["B7"] = str(remplacer_na(infos_header.get("date_coulee"), "-"))
    format_cell(ws["B7"], font_bold, align_center)

    ws.merge_cells("C7:D7")
    ws["C7"] = "Lieu de\nprélèvement"
    format_cell(ws["C7"], font_bold, align_center)

    ws.merge_cells("E7:H7")
    ws["E7"] = remplacer_na(infos_header.get("lieu_prelevement", infos_header.get("ouvrage")), "-")
    format_cell(ws["E7"], font_regular, align_center)

    ws["A8"] = "Chantier"
    format_cell(ws["A8"], font_bold, align_center)

    ws.merge_cells("B8:D8")
    ws["B8"] = remplacer_na(infos_header.get("chantier"), "Augmentation de la capacité ferroviaire entre Kenitra et Marrakech et au niveau du hub de Casablanca\nTravaux d'exécution de terrassement, ouvrages d'art et rétablissement de communication entre PK 5+450 et PK 10+000-GARE CASA SUD")
    format_cell(ws["B8"], font_small, align_center)

    ws.merge_cells("E8:F8")
    ws["E8"] = "Type de béton"
    format_cell(ws["E8"], font_bold, align_center)

    ws.merge_cells("G8:H8")
    classe_beton_val = str(remplacer_na(infos_header.get("classe_beton"), "C35/45")).upper()
    ws["G8"] = classe_beton_val
    format_cell(ws["G8"], font_bold, align_center)

    ws.merge_cells("A9:B9")
    ws["A9"] = remplacer_na(infos_header.get("centrale"), "Centrale à Béton")
    format_cell(ws["A9"], font_bold, align_center)

    ws["C9"] = "- Dimensions"
    format_cell(ws["C9"], font_regular, align_left)

    ws.merge_cells("D9:H9")
    ws["D9"] = remplacer_na(infos_header.get("forme"), "Cylindrique 150x300")
    format_cell(ws["D9"], font_bold, align_center)

    ws.merge_cells("A10:B10")
    ws["A10"] = "Affaissement au cône d'abrams NF EN 12350-2"
    format_cell(ws["A10"], font_small, align_center)

    ws["C10"] = str(remplacer_na(infos_header.get("affaissement"), "-"))
    format_cell(ws["C10"], font_bold, align_center)

    ws["D10"] = "- Mode confection"
    format_cell(ws["D10"], font_regular, align_left)

    ws.merge_cells("E10:H10")
    ws["E10"] = "Par vibration NF EN 12390-2 (2019)"
    format_cell(ws["E10"], font_bold, align_center)

    ws.merge_cells("A11:B11")
    ws["A11"] = "Température °C"
    format_cell(ws["A11"], font_regular, align_center)

    ws["C11"] = str(remplacer_na(infos_header.get("temperature"), "-"))
    format_cell(ws["C11"], font_bold, align_center)

    ws["D11"] = "- Mode conservation"
    format_cell(ws["D11"], font_regular, align_left)

    ws.merge_cells("E11:H11")
    ws["E11"] = "au laboratoire par immersion dans l'eau NF EN 12390-2 (2019) à 20°C ± 2°C"
    format_cell(ws["E11"], font_bold, align_center)

    tech_prelevement = remplacer_na(infos_header.get("technicien_prelevement") or infos_header.get("preleve_par") or infos_header.get("technicien"), "Technicien LPEE")
    ws.merge_cells("A12:C12")
    ws["A12"] = f"prélèvement effectué par {tech_prelevement}"
    format_cell(ws["A12"], font_small, align_center)

    ws.merge_cells("D12:E12")
    ws["D12"] = "N° de bon de livraison"
    format_cell(ws["D12"], font_regular, align_center)

    ws.merge_cells("F12:H12")
    ws["F12"] = default_bl
    format_cell(ws["F12"], font_bold, align_center)

    labels_coords = {"A7", "C7", "A8", "E8", "A9", "A10", "A11", "A12", "D12"}
    for r in range(7, 13):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.border = border_cell
            if cell.coordinate in labels_coords:
                cell.fill = fill_label

    # TABLEAU DES RÉSULTATS
    ws.merge_cells("A13:A14")
    ws["A13"] = "Réf,"
    ws.merge_cells("B13:C13")
    ws["B13"] = "Date"
    ws["B14"], ws["C14"] = "Fabri", "Essai"
    ws.merge_cells("D13:D14")
    ws["D13"] = "Age (jours)"
    ws.merge_cells("E13:E14")
    ws["E13"] = "Charge rupture(KN)"
    ws.merge_cells("F13:H13")
    ws["F13"] = "Résistance (MPa)"
    ws["F14"], ws["G14"], ws["H14"] = "Compression", "Traction", "Moyenne"

    for r in range(13, 15):
        for c in range(1, 9):
            format_cell(ws.cell(row=r, column=c), font=font_bold if r == 13 or c == 1 else font_regular, align=align_center, fill=fill_table)

    row_start = 15
    nb_total = len(export_data)
    groupes_lots = {}
    a_des_28j_ecrases, cellule_moyenne_28j = False, None

    for idx, item in enumerate(export_data):
        curr_row = row_start + idx
        try:
            f_kn_val = float(item.get("force_kn", 0.0))
        except (ValueError, TypeError):
            f_kn_val = 0.0
            
        is_en_cours = str(item.get("statut", "")).lower() == "en cours" or f_kn_val == 0.0
        dt_essai = item.get("date_essai")
        
        # Détection du type d'essai pour l'éprouvette
        type_essai = str(item.get("type_essai", type_essai_global)).lower()
        is_fendage = "fendage" in type_essai or "traction" in type_essai

        ws.cell(row=curr_row, column=1, value=str(item.get("repere_eprouvette", "B/01")))
        ws.cell(row=curr_row, column=2, value=str(remplacer_na(infos_header.get("date_coulee"), "-")))
        ws.cell(row=curr_row, column=3, value=str(dt_essai) if (is_en_cours and dt_essai and dt_essai != "-") else ("En cours" if is_en_cours else str(remplacer_na(dt_essai, "-"))))

        age_val = extraire_nb_jours(item.get("age"), default=7)
        ws.cell(row=curr_row, column=4, value=age_val)

        if is_en_cours:
            ws.cell(row=curr_row, column=5, value="En cours")
            ws.cell(row=curr_row, column=6, value="En cours")
            ws.cell(row=curr_row, column=7, value="-")
        else:
            sec_val = float(item.get("section", 17671.46))
            if sec_val < 1000:  # conversion si entrée en cm² ou mm² mal renseignée
                sec_val = 17671.46

            # Calcul de la résistance dynamique selon type d'essai
            res_val = item.get("fc_mpa")
            if not res_val or float(res_val) == 0.0:
                res_val = calculer_resistance(f_kn_val, type_essai=type_essai, section=sec_val)
            else:
                res_val = float(res_val)

            ws.cell(row=curr_row, column=5, value=f_kn_val).number_format = "0.0"

            # Alignement selon type d'essai : Colonne F (Compression) vs Colonne G (Traction)
            if is_fendage:
                ws.cell(row=curr_row, column=6, value="-")
                ws.cell(row=curr_row, column=7, value=res_val).number_format = "0.00"
            else:
                ws.cell(row=curr_row, column=6, value=res_val).number_format = "0.0"
                ws.cell(row=curr_row, column=7, value="-")

        for c in range(1, 9):
            format_cell(ws.cell(row=curr_row, column=c), font=font_regular, align=align_center)

        cle_lot = f"{item.get('age')}_{item.get('date_essai')}"
        if cle_lot not in groupes_lots:
            groupes_lots[cle_lot] = {"lignes": [], "en_cours": is_en_cours, "age": age_val, "is_fendage": is_fendage}
        elif is_en_cours:
            groupes_lots[cle_lot]["en_cours"] = True
        groupes_lots[cle_lot]["lignes"].append(curr_row)

    for cle_lot, data_lot in groupes_lots.items():
        lignes = data_lot["lignes"]
        start_r, end_r = min(lignes), max(lignes)
        cell_h = ws[f"H{start_r}"]
        col_source = "G" if data_lot["is_fendage"] else "F"

        if data_lot["en_cours"]:
            if start_r != end_r: ws.merge_cells(f"H{start_r}:H{end_r}")
            cell_h.value = "En cours"
        else:
            if start_r == end_r:
                cell_h.value = f"=ROUND({col_source}{start_r}, 1)"
            else:
                ws.merge_cells(f"H{start_r}:H{end_r}")
                cell_h.value = f"=ROUND(AVERAGE({col_source}{start_r}:{col_source}{end_r}), 1)"
            cell_h.number_format = "0.0"
            if extraire_nb_jours(data_lot["age"]) >= 28:
                a_des_28j_ecrases, cellule_moyenne_28j = True, f"H{start_r}"

        format_cell(cell_h, font=font_bold, align=align_center)

    next_row = row_start + nb_total
    ws.cell(row=next_row, column=1, value="Commentaire :")
    format_cell(ws.cell(row=next_row, column=1), font=font_bold, align=align_left, fill=fill_label)

    ws.merge_cells(f"B{next_row}:H{next_row}")
    obs_defaut = infos_header.get("observations") or "PERFORMANCES MECANIQUES A 28 JOURS SONT CONFORMES"

    seuil_min = 35.0
    for classe, seuil in [("C25/30", 25.0), ("C30/37", 30.0), ("C35/45", 35.0), ("C40/50", 40.0)]:
        if classe in classe_beton_val:
            seuil_min = seuil
            break

    if not a_des_28j_ecrases or not cellule_moyenne_28j:
        formule_commentaires = "PERFORMANCES MECANIQUES A 28 JOURS SERONT DONNES ULTERIEUREMENT."
    else:
        m_cell = cellule_moyenne_28j
        formule_commentaires = f'=IF(OR(ISBLANK({m_cell}), {m_cell}="En cours"), "PERFORMANCES MECANIQUES A 28 JOURS SERONT DONNES ULTERIEUREMENT.", IF({m_cell}>={seuil_min}, "{obs_defaut}", "PERFORMANCES MECANIQUES NON CONFORMES"))'

    format_cell(ws.cell(row=next_row, column=2, value=formule_commentaires), font=font_bold, align=align_left)
    for c in range(1, 9): ws.cell(row=next_row, column=c).border = border_cell

    r_sig_titre = next_row + 2
    r_sig_debut, r_sig_fin = r_sig_titre + 1, r_sig_titre + 4

    ws.merge_cells(start_row=r_sig_titre, start_column=2, end_row=r_sig_titre, end_column=4)
    format_cell(ws.cell(row=r_sig_titre, column=2, value="Visa Responsable d'essai"), font=font_bold, align=align_center)

    ws.merge_cells(start_row=r_sig_debut, start_column=2, end_row=r_sig_fin, end_column=4)
    format_cell(ws.cell(row=r_sig_debut, column=2, value="O.IKKEN"), font=font_bold, align=align_top_center)

    ws.merge_cells(start_row=r_sig_titre, start_column=6, end_row=r_sig_titre, end_column=8)
    format_cell(ws.cell(row=r_sig_titre, column=6, value="Visa Chef du laboratoire"), font=font_bold, align=align_center)

    ws.merge_cells(start_row=r_sig_debut, start_column=6, end_row=r_sig_fin, end_column=8)
    format_cell(ws.cell(row=r_sig_debut, column=6, value="H.BAALLAL"), font=font_bold, align=align_top_center)

    row_heights = {7: 32, 8: 48, 10: 23, 11: 23, 9: 15, 12: 15, 13: 15, 14: 15}
    for r in range(1, r_sig_fin + 1):
        if r in row_heights: ws.row_dimensions[r].height = row_heights[r]
        elif 15 <= r < (15 + nb_total) or r > 14: ws.row_dimensions[r].height = 28
        else: ws.row_dimensions[r].height = 16

    col_widths = {"A": 16, "B": 12, "C": 12, "D": 10, "E": 18, "F": 14, "G": 12, "H": 12}
    for col, width in col_widths.items(): ws.column_dimensions[col].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =========================================================
# FONCTIONS AUXILIAIRES DE SUPABASE & PARSING
# =========================================================
def exporter_dataframe_excel(df, date_chaine):
    """Génère un fichier Excel à partir d'un DataFrame."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=f"Planning_{date_chaine}"[:31])
    buffer.seek(0)
    return buffer


def obtenir_historique_betonnage(supabase, betonnage_id):
    if not betonnage_id: return []
    try:
        res = supabase.table("suivi_controle_beton").select("*").eq("betonnage_id", betonnage_id).order("id", desc=False).execute()
        return res.data or []
    except Exception as e:
        st.warning(f"Note : Historique du bétonnage #{betonnage_id} non disponible : {e}")
        return []


def obtenir_infos_betonnage_parent(supabase, betonnage_id):
    if not betonnage_id: return {}
    try:
        res = supabase.table("suivi_betonnage").select("*").eq("id", betonnage_id).execute()
        if res.data: return res.data[0]
    except Exception as e:
        st.warning(f"Note : Impossible de charger la fiche parent #{betonnage_id} : {e}")
        return {}


def obtenir_infos_betonnage_parents_bulk(supabase, betonnage_ids):
    """Charge en UNE seule requête les fiches parentes de plusieurs lots à la fois."""
    ids_valides = sorted({int(b) for b in betonnage_ids if b is not None})
    if not ids_valides:
        return {}
    try:
        res = supabase.table("suivi_betonnage").select("*").in_("id", ids_valides).execute()
        return {p["id"]: p for p in (res.data or [])}
    except Exception as e:
        st.warning(f"Note : Impossible de charger les fiches parentes en lot : {e}")
        return {}


def determiner_ref_controle(supabase, betonnage_id, info_betonnage, sample_ep):
    session_key = f"ref_controle_beton_{betonnage_id}"
    if st.session_state.get(session_key): return st.session_state[session_key]

    num_rec = (info_betonnage or {}).get("num_reception")
    if num_rec and str(num_rec).strip() not in ["", "-", "None", "NaN", "N/A"]:
        st.session_state[session_key] = str(num_rec).strip()
        return str(num_rec).strip()

    for candidate in [(info_betonnage or {}).get("ref_controle"), (sample_ep or {}).get("ref_controle")]:
        if candidate and str(candidate).strip():
            st.session_state[session_key] = str(candidate).strip()
            return str(candidate).strip()

    defaut = f"REF-{betonnage_id}-{(info_betonnage or {}).get('ouvrage', 'N/A')}"
    st.session_state[session_key] = defaut
    return defaut


def _format_ep_row(ep, date_ref=None):
    """Helper pour parser les données d'éprouvettes (Retard, Jour, Semaine)."""
    dt_coul, dt_ecras = ep.get("date_coulee"), ep.get("date_ecrasement")
    age_calc = "-"
    if dt_coul and dt_ecras:
        try:
            d_c, d_e = datetime.strptime(str(dt_coul)[:10], "%Y-%m-%d").date(), datetime.strptime(str(dt_ecras)[:10], "%Y-%m-%d").date()
            age_calc = f"{(d_e - d_c).days} jours" if not date_ref else f"{(date_ref - d_c).days} j (Aujourd'hui)"
        except Exception:
            age_calc = str(ep.get("echeance", "-"))

    ref_p = str(ep.get("ref_controle") or "").strip()
    rep_s = str(ep.get("repere_eprouvette", "")).strip()
    rep_complet = f"{ref_p}{rep_s}" if ref_p else rep_s
    try:
        f_kn = float(ep.get("force_kn") or 0.0)
    except (ValueError, TypeError):
        f_kn = 0.0

    return {
        "ID": ep.get("id"),
        "Référence / Repère": rep_complet,
        "N° BL": extraire_num_bl(ep),
        "Ouvrage": ep.get("ouvrage", "-"),
        "Classe Béton": ep.get("classe_beton", "-"),
        "Date Coulée": dt_coul,
        "Date Écrasement Prévue": dt_ecras,
        "Échéance Visée": ep.get("echeance", "-"),
        "Âge Théorique": age_calc,
        "Statut": "✅ Écrasée" if f_kn > 0 else "⏳ En attente"
    }


# =========================================================
# MODULE NOUVEAU : PHASE 3 - VALIDATION ADMIN (PVs)
# =========================================================
def afficher_module_validation_admin(supabase, est_admin=False):
    """Affiche le module d'approbation administrative et de signature des PVs."""
    st.subheader("🛡️ 3. Validation & Consultation des PVs")

    projet_id_actif = projets_config.projet_actif(st.session_state.get("user") or {})
    if not projet_id_actif:
        st.error("⚠️ Aucun projet ne vous est autorisé. Contactez un administrateur.")
        return

    if est_admin:
        st.info("💡 **Espace Administrateur BAALLAL** : vérifiez la conformité des écrasements et validez/signalez officiellement les PVs.")
    else:
        st.info("👁️ **Mode consultation** : cette phase est ouverte aux utilisateurs connectés. La validation officielle, le rejet, la signature et la modification des résultats sont réservés à l'administrateur BAALLAL.")

    try:
        res = supabase.table("suivi_controle_beton").select("*").eq("projet_id", projet_id_actif).not_.is_("force_kn", "null").gt("force_kn", 0).order("id", desc=True).execute()
        essais_realises = res.data or []
    except Exception as e:
        st.error(f"❌ Erreur de chargement des essais réalisés : {e}")
        return

    if not essais_realises:
        st.warning("ℹ️ Aucun essai écrasé n'est actuellement en attente de validation.")
        return

    # Regroupement par lot (betonnage_id)
    lots_dict = {}
    for ep in essais_realises:
        b_id = ep.get("betonnage_id")
        if b_id not in lots_dict:
            lots_dict[b_id] = []
        lots_dict[b_id].append(ep)

    def _est_pv_deja_valide(statut):
        if not statut:
            return False
        s = (
            unicodedata.normalize("NFKD", str(statut).strip().lower())
            .encode("ascii", "ignore")
            .decode("ascii")
        ).strip()
        if any(mot in s for mot in ["rejet", "invalide", "non valide"]):
            return False
        return "valide" in s

    options_valid = []
    parents_dict_admin = obtenir_infos_betonnage_parents_bulk(supabase, list(lots_dict.keys()))
    for b_id, list_ep in lots_dict.items():
        info_b = parents_dict_admin.get(b_id, {})
        statut_lot = info_b.get("statut_pv", "⏳ En attente de validation")
        if _est_pv_deja_valide(statut_lot):
            continue
        ref_ctrl = determiner_ref_controle(supabase, b_id, info_b, list_ep[0])
        bl_num = extraire_num_bl(list_ep[0], info_b or {})
        label = f"Réf: {ref_ctrl} | BL: {bl_num} | Ouvrage: {list_ep[0].get('ouvrage', '-')} | Statut: {statut_lot}"
        options_valid.append((label, b_id, list_ep, info_b))

    if not options_valid:
        st.success("✅ Tous les PV en attente ont été traités — aucun PV ne reste à valider pour le moment.")
        return

    choix_label, b_id_sel, ep_sel_list, info_b_sel = st.selectbox(
        "📦 Choisir le PV / Lot à réviser :",
        options=options_valid,
        format_func=lambda x: x[0],
        key="select_pv_admin"
    )

    st.markdown("---")
    c1, c2, c3 = st.columns(3)
    c1.metric("Nombre d'éprouvettes écrasées", len(ep_sel_list))
    c2.metric("Date de coulée", extraire_date_coulee(info_b_sel))
    c3.metric("Statut Actuel du PV", info_b_sel.get("statut_pv", "⏳ En attente"))

    # Préparation des données pour affichage / édition tableau
    df_key = f"admin_edit_pv_{b_id_sel}"
    if df_key not in st.session_state or st.session_state.get(f"{df_key}_len") != len(ep_sel_list):
        rows_val = []
        for ep in ep_sel_list:
            sec = float(ep.get("section") or 17671.46)
            f_kn = float(ep.get("force_kn") or 0.0)
            t_essai = ep.get("type_essai", "Compression")
            fc = float(ep.get("fc_mpa") or (calculer_resistance(f_kn, t_essai, section=sec) if f_kn > 0 else 0.0))
            rows_val.append({
                "ID": ep.get("id"),
                "Repère": ep.get("repere_eprouvette", "-"),
                "Type Essai": t_essai,
                "Échéance": ep.get("echeance", "-"),
                "Date Écrasement": ep.get("date_ecrasement", "-"),
                "Force (kN)": f_kn,
                "Résistance (MPa)": fc,
                "Opérateur": ep.get("technicien", "-"),
                "_section": sec,
                "_force_orig": f_kn,
                "_fc_orig": fc,
            })
        st.session_state[df_key] = pd.DataFrame(rows_val)
        st.session_state[f"{df_key}_len"] = len(ep_sel_list)

    editor_key = f"editor_{df_key}"

    def _maj_resistance_admin():
        """Recalcule Résistance (MPa) en direct selon le Type d'Essai."""
        editor_state = st.session_state.get(editor_key, {})
        for row_idx, updated_cols in editor_state.get("edited_rows", {}).items():
            if "Force (kN)" in updated_cols or "Type Essai" in updated_cols:
                try:
                    new_force = float(updated_cols.get("Force (kN)", st.session_state[df_key].at[row_idx, "Force (kN)"]))
                except (ValueError, TypeError):
                    new_force = 0.0
                
                t_essai = updated_cols.get("Type Essai", st.session_state[df_key].at[row_idx, "Type Essai"])
                sec = float(st.session_state[df_key].at[row_idx, "_section"])
                
                st.session_state[df_key].at[row_idx, "Force (kN)"] = new_force
                st.session_state[df_key].at[row_idx, "Type Essai"] = t_essai
                st.session_state[df_key].at[row_idx, "Résistance (MPa)"] = calculer_resistance(new_force, t_essai, section=sec)

    if est_admin:
        st.caption(
            "✏️ Mode administrateur : la **Force (kN)** et le **Type d'essai** sont modifiables"
            " ci-dessous — la Résistance (MPa) se recalcule automatiquement selon les formules."
        )

    st.data_editor(
        st.session_state[df_key],
        column_config={
            "ID": None,
            "_section": None,
            "_force_orig": None,
            "_fc_orig": None,
            "Repère": st.column_config.TextColumn("Repère", disabled=True),
            "Type Essai": st.column_config.SelectboxColumn("Type Essai", options=["Compression", "Traction par fendage"], disabled=not est_admin),
            "Échéance": st.column_config.TextColumn("Échéance", disabled=True),
            "Date Écrasement": st.column_config.TextColumn("Date Écrasement", disabled=True),
            "Force (kN)": st.column_config.NumberColumn(
                "⚡ Force (kN)", disabled=not est_admin,
                min_value=0.0, max_value=3000.0, step=0.1, format="%.1f",
            ),
            "Résistance (MPa)": st.column_config.NumberColumn("Résistance (MPa)", disabled=True, format="%.2f"),
            "Opérateur": st.column_config.TextColumn("Opérateur", disabled=True),
        },
        use_container_width=True, hide_index=True, key=editor_key, on_change=_maj_resistance_admin,
    )

    ep_28j = [ep for ep in ep_sel_list if extraire_nb_jours(ep.get("echeance")) == 28]
    if ep_28j:
        ep_7j = [ep for ep in ep_sel_list if extraire_nb_jours(ep.get("echeance")) == 7]
        ref_ctrl_sel = determiner_ref_controle(supabase, b_id_sel, info_b_sel, ep_sel_list[0])
        st.warning(
            f"🔔 **Rappel avant validation à 28 jours** — Résultats à 7 jours"
            f" déjà enregistrés pour ce même lot (Réf: {ref_ctrl_sel}) :"
        )
        if ep_7j:
            rows_7j = []
            for ep in ep_7j:
                sec = float(ep.get("section") or 17671.46)
                f_kn = float(ep.get("force_kn") or 0.0)
                t_e = ep.get("type_essai", "Compression")
                fc = float(ep.get("fc_mpa") or calculer_resistance(f_kn, t_e, section=sec))
                rows_7j.append({
                    "Repère": ep.get("repere_eprouvette", "-"),
                    "Type Essai": t_e,
                    "Date Écrasement": ep.get("date_ecrasement", "-"),
                    "Force (kN)": f_kn,
                    "Résistance (MPa)": fc,
                })
            st.dataframe(pd.DataFrame(rows_7j), use_container_width=True, hide_index=True)
        else:
            st.caption("Aucun écrasement à 7 jours n'a encore été enregistré pour ce lot.")

    st.markdown("---")
    if not est_admin:
        st.warning("🔐 **Validation officielle désactivée pour votre compte.** Seul l'administrateur **BAALLAL** peut enregistrer une décision, modifier les forces ou signer le PV.")
        st.markdown("### 📄 Statut du PV")
        st.write(info_b_sel.get("statut_pv", "⏳ En attente de validation"))
        if info_b_sel.get("visa_resp"):
            st.write(f"**Visa Responsable d'essai :** {info_b_sel.get('visa_resp')}")
        if info_b_sel.get("visa_chef"):
            st.write(f"**Visa Chef du laboratoire :** {info_b_sel.get('visa_chef')}")
        if info_b_sel.get("date_validation"):
            st.write(f"**Date de validation :** {info_b_sel.get('date_validation')}")
        if info_b_sel.get("observations_admin"):
            st.write(f"**Observations :** {info_b_sel.get('observations_admin')}")
    else:
        with st.form("form_valider_pv"):
            st.markdown("##### ✍️ Décision & Signatures Officielles")
            col_sig1, col_sig2 = st.columns(2)
            resp_essai = col_sig1.text_input("Visa Responsable d'essai", value=info_b_sel.get("visa_resp", "O.IKKEN"))
            chef_labo = col_sig2.text_input("Visa Chef du laboratoire", value=info_b_sel.get("visa_chef", "H.BAALLAL"))

            statut_decision = st.radio(
                "Décision d'approbation :",
                ["✅ Valider et Signer le PV", "⚠️ Remettre en Révision / Rejeter"],
                horizontal=True
            )
            comm_admin = st.text_area("Observations / Instructions complémentaires", value=info_b_sel.get("observations_admin", "Conforme aux spécifications NF EN 12390."))

            submit_val = st.form_submit_button("💾 Enregistrer la décision de validation", type="primary", use_container_width=True)

            if submit_val:
                nouveau_statut = "✅ Validé & Signé" if "Valider" in statut_decision else "❌ Rejeté / En Révision"
                anciennes_val_pv = {
                    "statut_pv": info_b_sel.get("statut_pv"),
                    "visa_resp": info_b_sel.get("visa_resp"),
                    "visa_chef": info_b_sel.get("visa_chef"),
                    "observations_admin": info_b_sel.get("observations_admin"),
                }
                update_payload = {
                    "statut_pv": nouveau_statut,
                    "visa_resp": resp_essai,
                    "visa_chef": chef_labo,
                    "observations_admin": comm_admin,
                    "date_validation": str(date.today())
                }
                try:
                    supabase.table("suivi_betonnage").update(update_payload).eq("id", b_id_sel).execute()
                    enregistrer_modification(
                        supabase,
                        table_concernee="suivi_betonnage",
                        enregistrement_id=b_id_sel,
                        action="VALIDATION",
                        anciennes_valeurs=anciennes_val_pv,
                        nouvelles_valeurs={
                            "statut_pv": nouveau_statut,
                            "visa_resp": resp_essai,
                            "visa_chef": chef_labo,
                            "observations_admin": comm_admin,
                        },
                    )

                    df_edit = st.session_state.get(df_key)
                    if df_edit is not None:
                        for _, r in df_edit.iterrows():
                            try:
                                ep_id_maj = int(r["ID"])
                                anciennes_force = {
                                    "force_kn": float(r.get("_force_orig", r["Force (kN)"])),
                                    "fc_mpa": float(r.get("_fc_orig", r["Résistance (MPa)"])),
                                }
                                nouvelles_force = {
                                    "type_essai": str(r.get("Type Essai", "Compression")),
                                    "force_kn": float(r["Force (kN)"]),
                                    "fc_mpa": float(r["Résistance (MPa)"]),
                                }
                                supabase.table("suivi_controle_beton").update(nouvelles_force).eq("id", ep_id_maj).execute()
                                enregistrer_modification(
                                    supabase,
                                    table_concernee="suivi_controle_beton",
                                    enregistrement_id=ep_id_maj,
                                    action="MODIFICATION",
                                    anciennes_valeurs=anciennes_force,
                                    nouvelles_valeurs=nouvelles_force,
                                    commentaire="Correction de force / type d'essai par l'administrateur lors de la validation du PV",
                                )
                            except Exception as e_row:
                                st.warning(f"⚠️ Éprouvette {r.get('Repère', '-')} : mise à jour de la force impossible ({e_row}).")

                    st.session_state.pop(df_key, None)
                    st.session_state.pop(f"{df_key}_len", None)
                    st.success(f"✅ Le statut du PV a été mis à jour avec succès : **{nouveau_statut}**")
                    st.balloons()
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Erreur lors de la mise à jour du statut : {e}")

    if est_admin:
        afficher_historique_modifications(supabase, "suivi_betonnage", b_id_sel)


# =========================================================
# 3. APPLICATION PRINCIPALE STREAMLIT
# =========================================================
def show(supabase):
    st.title("🧪 Contrôle & Écrasement du Béton (NF EN 12390)")

    st.session_state.setdefault("phase_actuelle", OPTIONS_ONGLETS[0])
    st.session_state.setdefault("nav_widget_seed", 0)

    qr_en_attente = bool(st.session_state.get("pending_qr_rec") or st.session_state.get("pending_qr_bid"))
    forcer_phase2_ce_run = qr_en_attente and not st.session_state.get("qr_nav_applied", False)
    if forcer_phase2_ce_run:
        st.session_state["phase_actuelle"] = OPTIONS_ONGLETS[2]
        st.session_state["nav_widget_seed"] += 1
        st.session_state["qr_nav_applied"] = True

    user_info = st.session_state.get("user", {})
    role_user = str(st.session_state.get("user_role") or st.session_state.get("role") or user_info.get("role", "")).lower()
    can_edit = st.session_state.get("can_edit", False) or bool(user_info.get("can_edit", False))

    current_username = str(
        st.session_state.get("username")
        or user_info.get("username")
        or ""
    ).strip().upper()
    is_baallal_admin = current_username == "BAALLAL" and (
        role_user == "admin" or st.session_state.get("is_admin", False)
    )

    projet_id_actif = projets_config.projet_actif(user_info)
    if not projet_id_actif:
        st.error("⚠️ Aucun projet ne vous est autorisé. Contactez un administrateur.")
        return
    st.caption(f"📁 Projet actif : **{projets_config.nom_projet(projet_id_actif)}**")

    mode_admin = False
    if is_baallal_admin:
        st.sidebar.markdown("---")
        st.sidebar.subheader("🔒 Mode Administration / Édition")
        mode_admin = st.sidebar.checkbox("Activer le Mode Admin / Édition", value=False)
        if mode_admin:
            st.sidebar.warning("⚠️ Mode Édition / Administrateur Actif.")

    # NAVIGATION INTERACTIVE DES PHASES
    widget_key = f"nav_phase_widget_{st.session_state['nav_widget_seed']}"
    try:
        onglet_courant = st.segmented_control(
            "Navigation entre phases :",
            OPTIONS_ONGLETS,
            default=st.session_state["phase_actuelle"],
            key=widget_key,
        )
    except AttributeError:
        onglet_courant = st.radio(
            "Navigation entre phases :",
            OPTIONS_ONGLETS,
            index=OPTIONS_ONGLETS.index(st.session_state["phase_actuelle"]),
            horizontal=True,
            key=widget_key,
        )

    if not onglet_courant:
        onglet_courant = st.session_state.get("phase_actuelle", OPTIONS_ONGLETS[0])

    if forcer_phase2_ce_run:
        onglet_courant = OPTIONS_ONGLETS[2]

    st.session_state["phase_actuelle"] = onglet_courant
    st.session_state["onglet_actif"] = onglet_courant

    betonnages_preleves = []
    try:
        res_b = supabase.table("suivi_betonnage").select("*").eq("projet_id", projet_id_actif).order("id", desc=True).execute()
        if res_b.data:
            betonnages_preleves = [b for b in res_b.data if b.get("prelevement") and str(b.get("prelevement")).upper().startswith("OUI")]
    except Exception as e:
        st.error(f"❌ Erreur lors du chargement des bétonnages : {e}")

    map_betonnages = {b.get("id"): b for b in betonnages_preleves}

    # =========================================================
    # PHASE 0 : RÉCEPTION & SAISIE DU NUMÉRO DE RÉCEPTION + QR CODES
    # =========================================================
    if onglet_courant == OPTIONS_ONGLETS[0]:
        st.subheader("📋 0. Réception & Validation des Bétons")
        st.info("💡 **Condition requise** : Saisissez manuellement le **N° Réception**. Une fois enregistré, le numéro débloquera la Phase 1 et permettra la génération d'étiquettes **QR Code** étanches.")

        if not betonnages_preleves:
            st.info("ℹ️ Aucun bétonnage prélevé dans la base de données.")
        else:
            rows_reception = []
            for item in betonnages_preleves:
                aff_val = item.get("affaissement") or item.get("slump") or "-"
                temp_val = item.get("temperature") or item.get("temp_beton") or "-"
                dt_livraison_exacte = extraire_date_coulee(item)
                rows_reception.append({
                    "_id_beton": item.get("id"),
                    "1-Numero de reception": str(item.get("num_reception") or ""),
                    "2-Date de livraison": dt_livraison_exacte,
                    "3-Nb d'éprouvettes": int(item.get("nb_eprouvettes") or 12),
                    "4-Classe de béton": item.get("classe_beton") or item.get("classe") or "-",
                    "5-Ouvrage": item.get("ouvrage") or "-",
                    "6-Affaissement": f"{aff_val} mm" if str(aff_val) != "-" and "mm" not in str(aff_val) else str(aff_val),
                    "7-Temperature de béton frais": f"{temp_val} °C" if str(temp_val) != "-" and "°C" not in str(temp_val) else str(temp_val),
                })

            df_edited = st.data_editor(
                pd.DataFrame(rows_reception),
                column_config={
                    "_id_beton": None,
                    "1-Numero de reception": st.column_config.TextColumn("1-N° Réception", help="Saisissez le N° de Réception ici", required=False),
                    "2-Date de livraison": st.column_config.TextColumn("2-Date de livraison", disabled=True),
                    "3-Nb d'éprouvettes": st.column_config.NumberColumn("3-Nb d'éprouvettes", disabled=True),
                    "4-Classe de béton": st.column_config.TextColumn("4-Classe de béton", disabled=True),
                    "5-Ouvrage": st.column_config.TextColumn("5-Ouvrage", disabled=True),
                    "6-Affaissement": st.column_config.TextColumn("6-Affaissement", disabled=True),
                    "7-Temperature de béton frais": st.column_config.TextColumn("7-Température béton", disabled=True),
                },
                use_container_width=True, hide_index=True, key="editor_reception_phase0",
            )

            col_rec1, col_rec2 = st.columns(2)
            with col_rec1:
                if st.button("💾 Enregistrer les N° de Réception", type="primary", use_container_width=True):
                    succes_rec, bloque_doublon, nouveaux_numeros = 0, False, {}
                    for _, row in df_edited.iterrows():
                        b_id, n_rec = int(row["_id_beton"]), str(row["1-Numero de reception"]).strip()
                        if n_rec and n_rec not in ["-", ""]:
                            if n_rec in nouveaux_numeros.values():
                                st.error(f"❌ **Saisie Bloquée** : Le N° de réception `{n_rec}` est saisi en double !")
                                bloque_doublon = True
                                break
                            if verifier_doublon_num_reception(supabase, n_rec, current_beton_id=b_id, projet_id=projet_id_actif):
                                st.error(f"❌ **Enregistrement Bloqué** : Le N° de réception `{n_rec}` existe déjà dans la base !")
                                bloque_doublon = True
                                break
                            nouveaux_numeros[b_id] = n_rec

                    if not bloque_doublon:
                        for b_id, n_rec in nouveaux_numeros.items():
                            try:
                                supabase.table("suivi_betonnage").update({"num_reception": n_rec}).eq("id", b_id).execute()
                                succes_rec += 1
                            except Exception:
                                try:
                                    supabase.table("suivi_betonnage").update({"n_reception": n_rec}).eq("id", b_id).execute()
                                    succes_rec += 1
                                except Exception as err:
                                    st.error(f"Erreur d'enregistrement pour #{b_id} : {err}")
                        if succes_rec > 0:
                            st.success(f"✅ {succes_rec} N° de Réception enregistré(s) !")
                            st.rerun()

            with col_rec2:
                df_visibles = df_edited[df_edited["1-Numero de reception"].str.strip().ne("") & df_edited["1-Numero de reception"].str.strip().ne("-")]
                st.download_button(
                    label="📊 Télécharger la liste des réceptions en Excel",
                    data=exporter_dataframe_excel(df_visibles.drop(columns=["_id_beton"]), "Phase_0"),
                    file_name=f"Reception_Beton_Phase_0_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="btn_download_reception_excel",
                )

            # --- MODULE ÉTIQUETTES QR CODE ---
            st.divider()
            st.subheader("📱 Étiquettes QR Code")
            
            receptions_validees = [b for b in betonnages_preleves if b.get("num_reception") and str(b.get("num_reception")).strip() not in ["", "-", "None"]]

            if not receptions_validees:
                st.warning("⚠️ Renseignez et enregistrez un **N° de Réception** dans le tableau ci-dessus pour générer les QR Codes.")
            else:
                with st.expander("🖨️ Afficher & Imprimer les QR Codes des Éprouvettes", expanded=True):
                    opt_qr = {f"Réception N° {b.get('num_reception')} | {b.get('ouvrage', '-')} ({extraire_date_coulee(b)})": b for b in receptions_validees}
                    choix_qr_lab = st.selectbox("Choisir la Réception à étiqueter :", list(opt_qr.keys()), key="select_qr_reception")
                    b_qr = opt_qr[choix_qr_lab]
                    rec_num = b_qr.get("num_reception")
                    nb_ep = int(b_qr.get("nb_eprouvettes") or 12)

                    st.markdown(f"**Génération pour `{rec_num}` ({nb_ep} éprouvettes) :**")

                    base_url = "https://smart-control-beton-lt7pusyvxjehm5kphd7hru.streamlit.app"

                    qr_items = []
                    cols_qr = st.columns(3)
                    for i in range(1, nb_ep + 1):
                        qr_payload = f"{base_url}/?rec={rec_num}&beton_id={b_qr.get('id')}&ep={i}"
                        qr_bytes = generer_qr_code(qr_payload)
                        qr_items.append((i, qr_bytes))
                        with cols_qr[(i - 1) % 3]:
                            st.caption(f"Éprouvette #{i} / {rec_num}")
                            st.image(qr_bytes, width=140)
                            st.download_button(
                                label=f"📥 QR Épr. #{i}",
                                data=qr_bytes,
                                file_name=f"QR_{str(rec_num).replace('/', '_')}_Ep{i}.png",
                                mime="image/png",
                                key=f"btn_qr_{b_qr.get('id')}_{i}",
                                use_container_width=True
                            )

                    etiquettes_html = "".join(
                        f"""
                        <div class="etiquette">
                          <img src="data:image/png;base64,{base64.b64encode(qr_bytes).decode()}" />
                          <div class="legende">Éprouvette #{num} / {rec_num}</div>
                        </div>
                        """
                        for num, qr_bytes in qr_items
                    )
                    page_impression = f"""
                    <html>
                    <head>
                    <style>
                      body {{ font-family: Arial, sans-serif; margin: 0; padding: 12px; }}
                      .barre-actions {{ margin-bottom: 14px; }}
                      .barre-actions button {{ background: #FF4B4B; color: white; border: none; padding: 10px 18px; border-radius: 6px; font-size: 15px; cursor: pointer; }}
                      .grille {{ display: flex; flex-wrap: wrap; gap: 18px; }}
                      .etiquette {{ text-align: center; width: 150px; break-inside: avoid; page-break-inside: avoid; }}
                      .etiquette img {{ width: 140px; height: 140px; }}
                      .legende {{ font-size: 12px; margin-top: 4px; }}
                      @media print {{ .barre-actions {{ display: none; }} }}
                    </style>
                    </head>
                    <body>
                      <div class="barre-actions">
                        <button onclick="window.print()">🖨️ Imprimer les étiquettes QR</button>
                      </div>
                      <div class="grille">{etiquettes_html}</div>
                    </body>
                    </html>
                    """
                    st.caption("Le bouton ci-dessous n'imprime que les étiquettes QR :")
                    components.html(page_impression, height=min(250 + 220 * ((nb_ep - 1) // 3 + 1), 900), scrolling=True)

    # =========================================================
    # PHASE 1 : PROGRAMMATION DES ÉCHÉANCES
    # =========================================================
    elif onglet_courant == OPTIONS_ONGLETS[1]:
        st.subheader("📅 1. Programmer les Échéances d'Écrasement")

        if can_edit:
            with st.expander("🔧 Corriger en masse les dates d'écrasement incohérentes", expanded=False):
                st.caption("Recalcule `Date Écrasement Prévue = Date Coulée + Échéance Visée` pour TOUTES les éprouvettes.")
                if st.button("🔧 Recalculer et corriger toutes les dates d'écrasement", key="btn_fix_toutes_dates_ecrasement"):
                    try:
                        res_fix = supabase.table("suivi_controle_beton").select("*").eq("projet_id", projet_id_actif).execute()
                        toutes_eprouvettes = res_fix.data or []
                    except Exception as e:
                        toutes_eprouvettes = []
                        st.error(f"Erreur lors du chargement : {e}")

                    nb_corrigees, nb_ignorees = 0, 0
                    for ep_fix in toutes_eprouvettes:
                        dt_coulee_fix = str(ep_fix.get("date_coulee") or "").strip()
                        echeance_fix = str(ep_fix.get("echeance") or "").strip()
                        date_actuelle_fix = str(ep_fix.get("date_ecrasement") or "").strip()
                        nb_j_fix = extraire_nb_jours(echeance_fix, default=None)
                        if nb_j_fix is None or not dt_coulee_fix:
                            nb_ignorees += 1
                            continue
                        try:
                            dt_c_fix = datetime.strptime(dt_coulee_fix[:10], "%Y-%m-%d").date()
                            date_correcte_fix = str(dt_c_fix + timedelta(days=nb_j_fix))
                        except (ValueError, TypeError):
                            nb_ignorees += 1
                            continue

                        if date_correcte_fix != date_actuelle_fix[:10]:
                            try:
                                supabase.table("suivi_controle_beton").update(
                                    {"date_ecrasement": date_correcte_fix}
                                ).eq("id", ep_fix["id"]).execute()
                                nb_corrigees += 1
                            except Exception as err_fix:
                                st.error(f"Erreur pour #{ep_fix.get('id')} : {err_fix}")

                    if nb_corrigees > 0:
                        st.success(f"✅ {nb_corrigees} date(s) d'écrasement corrigée(s) !")
                        st.rerun()
                    else:
                        st.info("👍 Toutes les dates d'écrasement étaient déjà cohérentes.")

            with st.expander("✏️ Modification / Ajustement d'une Programmation Existante", expanded=False):
                try:
                    res_p = supabase.table("suivi_controle_beton").select("*").eq("projet_id", projet_id_actif).order("id", desc=True).execute()
                    eprouvettes_enregistrees = res_p.data or []
                except Exception as e:
                    eprouvettes_enregistrees = []
                    st.error(f"Erreur lors du chargement des programmations : {e}")

                if eprouvettes_enregistrees:
                    dates_corrigees_count = 0
                    for ep in eprouvettes_enregistrees:
                        parent_beton = map_betonnages.get(ep.get("betonnage_id"))
                        if parent_beton:
                            date_coulee_correcte = extraire_date_coulee(parent_beton)
                            dt_coulee_ep = str(ep.get("date_coulee") or "").strip()
                            
                            if dt_coulee_ep != date_coulee_correcte:
                                ep["date_coulee"] = date_coulee_correcte
                                nb_j = extraire_nb_jours(ep.get("echeance"), default=28)
                                
                                try:
                                    dt_c = datetime.strptime(date_coulee_correcte[:10], "%Y-%m-%d").date()
                                    nouvelle_date_ecrasement = str(dt_c + timedelta(days=nb_j))
                                    ep["date_ecrasement"] = nouvelle_date_ecrasement
                                except Exception:
                                    nouvelle_date_ecrasement = ep.get("date_ecrasement")

                                try:
                                    supabase.table("suivi_controle_beton").update({
                                        "date_coulee": date_coulee_correcte,
                                        "date_ecrasement": nouvelle_date_ecrasement
                                    }).eq("id", ep["id"]).execute()
                                    dates_corrigees_count += 1
                                except Exception:
                                    pass

                    if dates_corrigees_count > 0:
                        st.success(f"🔄 **Synchronisation effectuée** : {dates_corrigees_count} date(s) de coulée réalignée(s) !")

                    df_edit_prog = pd.DataFrame(eprouvettes_enregistrees)
                    cols_ed = [c for c in ["id", "betonnage_id", "type_essai", "ref_controle", "repere_eprouvette", "echeance", "date_ecrasement", "date_coulee", "ouvrage", "classe_beton"] if c in df_edit_prog.columns]
                    df_display_prog = df_edit_prog[cols_ed].copy()

                    etat_editeur_prog = st.session_state.get("editor_modification_phase1", {})
                    edited_rows_prog = etat_editeur_prog.get("edited_rows", {})
                    lignes_avec_date_manuelle = set()
                    for idx_pos, changements in edited_rows_prog.items():
                        idx_pos = int(idx_pos)
                        if idx_pos < len(df_display_prog):
                            for col_maj, val_maj in changements.items():
                                if col_maj in df_display_prog.columns:
                                    df_display_prog.iat[idx_pos, df_display_prog.columns.get_loc(col_maj)] = val_maj
                            if "date_ecrasement" in changements:
                                lignes_avec_date_manuelle.add(idx_pos)

                    if "date_ecrasement" in df_display_prog.columns and "date_coulee" in df_display_prog.columns:
                        col_idx_ecras = df_display_prog.columns.get_loc("date_ecrasement")
                        for idx_pos in range(len(df_display_prog)):
                            if idx_pos in lignes_avec_date_manuelle:
                                continue
                            ech_val = df_display_prog.iloc[idx_pos].get("echeance")
                            coulee_val = df_display_prog.iloc[idx_pos].get("date_coulee")
                            nb_j_apercu = extraire_nb_jours(ech_val, default=28)
                            try:
                                dt_c_apercu = datetime.strptime(str(coulee_val)[:10], "%Y-%m-%d").date()
                                df_display_prog.iat[idx_pos, col_idx_ecras] = str(dt_c_apercu + timedelta(days=nb_j_apercu))
                            except (ValueError, TypeError):
                                pass

                    df_prog_modifiee = st.data_editor(
                        df_display_prog,
                        column_config={
                            "id": st.column_config.NumberColumn("ID", disabled=True),
                            "betonnage_id": None,
                            "type_essai": st.column_config.SelectboxColumn("Type d'essai", options=["Compression", "Traction par fendage"]),
                            "ref_controle": st.column_config.TextColumn("Réf. Contrôle (N° Réception)"),
                            "echeance": st.column_config.SelectboxColumn("Échéance Visée", options=["3 jours", "7 jours", "28 jours", "90 jours"]),
                            "date_coulee": st.column_config.TextColumn("Date Coulée"),
                            "date_ecrasement": st.column_config.TextColumn("Date Écrasement Prévue"),
                            "ouvrage": st.column_config.TextColumn("Ouvrage", disabled=True),
                            "classe_beton": st.column_config.TextColumn("Classe Béton", disabled=True),
                        },
                        use_container_width=True, hide_index=True, key="editor_modification_phase1",
                    )

                    if st.button("💾 Enregistrer les Modifications de Programmation", type="primary", use_container_width=True, key="btn_save_mod_prog"):
                        bloque_mod = False
                        for _, r_m in df_prog_modifiee.iterrows():
                            ref_ctrl = str(r_m.get("ref_controle", "")).strip()
                            if ref_ctrl and verifier_doublon_num_reception(supabase, ref_ctrl, current_beton_id=r_m.get("betonnage_id"), projet_id=projet_id_actif):
                                st.error(f"❌ **Modification Bloquée** : La Réf `{ref_ctrl}` existe déjà !")
                                bloque_mod = True
                                break

                        if not bloque_mod:
                            orig_par_id_p1 = {ep["id"]: ep for ep in eprouvettes_enregistrees}
                            nb_succes = 0
                            for _, r_m in df_prog_modifiee.iterrows():
                                ep_id, b_id = int(r_m["id"]), r_m.get("betonnage_id")
                                ref_ctrl = str(r_m.get("ref_controle", "")).strip()
                                ech_str = str(r_m.get("echeance", "")).strip()
                                dt_coulee_str = str(r_m.get("date_coulee", "")).strip()
                                type_essai_m = str(r_m.get("type_essai", "Compression")).strip()

                                dt_ecrasement_val = str(r_m.get("date_ecrasement", "")).strip()
                                if not dt_ecrasement_val or dt_ecrasement_val.lower() in ["none", "nan", "-", ""]:
                                    nb_j = extraire_nb_jours(ech_str, default=28)
                                    try:
                                        dt_c = datetime.strptime(dt_coulee_str[:10], "%Y-%m-%d").date()
                                        dt_ecrasement_val = str(dt_c + timedelta(days=nb_j))
                                    except (ValueError, TypeError):
                                        dt_ecrasement_val = dt_coulee_str

                                pay = {
                                    "ref_controle": ref_ctrl,
                                    "type_essai": type_essai_m,
                                    "repere_eprouvette": str(r_m.get("repere_eprouvette", "")).strip(),
                                    "echeance": ech_str,
                                    "date_coulee": dt_coulee_str,
                                    "date_ecrasement": dt_ecrasement_val,
                                }
                                try:
                                    orig_row_p1 = orig_par_id_p1.get(ep_id, {})
                                    supabase.table("suivi_controle_beton").update(pay).eq("id", ep_id).execute()
                                    enregistrer_modification(
                                        supabase,
                                        table_concernee="suivi_controle_beton",
                                        enregistrement_id=ep_id,
                                        action="MODIFICATION",
                                        anciennes_valeurs={k: orig_row_p1.get(k) for k in pay},
                                        nouvelles_valeurs=pay,
                                        commentaire="Ajustement de programmation (Phase 1)",
                                    )
                                    if b_id:
                                        try: supabase.table("suivi_betonnage").update({"num_reception": ref_ctrl}).eq("id", b_id).execute()
                                        except Exception: pass
                                    nb_succes += 1
                                except Exception as err:
                                    st.error(f"Erreur pour #{ep_id} : {err}")
                            if nb_succes > 0:
                                st.success(f"✅ {nb_succes} programmation(s) mise(s) à jour !")
                                st.rerun()

                    st.markdown("---")
                    st.markdown("##### 🗑️ Supprimer une ou plusieurs éprouvettes programmées par erreur")
                    
                    lots_pour_suppression = {}
                    for ep in eprouvettes_enregistrees:
                        cle_lot_suppr = f"{ep.get('ref_controle', '-')} — {ep.get('ouvrage', '-')} (Lot #{ep.get('betonnage_id')})"
                        lots_pour_suppression.setdefault(cle_lot_suppr, []).append(ep)

                    labels_lots_suppr = [f"{cle} — {len(eps)} éprouvette(s)" for cle, eps in lots_pour_suppression.items()]
                    mapping_label_vers_cle_lot = dict(zip(labels_lots_suppr, lots_pour_suppression.keys()))

                    lot_choisi_suppr = st.selectbox("1️⃣ Choisir le lot", options=labels_lots_suppr, key="select_lot_suppr_prog")
                    eprouvettes_du_lot_suppr = lots_pour_suppression.get(mapping_label_vers_cle_lot.get(lot_choisi_suppr), [])

                    options_suppr = {
                        f"#{ep['id']} — {ep.get('repere_eprouvette', '-')} ({ep.get('echeance', '-')}, prévu le {ep.get('date_ecrasement', '-')})": ep["id"]
                        for ep in eprouvettes_du_lot_suppr
                    }
                    choix_suppr = st.multiselect("2️⃣ Choisir la ou les éprouvette(s) à supprimer dans ce lot", options=list(options_suppr.keys()), key="multiselect_suppr_prog")
                    if choix_suppr:
                        st.warning(f"⚠️ {len(choix_suppr)} éprouvette(s) sélectionnée(s) pour suppression définitive.")
                        confirmer_suppr = st.checkbox("Je confirme vouloir supprimer définitivement ces éprouvettes", key="confirm_suppr_prog")
                        if st.button("🗑️ Supprimer les éprouvettes sélectionnées", type="primary", disabled=not confirmer_suppr, key="btn_suppr_prog"):
                            orig_par_id_suppr = {ep["id"]: ep for ep in eprouvettes_enregistrees}
                            nb_suppr = 0
                            for label in choix_suppr:
                                ep_id_suppr = options_suppr[label]
                                try:
                                    ep_avant_suppr = orig_par_id_suppr.get(ep_id_suppr, {})
                                    supabase.table("suivi_controle_beton").delete().eq("id", ep_id_suppr).execute()
                                    enregistrer_modification(
                                        supabase,
                                        table_concernee="suivi_controle_beton",
                                        enregistrement_id=ep_id_suppr,
                                        action="SUPPRESSION",
                                        anciennes_valeurs={k: v for k, v in ep_avant_suppr.items() if k != "id"},
                                        commentaire="Suppression d'une éprouvette programmée par erreur (Phase 1)",
                                    )
                                    nb_suppr += 1
                                except Exception as err_suppr:
                                    st.error(f"Erreur lors de la suppression de {label} : {err_suppr}")
                            if nb_suppr > 0:
                                st.success(f"✅ {nb_suppr} éprouvette(s) supprimée(s).")
                                st.rerun()
        else:
            st.info("🔒 **Accès restreint :** La modification nécessite le droit `can_edit`.")

        st.divider()
        st.subheader("➕ Ajouter une Nouvelle Programmation")

        prog_counts = {}
        try:
            res_deja = supabase.table("suivi_controle_beton").select("betonnage_id").eq("projet_id", projet_id_actif).execute()
            for r in (res_deja.data or []):
                b_id_v = r.get("betonnage_id")
                if b_id_v: prog_counts[b_id_v] = prog_counts.get(b_id_v, 0) + 1
        except Exception as e:
            st.warning(f"Note lors du contrôle des quotas : {e}")

        betonnages_non_programmes, fiches_sans_num_reception = [], []
        for b in betonnages_preleves:
            num_rec = str(b.get("num_reception") or b.get("n_reception") or "").strip()
            if not num_rec or num_rec.upper() in ["-", "NONE", "NAN", "N/A", "NULL"]:
                fiches_sans_num_reception.append(b)
                continue
            total_prevu = int(b.get("nb_eprouvettes") or 12)
            if (total_prevu - prog_counts.get(b.get("id"), 0)) > 0 or mode_admin:
                betonnages_non_programmes.append(b)

        if fiches_sans_num_reception:
            st.warning(f"⚠️ **{len(fiches_sans_num_reception)} fiche(s)** manquent de **N° de Réception**. Allez dans l'onglet Phase 0 pour le saisir.")

        if betonnages_non_programmes:
            options_beton = {
                f"N° Réception: {b.get('num_reception') or b.get('n_reception')} | Date Coulée: {extraire_date_coulee(b)} | Classe: {b.get('classe_beton', 'N/A')} | Ouvrage: {b.get('ouvrage', 'N/A')} | BL: {extraire_num_bl(b)}": b
                for b in betonnages_non_programmes
            }
            choix_label_p = st.selectbox("Sélectionner la fiche de bétonnage :", list(options_beton.keys()), key="prog_beton_select")
            beton_p = options_beton[choix_label_p]
            b_id = beton_p.get("id")

            num_reception_p = beton_p.get("num_reception") or beton_p.get("n_reception")
            num_bl_p = extraire_num_bl(beton_p, choix_label_p)
            ouvrage_p, classe_beton_p = str(beton_p.get("ouvrage") or "-"), str(beton_p.get("classe_beton") or "-")
            total_eprouvettes_prevues = int(beton_p.get("nb_eprouvettes") or 12)
            eprouvettes_deja_prog = prog_counts.get(b_id, 0)
            solde_disponible = max(0, total_eprouvettes_prevues - eprouvettes_deja_prog)

            date_coulee_raw = extraire_date_coulee(beton_p)
            try: date_coulee_p = datetime.strptime(date_coulee_raw, "%Y-%m-%d").date()
            except Exception: date_coulee_p = date.today()

            st.markdown("---")
            st.info(f"📌 **N° Réception : {num_reception_p}** | Total prévu : **{total_eprouvettes_prevues}** | Déjà programmée(s) : **{eprouvettes_deja_prog}** | Reste : **{solde_disponible}**")

            ref_value = num_reception_p if num_reception_p and str(num_reception_p).strip() not in ["", "-", "None", "NaN", "N/A"] else determiner_ref_controle(supabase, b_id, beton_p, {})
            ref_controle_p = st.text_input("🏷️ Référence de Contrôle", value=ref_value, disabled=True, key=f"p_ref_ctrl_{b_id}")
            st.session_state[f"ref_controle_beton_{b_id}"] = ref_controle_p

            st.markdown("---")
            type_essai_sel = st.selectbox("🔬 Type d'Essai", ["Compression", "Traction par fendage"], key=f"p_type_essai_{b_id}")

            col1, col2, col3 = st.columns(3)
            col1.text_input("N° Bon de Livraison (BL)", value=num_bl_p, disabled=True, key=f"p_bl_{b_id}")
            col2.text_input("Ouvrage / Élément", value=ouvrage_p, disabled=True, key=f"p_ouv_{b_id}")
            col3.text_input("Classe de Béton Spécifiée", value=classe_beton_p, disabled=True, key=f"p_classe_{b_id}")

            options_echeances = ["3 jours", "7 jours", "28 jours", "90 jours"]
            echeance_p = st.selectbox("Âge / Échéance visée", options_echeances, key=f"p_echeance_{b_id}")
            
            df_calcul_single = pd.DataFrame([{
                'Date Coulée': str(date_coulee_p),
                'Échéance Visée': echeance_p
            }])
            df_calcul_single_res = calculer_date_ecrasement(df_calcul_single)
            date_ecrasement_calculee = datetime.strptime(df_calcul_single_res.at[0, 'Date Écrasement Prévue'], "%Y-%m-%d").date()

            nb_j = extraire_nb_jours(echeance_p, default=28)

            col_e1, col_e2, col_e3 = st.columns(3)
            col_e1.date_input("Date de Coulée (Phase 0)", value=date_coulee_p, disabled=True, key=f"p_date_coul_{b_id}")
            date_ecrasement_prevue = col_e2.date_input("Date d'Écrasement Prévue", value=date_ecrasement_calculee, key=f"p_date_ecras_{b_id}_{nb_j}j")
            
            max_allowed = solde_disponible if not mode_admin else 50
            nb_eprouvettes_p = col_e3.number_input("Nombre d'éprouvettes", min_value=(1 if max_allowed > 0 else 0), max_value=max_allowed, value=min(3, max_allowed) if max_allowed >= 3 else max_allowed, key=f"p_nb_ep_{b_id}_{nb_j}j")

            forme_p = st.selectbox("Type / Forme d'éprouvette", ["Cylindrique 150x300", "Cylindrique 160x320", "Cylindrique 100x200"], key=f"p_forme_{b_id}")
            
            # Dimensions standard 150x300 -> d=150mm, L=300mm -> Section 17671.46 mm²
            sect_def = 17671.46 if "150x300" in forme_p else (20106.19 if "160x320" in forme_p else 7853.98)

            if int(nb_eprouvettes_p) > 0:
                st.markdown("##### 🏷️ Repères codés des éprouvettes")
                reperes_p = []
                cols_rep = st.columns(min(int(nb_eprouvettes_p), 6))
                for i in range(int(nb_eprouvettes_p)):
                    with cols_rep[i % 6]:
                        rep_val = st.text_input(f"Repère #{eprouvettes_deja_prog + i + 1}", value=f"/{eprouvettes_deja_prog + i + 1}", key=f"prog_rep_{b_id}_{nb_j}j_{i}")
                        reperes_p.append(rep_val)

                if st.button("📌 Enregistrer la Programmation", type="primary", use_container_width=True, key=f"btn_save_prog_{b_id}_{nb_j}j"):
                    try: supabase.table("suivi_betonnage").update({"ref_controle": ref_controle_p}).eq("id", b_id).execute()
                    except Exception: pass

                    succes_cnt = 0
                    for rep in reperes_p:
                        pay = {
                            "betonnage_id": b_id, "type_essai": type_essai_sel, "num_bl": num_bl_p, "ouvrage": ouvrage_p, "classe_beton": classe_beton_p,
                            "date_coulee": str(date_coulee_p), "echeance": echeance_p, "date_ecrasement": str(date_ecrasement_prevue),
                            "ref_controle": ref_controle_p, "repere_eprouvette": rep, "forme": forme_p, "section": float(sect_def),
                            "projet_id": projet_id_actif,
                        }
                        try:
                            res_ins_prog = supabase.table("suivi_controle_beton").insert(pay).execute()
                            if res_ins_prog.data:
                                succes_cnt += 1
                                nouvel_id_prog = res_ins_prog.data[0].get("id")
                                enregistrer_modification(
                                    supabase,
                                    table_concernee="suivi_controle_beton",
                                    enregistrement_id=nouvel_id_prog,
                                    action="CREATION",
                                    nouvelles_valeurs=pay,
                                    commentaire="Programmation d'une nouvelle éprouvette",
                                )
                        except Exception as err: st.error(f"Erreur pour {rep} : {err}")

                    if succes_cnt > 0:
                        st.success(f"✅ {succes_cnt} éprouvette(s) programmée(s) pour le {date_ecrasement_prevue} ({echeance_p}) en mode {type_essai_sel} !")
                        st.rerun()

    # =========================================================
    # PHASE 2 : PLANNING & SAISIE DES ÉCRASEMENTS
    # =========================================================
    elif onglet_courant == OPTIONS_ONGLETS[2]:
        st.subheader("💥 2. Planning des Échéances & Saisie des Écrasements")

        scan_rec = st.session_state.get("pending_qr_rec")
        scan_b_id = st.session_state.get("pending_qr_bid")
        if scan_rec or scan_b_id:
            col_qr_info, col_qr_btn = st.columns([4, 1])
            col_qr_info.success(f"🎯 **Accès direct QR Code actif** (Réception / ID : `{scan_rec or scan_b_id}`)")
            if col_qr_btn.button("✖ Revenir au normal", use_container_width=True):
                st.session_state.pop("pending_qr_rec", None)
                st.session_state.pop("pending_qr_bid", None)
                st.session_state.pop("qr_nav_applied", None)
                st.rerun()

        today_date = date.today()
        today_str = str(today_date)

        date_filtre = st.date_input("📅 Choisir une date à consulter", value=today_date, key="filtre_date_planning")
        date_filtre_str = str(date_filtre)
        debut_semaine = date_filtre - timedelta(days=date_filtre.weekday())
        fin_semaine = debut_semaine + timedelta(days=6)

        try:
            res_retards = supabase.table("suivi_controle_beton").select("*").eq("projet_id", projet_id_actif).lte("date_ecrasement", today_str).or_("force_kn.is.null,force_kn.eq.0").order("date_ecrasement", desc=False).execute()
            retards_list = res_retards.data or []
        except Exception as e:
            retards_list = []
            st.warning(f"Note retards : {e}")

        if retards_list:
            st.error(f"🚨 **ATTENTION : {len(retards_list)} éprouvette(s) non écrasée(s) ont atteint ou dépassé leur date d'échéance !**")
            rows_retard = []
            for ep in retards_list:
                row = _format_ep_row(ep, date_ref=today_date)
                dt_e = datetime.strptime(str(ep.get("date_ecrasement"))[:10], "%Y-%m-%d").date() if ep.get("date_ecrasement") else today_date
                priorite = f"⚠️ En Retard ({(today_date - dt_e).days} jour(s))" if dt_e < today_date else "🔥 Prévu Aujourd'hui"
                rows_retard.append({
                    "Priorité": priorite, "Date Écrasement Prévue": row["Date Écrasement Prévue"],
                    "Référence / Repère": row["Référence / Repère"], "N° BL": row["N° BL"], "Ouvrage": row["Ouvrage"],
                    "Classe Béton": row["Classe Béton"], "Date Coulée": row["Date Coulée"], "Échéance Visée": row["Échéance Visée"], "Âge Actuel Réel": row["Âge Théorique"]
                })
            st.dataframe(pd.DataFrame(rows_retard), use_container_width=True, hide_index=True)
            st.markdown("---")

        try:
            eprouvettes_date_sel = supabase.table("suivi_controle_beton").select("*").eq("projet_id", projet_id_actif).eq("date_ecrasement", date_filtre_str).order("id", desc=False).execute().data or []
            eprouvettes_semaine = supabase.table("suivi_controle_beton").select("*").eq("projet_id", projet_id_actif).gte("date_ecrasement", str(debut_semaine)).lte("date_ecrasement", str(fin_semaine)).order("date_ecrasement", desc=False).execute().data or []
        except Exception as e:
            eprouvettes_date_sel, eprouvettes_semaine = [], []
            st.warning(f"Note chargement planning : {e}")

        df_sel = pd.DataFrame([_format_ep_row(ep) for ep in eprouvettes_date_sel])
        df_semaine = pd.DataFrame([_format_ep_row(ep) for ep in eprouvettes_semaine])

        with st.expander(f"📆 Éprouvettes programmées spécifiquement pour le : {date_filtre_str} ({len(eprouvettes_date_sel)} éprouvette(s))", expanded=True):
            if not df_sel.empty: st.dataframe(df_sel, use_container_width=True, hide_index=True)
            else: st.info(f"ℹ️ Aucune éprouvette programmée pour le {date_filtre_str}.")

            st.markdown("---")
            col_exp1, col_exp2 = st.columns(2)
            with col_exp1:
                if not df_sel.empty:
                    st.download_button(f"📊 Télécharger la liste du jour ({date_filtre_str})", exporter_dataframe_excel(df_sel, date_filtre_str), file_name=f"Planning_Ecrasement_{date_filtre_str}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, key="btn_download_planning_excel")
            with col_exp2:
                if not df_semaine.empty:
                    st.download_button(f"📅 Télécharger la liste de la semaine ({debut_semaine} au {fin_semaine})", exporter_dataframe_excel(df_semaine, f"Sem_{debut_semaine}"), file_name=f"Planning_Semaine_{debut_semaine}_au_{fin_semaine}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, key="btn_download_planning_semaine_excel")

        st.markdown("---")

        try:
            if mode_admin:
                res_att = supabase.table("suivi_controle_beton").select("*").eq("projet_id", projet_id_actif).order("id", desc=False).execute()
                eprouvettes_en_attente = res_att.data or []
            else:
                res_att = (
                    supabase.table("suivi_controle_beton")
                    .select("*")
                    .eq("projet_id", projet_id_actif)
                    .or_("force_kn.is.null,force_kn.eq.0")
                    .order("id", desc=False)
                    .execute()
                )
                eprouvettes_en_attente = res_att.data or []
        except Exception as e:
            eprouvettes_en_attente = []
            st.error(f"Erreur de chargement des essais : {e}")

        if not eprouvettes_en_attente:
            st.info("👍 Aucune éprouvette en attente de saisie.")
        else:
            groupes_lots = {}
            index_selectionne = 0

            parents_dict = obtenir_infos_betonnage_parents_bulk(
                supabase, [ep.get("betonnage_id") for ep in eprouvettes_en_attente]
            )

            for idx_key, ep in enumerate(eprouvettes_en_attente):
                dt_ecras_str = str(ep.get("date_ecrasement") or "")[:10]
                
                if not mode_admin and dt_ecras_str:
                    try:
                        dt_ecras_obj = datetime.strptime(dt_ecras_str, "%Y-%m-%d").date()
                        if dt_ecras_obj > today_date:
                            is_scanned = (scan_rec and str(scan_rec).strip().lower() in str(ep.get("ref_controle") or "").lower()) or \
                                         (scan_b_id and str(scan_b_id).strip() == str(ep.get("betonnage_id")).strip())
                            if not is_scanned:
                                continue
                    except Exception:
                        pass

                b_id_ep = ep.get("betonnage_id")
                info_b_temp = parents_dict.get(b_id_ep, {})
                ref_ctrl = determiner_ref_controle(supabase, b_id_ep, info_b_temp, ep)
                num_rec_parent = str((info_b_temp or {}).get("num_reception") or "").strip()
                classe_ep = ep.get("classe_beton") or (info_b_temp.get("classe_beton") if info_b_temp else "-")
                t_essai = ep.get("type_essai", "Compression")
                
                prefixe_retard = ""
                if dt_ecras_str:
                    try:
                        dt_e = datetime.strptime(dt_ecras_str, "%Y-%m-%d").date()
                        if dt_e < today_date:
                            prefixe_retard = f"🚨 [RETARD {(today_date - dt_e).days}j] "
                        elif dt_e == today_date:
                            prefixe_retard = "⚠️ [À ÉCRASER AUJOURD'HUI] "
                    except Exception:
                        pass

                cle_groupe = f"{prefixe_retard}Réf: {ref_ctrl} | Essai: {t_essai} | Classe: {classe_ep} | Ouvrage: {ep.get('ouvrage', '-')} | Échéance: {ep.get('echeance', '28 jours')} | Lot #{b_id_ep}"

                if cle_groupe not in groupes_lots:
                    if scan_rec and (str(scan_rec).strip().lower() in ref_ctrl.lower() or str(scan_rec).strip().lower() in num_rec_parent.lower()):
                        index_selectionne = len(groupes_lots)
                    elif scan_b_id and str(scan_b_id).strip() == str(b_id_ep).strip():
                        index_selectionne = len(groupes_lots)
                    groupes_lots[cle_groupe] = []

                groupes_lots[cle_groupe].append(ep)

            options_lots = list(groupes_lots.keys())

            if not options_lots:
                st.info("ℹ️ Aucune éprouvette à écraser pour la date d'aujourd'hui ou en retard.")
            else:
                index_defaut = min(index_selectionne, len(options_lots) - 1) if options_lots else 0

                choix_lot = st.selectbox("📦 Sélectionner le lot d'éprouvettes à écraser / modifier :", options_lots, index=index_defaut, key="select_lot_saisie")
                lot_selected = groupes_lots[choix_lot]
                sample = lot_selected[0]
                betonnage_id = sample.get("betonnage_id")

                info_betonnage = parents_dict.get(betonnage_id, {}) or obtenir_infos_betonnage_parent(supabase, betonnage_id)
                historique_complet = obtenir_historique_betonnage(supabase, betonnage_id)
                exact_bl_phase1 = extraire_num_bl(sample, info_betonnage or {}, choix_lot)
                num_reception_affiche = sample.get("num_reception") or sample.get("n_reception") or ((info_betonnage or {}).get("num_reception") or "-")

                col_l1, col_l2, col_l3, col_l4 = st.columns(4)
                col_l1.metric("N° Réception", str(num_reception_affiche))
                col_l2.metric("N° Bon Livraison", exact_bl_phase1)
                col_l3.metric("Type d'essai", str(sample.get("type_essai", "Compression")))
                col_l4.metric("Échéance Visée", str(sample.get("echeance", "-")))

                st.markdown("---")
                st.markdown("##### 📝 Saisie des résultats de rupture (Force kN)")

                with st.form("form_saisie_ecrasement"):
                    technicien_saisie = st.text_input("Technicien / Opérateur d'essai", value=st.session_state.get("username", "Technicien LPEE"))
                    
                    saisies = []
                    for idx_ep, ep_item in enumerate(lot_selected):
                        c_ep1, c_ep2, c_ep3, c_ep4 = st.columns([2, 2, 2, 2])
                        c_ep1.write(f"**Éprouvette {ep_item.get('repere_eprouvette', '-')}**")
                        
                        f_curr = float(ep_item.get("force_kn") or 0.0)
                        f_kn_in = c_ep2.number_input(f"Force (kN) #{idx_ep+1}", min_value=0.0, max_value=3000.0, value=f_curr, step=0.1, key=f"f_kn_{ep_item['id']}")
                        
                        sec_ep = float(ep_item.get("section") or 17671.46)
                        t_essai_ep = ep_item.get("type_essai", "Compression")
                        
                        # Calcul dynamique de la résistance
                        rc_calc = calculer_resistance(f_kn_in, t_essai_ep, section=sec_ep)
                        
                        c_ep3.write(f"Résistance : **{rc_calc} MPa**")
                        c_ep4.write(f"Type : *{t_essai_ep}*")
                        
                        saisies.append({
                            "id": ep_item["id"],
                            "force_kn": f_kn_in,
                            "fc_mpa": rc_calc,
                            "technicien": technicien_saisie,
                            "date_essai": today_str,
                            "orig_f": f_curr
                        })

                    submit_saisie = st.form_submit_button("💾 Enregistrer les Écrasements du Lot", type="primary", use_container_width=True)

                    if submit_saisie:
                        nb_saisis = 0
                        for s_item in saisies:
                            if s_item["force_kn"] > 0:
                                pay_ecras = {
                                    "force_kn": s_item["force_kn"],
                                    "fc_mpa": s_item["fc_mpa"],
                                    "technicien": s_item["technicien"],
                                    "date_essai": s_item["date_essai"]
                                }
                                try:
                                    supabase.table("suivi_controle_beton").update(pay_ecras).eq("id", s_item["id"]).execute()
                                    enregistrer_modification(
                                        supabase,
                                        table_concernee="suivi_controle_beton",
                                        enregistrement_id=s_item["id"],
                                        action="ECRASEMENT",
                                        anciennes_valeurs={"force_kn": s_item["orig_f"]},
                                        nouvelles_valeurs=pay_ecras,
                                        commentaire="Saisie d'écrasement de béton",
                                    )
                                    nb_saisis += 1
                                except Exception as err_s:
                                    st.error(f"Erreur d'enregistrement pour ID #{s_item['id']} : {err_s}")

                        if nb_saisis > 0:
                            st.success(f"✅ {nb_saisis} écrasement(s) enregistré(s) avec succès !")
                            st.balloons()
                            st.rerun()

                # EXPORTATION DU PV EXCEL
                st.markdown("---")
                st.markdown("##### 📄 Exportation du Procès-Verbal (PV Excel)")
                
                # Chargement de toutes les éprouvettes du lot parent pour PV
                res_pv_all = supabase.table("suivi_controle_beton").select("*").eq("betonnage_id", betonnage_id).order("id", desc=False).execute()
                donnees_lot_pv = res_pv_all.data or []

                if donnees_lot_pv:
                    pv_excel_data = generer_pv_excel(donnees_lot_pv, info_betonnage)
                    st.download_button(
                        label="📥 Télécharger le PV d'Écrasement LPEE (.xlsx)",
                        data=pv_excel_data,
                        file_name=f"PV_Ecrasement_Beton_{num_reception_affiche}_{date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"btn_dl_pv_{betonnage_id}"
                    )

    # =========================================================
    # PHASE 3 : VALIDATION ADMIN (PVs)
    # =========================================================
    elif onglet_courant == OPTIONS_ONGLETS[3]:
        afficher_module_validation_admin(supabase, est_admin=is_baallal_admin or mode_admin)


if __name__ == "__main__":
    if Client:
        url = st.secrets.get("SUPABASE_URL")
        key = st.secrets.get("SUPABASE_KEY")
        if url and key:
            supabase_client = create_client(url, key)
            if st.session_state.get("user_logged"):
                show(supabase_client)
            else:
                afficher_ecran_connexion(supabase_client)
        else:
            st.error("Secrets SUPABASE_URL ou SUPABASE_KEY manquants.")
    else:
        st.error("Le module `supabase` n'est pas installé.")
