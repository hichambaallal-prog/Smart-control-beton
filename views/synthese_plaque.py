import streamlit as st
import pandas as pd
from datetime import datetime, date
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_excel_a4(df_filtered, filter_title="Synthèse Générale"):
    """
    Génère un fichier Excel professionnel mis en page pour impression A4 Portrait
    avec une police de taille 12, un espacement de ligne de 34, et les blocs de signature.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Synthèse Essais Plaque"

    # --- CONFIGURATION IMPRESSION A4 PORTRAIT ---
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # --- EN-TÊTE ET PIED DE PAGE D'IMPRESSION ---
    ws.oddHeader.left.text = "&\"Calibri,Bold\"&10LABORATOIRE LPEE - CTR-CSB\nProjet: LGV CASA SUD | Client: TGCC"
    ws.oddHeader.center.text = f"&\"Calibri,Bold\"&12SYNTHÈSE DES ESSAIS À LA PLAQUE\n{filter_title}"
    ws.oddHeader.right.text = "&\"Calibri,Regular\"&9Edité le: &D"
    ws.oddFooter.center.text = "&\"Calibri,Bold\"&10Page &P sur &N"

    # --- PALETTE DE COULEURS ET STYLES ---
    NAVY_HEADER = "1F4E79"
    BLUE_SUBHEADER = "2F5597"
    ICE_BLUE_BG = "F2F5F9"
    BORDER_COLOR = "D9D9D9"
    GREEN_OK = "E2EFDA"
    TEXT_GREEN = "276A3C"
    ORANGE_WARN = "FFF2CC"
    TEXT_ORANGE = "B25900"

    font_title = Font(name="Calibri", size=15, bold=True, color=NAVY_HEADER)
    font_th = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    
    fill_th = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
    fill_zebra = PatternFill(start_color=ICE_BLUE_BG, end_color=ICE_BLUE_BG, fill_type="solid")
    fill_kpi = PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )

    thick_top_bottom = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='medium', color=NAVY_HEADER),
        bottom=Side(style='double', color=NAVY_HEADER)
    )

    # --- 1. EN-TÊTE DU DOCUMENT ---
    ws.merge_cells("A1:G1")
    ws["A1"] = "LABORATOIRE LPEE — CENTRE TECHNIQUE RÉGIONAL"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = "Norme : NF P 94-117-1 (Plaque Ø 600 mm)"
    ws["A2"].font = Font(name="Calibri", size=15, bold=True, color=NAVY_HEADER)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:G3")
    ws["A3"] = "Projet : LGV CASA SUD  |  Client : TGCC"
    ws["A3"].font = Font(name="Calibri", size=14, bold=True, color=BLUE_SUBHEADER)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A4:G4")
    ws["A4"] = f"SYNTHÈSE DES ESSAIS DE PORTANCE À LA PLAQUE — {filter_title.upper()}"
    ws["A4"].font = Font(name="Calibri", size=12, italic=True, color="595959")
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 8

    # --- 2. EN-TÊTES DE TABLEAU ---
    headers = [
        "Date Essai", "Couche", "Emplacement", 
        "PK / Profil", "EV1 (MPa)", "EV2 (MPa)", "K (EV2/EV1)"
    ]

    ws.row_dimensions[6].height = 30
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=text)
        cell.font = font_th
        cell.fill = fill_th
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # --- 3. REMPLISSAGE DES DONNÉES ---
    start_row = 7
    for r_idx, (_, row) in enumerate(df_filtered.iterrows(), start=start_row):
        ws.row_dimensions[r_idx].height = 34 

        is_even = (r_idx % 2 == 0)
        current_fill = fill_zebra if is_even else PatternFill(fill_type=None)

        k_val = float(row.get("k", 0.0) or 0.0)

        values = [
            str(row.get("date_essai", "") or ""),
            str(row.get("couche", "") or ""),
            str(row.get("emplacement", "") or ""),
            str(row.get("pk_profil", "") or ""),
            float(row.get("ev1", 0.0) or 0.0),
            float(row.get("ev2", 0.0) or 0.0),
            k_val
        ]

        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Calibri", size=12) 
            cell.border = thin_border
            cell.fill = current_fill

            if c_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx in [2, 3, 4]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_idx in [5, 6]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0.00"
            elif c_idx == 7:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "0.00"
                if k_val >= 1.5:
                    cell.fill = PatternFill(start_color=GREEN_OK, end_color=GREEN_OK, fill_type="solid")
                    cell.font = Font(name="Calibri", size=12, bold=True, color=TEXT_GREEN)
                else:
                    cell.fill = PatternFill(start_color=ORANGE_WARN, end_color=ORANGE_WARN, fill_type="solid")
                    cell.font = Font(name="Calibri", size=12, bold=True, color=TEXT_ORANGE)

    end_row = start_row + len(df_filtered) - 1

    # --- 4. LIGNE DE MOYENNE AUTOMATIQUE ---
    if len(df_filtered) > 0:
        stat_row = end_row + 1
        ws.row_dimensions[stat_row].height = 26

        ws.merge_cells(start_row=stat_row, start_column=1, end_row=stat_row, end_column=4)
        lbl_cell = ws.cell(row=stat_row, column=1, value="MOYENNE DES ESSAIS")
        lbl_cell.font = Font(name="Calibri", size=11, bold=True, color=NAVY_HEADER)
        lbl_cell.alignment = Alignment(horizontal="right", vertical="center")

        for col_idx in range(1, 5):
            ws.cell(row=stat_row, column=col_idx).border = thick_top_bottom
            ws.cell(row=stat_row, column=col_idx).fill = fill_kpi

        formulas = [
            (5, f"=AVERAGE(E{start_row}:E{end_row})", "#,##0.00"),
            (6, f"=AVERAGE(F{start_row}:F{end_row})", "#,##0.00"),
            (7, f"=AVERAGE(G{start_row}:G{end_row})", "0.00")
        ]

        for c_idx, form, num_fmt in formulas:
            c = ws.cell(row=stat_row, column=c_idx, value=form)
            c.font = Font(name="Calibri", size=11, bold=True, color=NAVY_HEADER)
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = thick_top_bottom
            c.fill = fill_kpi
            c.number_format = num_fmt

        # --- 5. SYNTHÈSE QUALITÉ ---
        synth_start = stat_row + 2
        ws.cell(row=synth_start, column=1, value="RÉSUMÉ STATISTIQUE QUALITÉ").font = Font(name="Calibri", size=12, bold=True, color=NAVY_HEADER)

        summary_headers = ["Indicateur", "EV1 (MPa)", "EV2 (MPa)", "Ratio K (EV2/EV1)"]
        ws.row_dimensions[synth_start+1].height = 24

        for idx, header in enumerate(summary_headers, start=1):
            c = ws.cell(row=synth_start+1, column=idx, value=header)
            c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color=BLUE_SUBHEADER, end_color=BLUE_SUBHEADER, fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border

        metrics = [
            ("Valeur Minimale", f"=MIN(E{start_row}:E{end_row})", f"=MIN(F{start_row}:F{end_row})", f"=MIN(G{start_row}:G{end_row})"),
            ("Valeur Maximale", f"=MAX(E{start_row}:E{end_row})", f"=MAX(F{start_row}:F{end_row})", f"=MAX(G{start_row}:G{end_row})"),
            ("Moyenne Générale", f"=AVERAGE(E{start_row}:E{end_row})", f"=AVERAGE(F{start_row}:F{end_row})", f"=AVERAGE(G{start_row}:G{end_row})"),
            ("Nombre d'essais", f"=COUNT(E{start_row}:E{end_row})", f"=COUNT(F{start_row}:F{end_row})", f"=COUNT(G{start_row}:G{end_row})")
        ]

        for idx, (label, ev1_f, ev2_f, k_f) in enumerate(metrics, start=synth_start+2):
            ws.row_dimensions[idx].height = 22
            c1 = ws.cell(row=idx, column=1, value=label)
            c2 = ws.cell(row=idx, column=2, value=ev1_f)
            c3 = ws.cell(row=idx, column=3, value=ev2_f)
            c4 = ws.cell(row=idx, column=4, value=k_f)
            
            c1.font = Font(name="Calibri", size=11, bold=True)
            c1.border = thin_border
            c1.alignment = Alignment(horizontal="left", vertical="center")
            
            for c, fmt in zip([c2, c3, c4], ["#,##0.00", "#,##0.00", "0.00" if "Nombre" not in label else "0"]):
                c.font = Font(name="Calibri", size=11)
                c.border = thin_border
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = fmt

        # --- 6. BLOCS DE SIGNATURES ---
        sig_start = synth_start + 7
        ws.row_dimensions[sig_start].height = 24

        ws.merge_cells(start_row=sig_start, start_column=2, end_row=sig_start, end_column=3)
        c_resp = ws.cell(row=sig_start, column=2, value="Responsable d'essai")
        c_resp.font = Font(name="Calibri", size=11, bold=True, color=NAVY_HEADER)
        c_resp.alignment = Alignment(horizontal="center", vertical="center")
        c_resp.border = thin_border

        ws.merge_cells(start_row=sig_start, start_column=5, end_row=sig_start, end_column=6)
        c_chef = ws.cell(row=sig_start, column=5, value="Chef du Laboratoire")
        c_chef.font = Font(name="Calibri", size=11, bold=True, color=NAVY_HEADER)
        c_chef.alignment = Alignment(horizontal="center", vertical="center")
        c_chef.border = thin_border

        for r in range(sig_start + 1, sig_start + 4):
            ws.row_dimensions[r].height = 24
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
            
            for col in range(2, 4):
                ws.cell(row=r, column=col).border = thin_border
            for col in range(5, 7):
                ws.cell(row=r, column=col).border = thin_border

    col_widths = {
        'A': 14, 'B': 18, 'C': 20, 'D': 15,
        'E': 14, 'F': 14, 'G': 14
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def show(supabase):
    col_header1, col_header2 = st.columns([2, 1])
    with col_header1:
        st.title("📊 Synthèse Essais à la Plaque")
    with col_header2:
        st.markdown(
            "<div style='text-align: right; padding-top: 15px; font-weight: bold; color: #1F4E79; font-size: 1.1em;'>"
            "📋 Norme : NF P 94-117-1"
            "</div>", 
            unsafe_allow_html=True
        )

    st.markdown("---")

    if not supabase:
        st.error("❌ Connexion Supabase indisponible.")
        return

    try:
        res = supabase.table("essais_plaque").select("*").order("date_essai", desc=True).execute()
        data = res.data if res else []
    except Exception as e:
        st.error(f"Erreur lors de la connexion à la base de données : {e}")
        return

    if not data:
        st.info("Aucun essai enregistré dans la base de données pour le moment.")
        return

    df = pd.DataFrame(data)

    # --- SÉCURISATION DES COLONNES PAR DÉFAUT ---
    expected_columns = ["date_essai", "couche", "emplacement", "pk_profil", "ev1", "ev2", "k"]
    for col in expected_columns:
        if col not in df.columns:
            df[col] = None

    df['date_essai_dt'] = pd.to_datetime(df['date_essai'], errors='coerce')

    # --- FILTRES DE RECHERCHE ---
    st.markdown("### 🔍 Filtres de Recherche")
    col_f1, col_f2, col_f3, col_f4 = st.columns(4)

    with col_f1:
        type_recap = st.selectbox("Période", ["Tous les essais", "Journalier", "Mensuel", "Période Personnalisée"])

    filtered_df = df.copy()
    filter_label = "Général"

    if type_recap == "Journalier":
        with col_f2:
            date_choisie = st.date_input("Date", value=date.today())
            filtered_df = df[df['date_essai_dt'].dt.date == date_choisie]
            filter_label = f"Journalier du {date_choisie.strftime('%d/%m/%Y')}"

    elif type_recap == "Mensuel":
        with col_f2:
            mois_choisi = st.date_input("Choisir le mois", value=date.today())
            filtered_df = df[
                (df['date_essai_dt'].dt.year == mois_choisi.year) & 
                (df['date_essai_dt'].dt.month == mois_choisi.month)
            ]
            filter_label = f"Mensuel - {mois_choisi.strftime('%m/%Y')}"

    elif type_recap == "Période Personnalisée":
        with col_f2:
            d_start = st.date_input("Du", value=date.today())
        with col_f3:
            d_end = st.date_input("Au", value=date.today())
            filtered_df = df[
                (df['date_essai_dt'].dt.date >= d_start) & 
                (df['date_essai_dt'].dt.date <= d_end)
            ]
            filter_label = f"Période du {d_start.strftime('%d/%m/%Y')} au {d_end.strftime('%d/%m/%Y')}"
    else:
        filter_label = "Historique Complet"

    # Filtres secondaires
    with col_f3 if type_recap != "Période Personnalisée" else col_f4:
        emplacements = ["Tous les emplacements"] + sorted([str(x) for x in df['emplacement'].dropna().unique() if str(x) not in ["None", "nan"]])
        emp_sel = st.selectbox("Emplacement", emplacements)
        if emp_sel != "Tous les emplacements":
            filtered_df = filtered_df[filtered_df['emplacement'] == emp_sel]
            filter_label += f" | Emplacement : {emp_sel}"

    with col_f4 if type_recap != "Période Personnalisée" else col_f1:
        couches = ["Toutes les couches"] + sorted([str(x) for x in df['couche'].dropna().unique() if str(x) not in ["None", "nan"]])
        couche_sel = st.selectbox("Type de couche", couches)
        if couche_sel != "Toutes les couches":
            filtered_df = filtered_df[filtered_df['couche'] == couche_sel]
            filter_label += f" | Couche : {couche_sel}"

    st.markdown("---")

    # --- KPI STATISTIQUES ---
    st.markdown(f"### 📈 Métriques Clés : **{filter_label}**")

    if filtered_df.empty:
        st.warning("⚠️ Aucun essai trouvé pour les filtres sélectionnés.")
    else:
        nb_essais = len(filtered_df)
        ev1_moyen = pd.to_numeric(filtered_df['ev1'], errors='coerce').fillna(0).mean()
        ev2_moyen = pd.to_numeric(filtered_df['ev2'], errors='coerce').fillna(0).mean()
        k_moyen = pd.to_numeric(filtered_df['k'], errors='coerce').fillna(0).mean()
        taux_conforme = (pd.to_numeric(filtered_df['k'], errors='coerce').fillna(0) >= 1.50).sum() / nb_essais * 100

        kpi1, kpi2, kpi3, kpi4, kpi5 = st.columns(5)
        kpi1.metric("Total Essais", f"{nb_essais}")
        kpi2.metric("EV1 Moyen", f"{ev1_moyen:.2f} MPa")
        kpi3.metric("EV2 Moyen", f"{ev2_moyen:.2f} MPa")
        kpi4.metric("K Moyen", f"{k_moyen:.2f}")
        kpi5.metric("Conformité (K ≥ 1.5)", f"{taux_conforme:.1f}%")

        # --- GRAPHIQUE ---
        st.markdown("#### 📊 Évolution des Modules EV1 et EV2")
        df_chart = filtered_df.copy()
        df_chart['pk_profil'] = df_chart['pk_profil'].fillna('N/A')
        st.line_chart(df_chart.set_index("pk_profil")[["ev1", "ev2"]])

        # --- TABLEAU INTERACTIF ---
        st.markdown("#### 📋 Détail des Essais Filtrés")
        
        df_display = filtered_df.copy()
        df_display['date_essai'] = df_display['date_essai_dt'].dt.strftime('%Y-%m-%d')

        cols_show = ["date_essai", "couche", "emplacement", "pk_profil", "ev1", "ev2", "k"]

        df_display_clean = df_display[cols_show].rename(columns={
            "date_essai": "Date", "couche": "Couche", "emplacement": "Emplacement",
            "pk_profil": "PK/Profil", "ev1": "EV1 (MPa)", "ev2": "EV2 (MPa)", "k": "Coeff K"
        })

        st.dataframe(df_display_clean, use_container_width=True, hide_index=True)

        # ==========================================================
        # --- ZONE ADMINISTRATION (Admin Only) ---
        # ==========================================================
        if st.session_state.get("role") == "admin":
            st.markdown("---")
            st.subheader("🛠️ Espace Administration (Gestion des données)")
            
            record_options = {f"ID {r['id']} - {r.get('date_essai', 'N/A')} - {r.get('pk_profil', '')}": r for r in data}
            selected_key = st.selectbox("Sélectionner l'essai à gérer", list(record_options.keys()))
            selected_item = record_options[selected_key]
            
            col_ed, col_del = st.columns(2)
            
            with col_ed:
                with st.expander("📝 Modifier cet essai"):
                    with st.form("edit_form"):
                        new_pk = st.text_input("PK / Profil", value=selected_item.get("pk_profil", ""))
                        new_ev1 = st.number_input("EV1 (MPa)", value=float(selected_item.get("ev1", 0) or 0))
                        new_ev2 = st.number_input("EV2 (MPa)", value=float(selected_item.get("ev2", 0) or 0))
                        
                        if st.form_submit_button("Enregistrer les modifications"):
                            try:
                                new_k = new_ev2 / new_ev1 if new_ev1 > 0 else 0
                                supabase.table("essais_plaque").update({
                                    "pk_profil": new_pk,
                                    "ev1": new_ev1,
                                    "ev2": new_ev2,
                                    "k": new_k
                                }).eq("id", selected_item["id"]).execute()
                                st.success("Données mises à jour avec succès !")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Erreur lors de la mise à jour : {e}")
                                
            with col_del:
                st.markdown("##### ⚠️ Suppression")
                if st.button("🗑️ Supprimer définitivement", type="primary"):
                    try:
                        supabase.table("essais_plaque").delete().eq("id", selected_item["id"]).execute()
                        st.success("Enregistrement supprimé avec succès.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erreur lors de la suppression : {e}")

        # --- TÉLÉCHARGEMENT EXCEL ---
        st.markdown("---")
        st.markdown("### 📥 Exportation Excel avec Mise en Page A4 Imprimable")
        
        excel_data = generate_excel_a4(filtered_df, filter_title=filter_label)
        file_name_clean = f"Synthese_Essais_Plaque_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        st.download_button(
            label="📄 Télécharger la Synthèse Excel (Format A4 Imprimable)",
            data=excel_data,
            file_name=file_name_clean,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
